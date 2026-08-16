"""Kintsugi vs Sales Tax Agent reconciliation.

Parses Kintsugi transaction detail XLSX (Jurisdiction Summary sheet)
and compares against our sales_by_state data for a configurable window.

Output: console report + CSV + optional Storage upload.
"""
from __future__ import annotations

import csv
import io
import json
from datetime import date, datetime, timezone
from pathlib import Path

from src.channels import normalize_channel, SHOPIFY, AMAZON
from src.db import fetch_all


# ── US state codes (for filtering territories/aggregates) ──

US_STATES = {
    "AL", "AK", "AZ", "AR", "CA", "CO", "CT", "DC", "DE", "FL",
    "GA", "HI", "ID", "IL", "IN", "IA", "KS", "KY", "LA", "ME",
    "MD", "MA", "MI", "MN", "MS", "MO", "MT", "NE", "NV", "NH",
    "NJ", "NM", "NY", "NC", "ND", "OH", "OK", "OR", "PA", "RI",
    "SC", "SD", "TN", "TX", "UT", "VT", "VA", "WA", "WV", "WI",
    "WY",
}


# ── Parse Kintsugi XLSX ────────────────────────────────────

def parse_jurisdiction_summary(filepath: str | Path) -> list[dict]:
    """Parse the Jurisdiction Summary sheet from a Kintsugi XLSX.

    Returns list of dicts: {state_code, state_name, txn_count, txn_amount,
                            tax_collected, taxable_sales}
    """
    import openpyxl
    wb = openpyxl.load_workbook(str(filepath), read_only=True, data_only=True)
    ws = wb["Jurisdiction Summary"]

    # Row 0 = org info, Row 1 = headers, Row 2+ = data
    rows = []
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i < 2:
            continue  # skip org info + header
        country = str(row[0] or "").strip()
        state = str(row[1] or "").strip().upper()
        if country != "US" or not state or state not in US_STATES:
            continue

        rows.append({
            "state_code": state,
            "state_name": str(row[2] or "").strip().title(),
            "txn_count": _to_int(row[3]),
            "txn_amount": _to_float(row[4]),
            "tax_collected": _to_float(row[7]),
            "taxable_sales": _to_float(row[10]),
        })

    wb.close()
    return rows


def parse_detail_channel_split(
    filepath: str | Path,
    states: list[str] | None = None,
) -> dict[str, dict]:
    """Parse detail sheets US-XX to get channel split for selected states.

    Returns {state_code: {amazon_amount, shopify_amount, amazon_count, shopify_count}}
    Only parses requested states to avoid reading the full 32MB file.
    """
    import openpyxl
    wb = openpyxl.load_workbook(str(filepath), read_only=True, data_only=True)

    if states is None:
        states = ["CA", "TX"]

    result = {}
    for sc in states:
        sheet_name = f"US-{sc}"
        if sheet_name not in wb.sheetnames:
            continue

        ws = wb[sheet_name]
        amazon_amt = 0.0
        shopify_amt = 0.0
        amazon_cnt = 0
        shopify_cnt = 0

        for i, row in enumerate(ws.iter_rows(values_only=True)):
            if i < 2:
                continue  # skip org info + header
            if not row or len(row) < 19:
                continue

            source = str(row[11] or "").strip().upper()
            amount = _to_float(row[18])

            if "AMAZON" in source:
                amazon_amt += amount
                amazon_cnt += 1
            elif "SHOPIFY" in source:
                shopify_amt += amount
                shopify_cnt += 1

        result[sc] = {
            "amazon_amount": round(amazon_amt, 2),
            "shopify_amount": round(shopify_amt, 2),
            "amazon_count": amazon_cnt,
            "shopify_count": shopify_cnt,
        }

    wb.close()
    return result


# ── Agent data query ───────────────────────────────────────

def query_agent_sales(
    window_start: str = "2024-01-01",
    window_end: str | None = None,
) -> dict[str, dict]:
    """Query sales_by_state and aggregate by state + channel.

    Returns {state_code: {shopify, amazon, total, shopify_txn, amazon_txn, total_txn}}
    """
    if window_end is None:
        window_end = date.today().isoformat()

    records = fetch_all("sales_by_state")
    by_state: dict[str, dict] = {}

    for rec in records:
        ps = str(rec.get("period_start", ""))
        if ps < window_start or ps > window_end:
            continue

        sc = rec.get("state_code")
        if not sc or sc not in US_STATES:
            continue

        ch = normalize_channel(rec.get("channel", "unknown"))
        amt = float(rec.get("gross_sales", 0) or 0)
        txns = int(rec.get("order_count", 0) or 0)

        if sc not in by_state:
            by_state[sc] = {
                "shopify": 0.0, "amazon": 0.0, "total": 0.0,
                "shopify_txn": 0, "amazon_txn": 0, "total_txn": 0,
            }

        if ch == SHOPIFY:
            by_state[sc]["shopify"] += amt
            by_state[sc]["shopify_txn"] += txns
        else:
            by_state[sc]["amazon"] += amt
            by_state[sc]["amazon_txn"] += txns
        by_state[sc]["total"] += amt
        by_state[sc]["total_txn"] += txns

    return by_state


# ── Comparison ─────────────────────────────────────────────

def build_comparison(
    kintsugi_file: str | Path,
    window_start: str = "2024-01-01",
    window_end: str | None = None,
    channel_split_states: list[str] | None = None,
) -> dict:
    """Build the full comparison object.

    Returns {rows, flags, channel_splits, totals, methodology, generated_at}
    """
    kintsugi = parse_jurisdiction_summary(kintsugi_file)
    agent = query_agent_sales(window_start, window_end)

    if channel_split_states is None:
        channel_split_states = ["CA", "TX"]

    channel_splits = parse_detail_channel_split(kintsugi_file, channel_split_states)

    # Union all states
    all_states = sorted(set(k["state_code"] for k in kintsugi) | set(agent.keys()))

    kintsugi_by_state = {k["state_code"]: k for k in kintsugi}

    rows = []
    flags = []
    for sc in all_states:
        k = kintsugi_by_state.get(sc, {})
        a = agent.get(sc, {})

        k_txn_amt = k.get("txn_amount", 0.0)
        k_taxable = k.get("taxable_sales", 0.0)
        k_txn_count = k.get("txn_count", 0)

        a_shopify = round(a.get("shopify", 0.0), 2)
        a_amazon = round(a.get("amazon", 0.0), 2)
        a_total = round(a.get("total", 0.0), 2)

        delta = round(a_total - k_txn_amt, 2)
        delta_pct = round(delta / k_txn_amt * 100, 1) if k_txn_amt else (
            100.0 if a_total > 0 else 0.0
        )

        notes = []
        if not k:
            notes.append("Not in Kintsugi")
        if not a:
            notes.append("Not in Agent")
        if k_taxable and k_txn_amt and abs(k_taxable - k_txn_amt) > 100:
            notes.append(f"K taxable ${k_taxable:,.0f} != txn ${k_txn_amt:,.0f}")

        row = {
            "state_code": sc,
            "state_name": k.get("state_name", sc),
            "k_txn_amount": k_txn_amt,
            "k_taxable_sales": k_taxable,
            "k_txn_count": k_txn_count,
            "k_tax_collected": k.get("tax_collected", 0.0),
            "agent_shopify": a_shopify,
            "agent_amazon": a_amazon,
            "agent_total": a_total,
            "agent_txn_count": a.get("total_txn", 0),
            "delta": delta,
            "delta_pct": delta_pct,
            "notes": "; ".join(notes),
        }
        rows.append(row)

        if abs(delta_pct) > 15 and abs(delta) > 5000:
            flags.append({
                "state_code": sc,
                "delta": delta,
                "delta_pct": delta_pct,
                "reason": "; ".join(notes) if notes else "Large variance",
            })

    # Channel split validation for CA, TX
    split_notes = {}
    for sc, cs in channel_splits.items():
        a = agent.get(sc, {})
        split_notes[sc] = {
            "kintsugi_amazon": cs["amazon_amount"],
            "kintsugi_shopify": cs["shopify_amount"],
            "agent_amazon": round(a.get("amazon", 0.0), 2),
            "agent_shopify": round(a.get("shopify", 0.0), 2),
            "amazon_delta": round(a.get("amazon", 0.0) - cs["amazon_amount"], 2),
            "shopify_delta": round(a.get("shopify", 0.0) - cs["shopify_amount"], 2),
        }

    k_grand = sum(r["k_txn_amount"] for r in rows)
    a_grand = sum(r["agent_total"] for r in rows)

    return {
        "generated_at": datetime.now(timezone.utc).isoformat(),
        "kintsugi_file": str(kintsugi_file),
        "window": {"start": window_start, "end": window_end or date.today().isoformat()},
        "rows": rows,
        "flags": flags,
        "channel_splits": split_notes,
        "totals": {
            "kintsugi_grand_total": round(k_grand, 2),
            "agent_grand_total": round(a_grand, 2),
            "grand_delta": round(a_grand - k_grand, 2),
            "grand_delta_pct": round((a_grand - k_grand) / k_grand * 100, 1) if k_grand else 0,
            "states_in_kintsugi": sum(1 for r in rows if r["k_txn_amount"] > 0),
            "states_in_agent": sum(1 for r in rows if r["agent_total"] > 0),
        },
        "methodology": [
            "Kintsugi: 'Transaction Amount' from Jurisdiction Summary sheet used as primary comparison basis.",
            "Agent: gross_sales from sales_by_state (Shopify + Amazon), deduplicated by source priority.",
            "Kintsugi 'Taxable Sales' may differ from Transaction Amount due to exemptions and deductions.",
            "Kintsugi 'Final Liability' is the tax owed, NOT the sales base — do not confuse with threshold amount.",
            "Amazon marketplace tax collected appears in Kintsugi but does NOT eliminate seller economic nexus obligations.",
            "Measurement windows may differ: Kintsugi aggregates all-time; Agent window is configurable.",
            "Source dedup: Agent prefers SP-API over CSV for same (state, channel, period).",
            f"Comparison window: {window_start} to {window_end or 'today'}.",
        ],
    }


def build_csv(comparison: dict) -> str:
    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow([
        "state_code", "state_name",
        "kintsugi_txn_amount", "kintsugi_taxable_sales", "kintsugi_txn_count",
        "kintsugi_tax_collected",
        "agent_shopify", "agent_amazon", "agent_total", "agent_txn_count",
        "delta", "delta_pct", "notes",
    ])
    for r in comparison["rows"]:
        writer.writerow([
            r["state_code"], r["state_name"],
            r["k_txn_amount"], r["k_taxable_sales"], r["k_txn_count"],
            r["k_tax_collected"],
            r["agent_shopify"], r["agent_amazon"], r["agent_total"],
            r["agent_txn_count"],
            r["delta"], r["delta_pct"], r["notes"],
        ])
    return output.getvalue()


def print_report(comparison: dict) -> str:
    """Format a console-friendly report. Returns the text."""
    lines = []
    t = comparison["totals"]
    lines.append("=" * 70)
    lines.append("KINTSUGI vs SALES TAX AGENT RECONCILIATION")
    lines.append("=" * 70)
    lines.append(f"Window: {comparison['window']['start']} to {comparison['window']['end']}")
    lines.append(f"File:   {comparison['kintsugi_file']}")
    lines.append("")
    lines.append(f"{'':>6s}  {'Kintsugi':>14s}  {'Agent':>14s}  {'Delta':>12s}  {'%':>7s}")
    lines.append(f"{'TOTAL':>6s}  ${t['kintsugi_grand_total']:>13,.0f}  ${t['agent_grand_total']:>13,.0f}  ${t['grand_delta']:>11,.0f}  {t['grand_delta_pct']:>6.1f}%")
    lines.append(f"States: {t['states_in_kintsugi']} in Kintsugi, {t['states_in_agent']} in Agent")
    lines.append("")

    # Top variances
    flagged = comparison["flags"]
    if flagged:
        lines.append(f"FLAGGED ({len(flagged)} states with |delta| > $5k AND |%| > 15%):")
        for f in sorted(flagged, key=lambda x: abs(x["delta"]), reverse=True):
            lines.append(f"  {f['state_code']}: ${f['delta']:>+12,.0f} ({f['delta_pct']:>+6.1f}%)  {f['reason']}")
        lines.append("")

    # Full table
    lines.append(f"{'State':<6s} {'K Txn $':>12s} {'K Taxable':>12s} {'Agent Shop':>12s} {'Agent Amz':>12s} {'Agent Tot':>12s} {'Delta':>10s} {'%':>7s}")
    lines.append("-" * 85)
    for r in sorted(comparison["rows"], key=lambda x: abs(x["delta"]), reverse=True):
        flag = " ***" if any(f["state_code"] == r["state_code"] for f in flagged) else ""
        lines.append(
            f"{r['state_code']:<6s} "
            f"${r['k_txn_amount']:>11,.0f} "
            f"${r['k_taxable_sales']:>11,.0f} "
            f"${r['agent_shopify']:>11,.0f} "
            f"${r['agent_amazon']:>11,.0f} "
            f"${r['agent_total']:>11,.0f} "
            f"${r['delta']:>+9,.0f} "
            f"{r['delta_pct']:>+6.1f}%{flag}"
        )

    # Channel splits
    if comparison["channel_splits"]:
        lines.append("")
        lines.append("CHANNEL SPLIT VALIDATION (from Kintsugi detail sheets):")
        for sc, cs in comparison["channel_splits"].items():
            lines.append(f"  {sc}:")
            lines.append(f"    Amazon:  Kintsugi ${cs['kintsugi_amazon']:>12,.0f}  Agent ${cs['agent_amazon']:>12,.0f}  Delta ${cs['amazon_delta']:>+10,.0f}")
            lines.append(f"    Shopify: Kintsugi ${cs['kintsugi_shopify']:>12,.0f}  Agent ${cs['agent_shopify']:>12,.0f}  Delta ${cs['shopify_delta']:>+10,.0f}")

    lines.append("")
    lines.append("METHODOLOGY:")
    for m in comparison["methodology"]:
        lines.append(f"  {m}")

    return "\n".join(lines)


# ── Helpers ────────────────────────────────────────────────

def _to_float(v) -> float:
    if v is None:
        return 0.0
    try:
        return float(v)
    except (ValueError, TypeError):
        return 0.0


def _to_int(v) -> int:
    if v is None:
        return 0
    try:
        return int(float(v))
    except (ValueError, TypeError):
        return 0
