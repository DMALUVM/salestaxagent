from __future__ import annotations

import signal
import subprocess
import sys
import time

import click


def _code_revision() -> str:
    try:
        return subprocess.check_output(
            ["git", "rev-parse", "--short", "HEAD"],
            stderr=subprocess.DEVNULL,
            text=True,
        ).strip()
    except Exception:
        return "unknown"


@click.group()
def cli():
    """Sales Tax Compliance Agent — Monitor nexus, track deadlines, stay compliant."""
    pass


@cli.command()
def analyze():
    """Run full nexus analysis (physical + economic) and print summary."""
    click.echo("Running nexus analysis...")

    click.echo("\n--- Physical Nexus (FBA Inventory) ---")
    from src.engines.physical_nexus import evaluate_physical_nexus
    phys = evaluate_physical_nexus()

    if phys.get("nexus_states"):
        click.echo(f"Physical nexus detected in {len(phys['nexus_states'])} states: {', '.join(phys['nexus_states'])}")
        if phys.get("new_nexus_states"):
            click.echo(f"  NEW nexus states: {', '.join(phys['new_nexus_states'])}")
    else:
        click.echo(phys.get("message", "No physical nexus detected."))

    if phys.get("franchise_flags"):
        click.echo(f"\n  Franchise tax flags: {len(phys['franchise_flags'])}")
        for flag in phys["franchise_flags"]:
            click.echo(f"    {flag['severity'].upper()}: {flag['state_code']} — {flag['description'][:100]}")

    click.echo("\n--- Economic Nexus (Sales Thresholds) ---")
    from src.engines.economic_nexus import evaluate_economic_nexus
    econ = evaluate_economic_nexus()

    if econ.get("exceeded_threshold"):
        click.echo(f"Economic nexus EXCEEDED: {', '.join(econ['exceeded_threshold'])}")
    if econ.get("approaching_threshold"):
        click.echo(f"Approaching threshold: {', '.join(econ['approaching_threshold'])}")
    if not econ.get("exceeded_threshold") and not econ.get("approaching_threshold"):
        click.echo(econ.get("message", "No states approaching economic nexus thresholds."))

    # Show top states by total sales so the user sees their Amazon data
    details = econ.get("details", {})
    if details:
        top = sorted(details.items(), key=lambda kv: kv[1].get("total_amount", 0), reverse=True)[:10]
        click.echo(f"\n  Top 10 states by total sales (Shopify + Amazon):")
        click.echo(f"  {'State':<6} {'Counted$':>12} {'Shopify':>12} {'Amazon':>12} {'Threshold':>12} {'$ Prog':>8} {'MP?':>4}")
        for sc, d in top:
            mp = "Y" if d.get("marketplace_included", True) else "N"
            click.echo(
                f"  {sc:<6} ${d['threshold_amount']:>11,.2f} ${d['shopify_amount']:>11,.2f} "
                f"${d['amazon_amount']:>11,.2f} ${(d.get('threshold_amount_cfg') or 0):>11,.0f} "
                f"{d['progress_percent']:>6.1f}% {mp:>4}"
            )
        click.echo(f"\n  MP? = marketplace sales count toward threshold (Y/N per state rule).")
        click.echo(f"  Remittance liability is always Shopify/direct only.")

    click.echo("\n--- Full Report ---")
    from src.reports.cli_reports import full_report
    click.echo(full_report())


@cli.command()
def status():
    """Show current nexus status summary."""
    from src.reports.cli_reports import nexus_summary
    click.echo(nexus_summary())


@cli.command()
@click.option("--days", default=30, help="Days ahead to check")
def deadlines(days):
    """Show upcoming filing deadlines."""
    from src.reports.cli_reports import deadlines_report
    click.echo(deadlines_report(days))


@cli.command()
def flags():
    """Show franchise / entity tax flags."""
    from src.reports.cli_reports import franchise_flags_report
    click.echo(franchise_flags_report())


@cli.command()
@click.option("--state", required=True, help="State code (e.g., TX)")
@click.option("--period", required=True, help="Period label (e.g., 2026-Q1)")
@click.option("--amount", type=float, default=None, help="Amount filed")
@click.option("--notes", default=None, help="Filing notes")
@click.option("--zero-return", is_flag=True, help="Mark as zero return")
def complete(state, period, amount, notes, zero_return):
    """Mark a filing as complete."""
    from src.calendar.filing_calendar import mark_filing_complete
    result = mark_filing_complete(state.upper(), period, amount, notes, zero_return)
    if result:
        click.echo(f"Filing marked complete: {state.upper()} {period}")
    else:
        click.echo(f"Filing not found: {state.upper()} {period}. Check state code and period label.")


@cli.command()
@click.option("--format", "fmt", type=click.Choice(["csv", "json"]), default="json")
@click.option("--output", default=None, help="Output file path")
def export(fmt, output):
    """Export nexus data for CPA review."""
    from src.reports.cli_reports import export_csv, export_json

    if output is None:
        from datetime import date
        output = f"sales_tax_report_{date.today().isoformat()}.{fmt}"

    if fmt == "csv":
        result = export_csv(output)
    else:
        result = export_json(output)
    click.echo(result)


@cli.command("test-alert")
def test_alert():
    """Send a test Telegram notification."""
    from src.alerts.telegram import send_test_alert
    result = send_test_alert()
    if result.get("sent"):
        click.echo("Test alert sent successfully! Check your Telegram.")
    else:
        click.echo(f"Failed to send alert: {result.get('error')}")


@cli.command("backfill-shopify-skus")
def backfill_shopify_skus():
    """Pull Shopify line items and populate sales_by_sku (monthly grain)."""
    click.echo("Fetching Shopify orders with line items...")
    from src.parsers.shopify_skus import fetch_shopify_skus
    result = fetch_shopify_skus()
    if result.get("error"):
        click.echo(f"Error: {result['error']}")
    else:
        click.echo(f"Orders: {result['orders_fetched']:,}")
        click.echo(f"SKU rows: {result['sku_rows']:,}")
        click.echo(f"Unique SKUs: {result['unique_skus']}")
        click.echo(f"Inserted: {result['rows_inserted']}")


@cli.command("backfill-amazon-skus")
@click.option("--start", "start_str", default=None,
              help="Start date (YYYY-MM-DD, default: 2025-01-01)")
@click.option("--end", "end_str", default=None,
              help="End date (YYYY-MM-DD, default: yesterday)")
@click.option("--dry-run", is_flag=True)
def backfill_amazon_skus(start_str, end_str, dry_run):
    """Pull Amazon SP-API orders and populate sales_by_sku (monthly grain)."""
    from datetime import date as d, timedelta
    start = d.fromisoformat(start_str) if start_str else d(2025, 1, 1)
    end = d.fromisoformat(end_str) if end_str else d.today() - timedelta(days=1)
    yesterday = d.today() - timedelta(days=1)
    if end > yesterday:
        end = yesterday

    if dry_run:
        click.echo("DRY RUN — no data will be written.\n")

    from src.rules import clamp_orders_report_range, orders_report_floor
    floor = orders_report_floor()
    clamped_start, clamped_end, floor_warning = clamp_orders_report_range(start, end)
    if floor_warning:
        click.echo(floor_warning)
    if clamped_start is None:
        click.echo(
            f"Skipped Amazon. All Orders floor is {floor.isoformat()}. "
            "No SKU rows will be written."
        )
        return
    start, end = clamped_start, clamped_end

    click.echo(f"Fetching Amazon SKU data: {start} to {end}")

    def _on_poll(status, elapsed):
        click.echo(f"  [{elapsed}s] {status}")

    from src.amazon_sp.reports import fetch_amazon_skus
    result = fetch_amazon_skus(start, end, dry_run=dry_run, on_poll=_on_poll)

    click.echo(f"\n  Amazon SKU Results:")
    click.echo(f"  Chunks:      {result.get('chunks', 0)}")
    click.echo(f"  Rows total:  {result.get('rows_total', 0):,}")
    click.echo(f"  Rows parsed: {result.get('rows_parsed', 0):,}")
    click.echo(f"  SKU rows:    {result.get('sku_rows', 0):,}")
    click.echo(f"  Unique SKUs: {result.get('unique_skus', 0)}")
    click.echo(f"  Inserted:    {result.get('rows_inserted', 0):,}")
    for w in result.get("warnings", []):
        click.echo(f"  Warning: {w}")


@cli.command("ads-import")
@click.argument("path", type=click.Path(exists=True))
@click.option("--dry-run", is_flag=True)
def ads_import_cmd(path, dry_run):
    """Load monthly Amazon ad spend from SKU Economics or an Ads Console CSV.

    Ads API only keeps ~95 days. Drop the file in incoming/amazon/ or pass
    the path here. SKU Economics must be Monthly aggregation (or one month
    per file) — a multi-month custom range is one lump and is refused.
    """
    from src.parsers.amazon_ads_spend import ingest_amazon_ads_spend
    result = ingest_amazon_ads_spend(path, dry_run=dry_run)
    click.echo(f"Kind:    {result.get('kind')}")
    click.echo(f"Months:  {result.get('months', 0)}  {result.get('month_starts')}")
    click.echo(f"Spend:   ${result.get('total_spend', 0):,.2f}")
    click.echo(f"Inserted:{result.get('rows_inserted', 0)}")
    for w in result.get("warnings") or []:
        click.echo(f"Warning: {w}")


@cli.command("ads-monthly-restore")
@click.option("--seed", "seed_path", type=click.Path(exists=True), default=None,
              help="Seed CSV (default: config/ads_monthly_spend_seed.csv)")
@click.option("--dry-run", is_flag=True)
def ads_monthly_restore(seed_path, dry_run):
    """Restore ads_monthly_spend from the committed seed file."""
    from src.parsers.amazon_ads_spend import restore_ads_monthly_from_seed
    result = restore_ads_monthly_from_seed(seed_path, dry_run=dry_run)
    if result.get("error"):
        click.echo(result["error"], err=True)
        return
    click.echo(f"Seed: {result.get('seed')}")
    click.echo(f"Rows: {result.get('rows', 0)}  Inserted: {result.get('inserted', 0)}")


@cli.command("ads-monthly-export-seed")
@click.option("--output", "seed_path", type=click.Path(), default=None,
              help="Output path (default: config/ads_monthly_spend_seed.csv)")
def ads_monthly_export_seed(seed_path):
    """Export ads_monthly_spend to the git-tracked seed CSV."""
    from src.parsers.amazon_ads_spend import export_ads_monthly_seed
    result = export_ads_monthly_seed(seed_path)
    click.echo(f"Wrote {result.get('rows', 0)} month(s) to {result.get('path')}")


@cli.command("warehouse-export")
@click.option("--output", "-o", type=click.Path(), default=None,
              help="Output .json.gz path (default: warehouse_snapshot_<date>.json.gz)")
@click.option("--table", "tables", multiple=True,
              help="Export only these tables (default: all configured)")
def warehouse_export_cmd(output, tables):
    """Export a gzip JSON bundle of all operational Supabase tables."""
    from datetime import date as d
    from src.maintenance.warehouse_snapshot import export_snapshot_gzip

    if output is None:
        output = f"warehouse_snapshot_{d.today().isoformat()}.json.gz"
    table_list = list(tables) if tables else None
    result = export_snapshot_gzip(output, tables_filter=table_list)
    click.echo(f"Wrote {result.get('bytes', 0):,} bytes to {result.get('path')}")
    click.echo(f"Tables OK: {result.get('tables_ok')}  Errors: {result.get('tables_error')}")
    click.echo(f"Total rows: {result.get('total_rows', 0):,}")


@cli.command("warehouse-restore")
@click.argument("path", type=click.Path(exists=True))
@click.option("--dry-run", is_flag=True, help="Validate and count rows without writing")
@click.option("--table", "tables", multiple=True,
              help="Restore only these tables from the bundle")
def warehouse_restore_cmd(path, dry_run, tables):
    """Restore operational tables from a warehouse snapshot .json.gz bundle."""
    from src.maintenance.warehouse_snapshot import restore_snapshot_file

    table_list = list(tables) if tables else None
    result = restore_snapshot_file(path, dry_run=dry_run, tables_filter=table_list)
    click.echo(f"Snapshot from: {result.get('exported_at')}")
    click.echo(f"Tables processed: {result.get('tables_processed')}")
    click.echo(f"Rows upserted: {result.get('total_upserted', 0):,}")
    if result.get("unknown_tables"):
        click.echo(f"Unknown tables in file (skipped): {', '.join(result['unknown_tables'])}")
    for err in result.get("errors") or []:
        click.echo(f"ERROR {err.get('table')}: {err.get('error')}", err=True)
    for row in result.get("tables") or []:
        if row.get("status") == "error":
            continue
        click.echo(
            f"  {row['table']}: {row.get('rows_in_backup', 0):,} in backup, "
            f"{row.get('upserted', 0):,} upserted"
        )


@cli.command("export-csv")
@click.option("--table", type=click.Choice([
    "sales_by_state", "sales_by_sku", "ads_monthly_spend",
]), default="sales_by_state", help="Table to export")
@click.option("--start", "start_str", default=None, help="Start date (YYYY-MM-DD)")
@click.option("--end", "end_str", default=None, help="End date (YYYY-MM-DD)")
@click.option("--output", default=None, help="Output file path")
def export_csv_cmd(table, start_str, end_str, output):
    """Export sales data as CSV for CPA review."""
    import csv as csvmod
    from datetime import date as d
    from src.db import fetch_all

    rows = fetch_all(table)
    date_key = "period_start"
    if start_str:
        rows = [r for r in rows if (r.get(date_key) or "") >= start_str]
    if end_str:
        rows = [r for r in rows if (r.get(date_key) or "") <= end_str]

    if not rows:
        click.echo("No rows match the filter.")
        return

    if output is None:
        output = f"{table}_{d.today().isoformat()}.csv"

    keys = sorted(rows[0].keys())
    with open(output, "w", newline="") as f:
        writer = csvmod.DictWriter(f, fieldnames=keys)
        writer.writeheader()
        writer.writerows(rows)

    click.echo(f"Exported {len(rows)} rows to {output}")


@cli.command("populate-calendar")
@click.option("--year", type=int, default=None, help="Calendar year (default: current + next)")
def populate_calendar(year):
    """Generate filing calendar entries for registered states."""
    from src.calendar.filing_calendar import populate_calendar_for_registered_states
    result = populate_calendar_for_registered_states(year)
    click.echo(f"States: {result.get('states_populated', [])}")
    click.echo(f"Entries: {result.get('entries_created', 0)}")
    if result.get("message"):
        click.echo(result["message"])


@cli.command("generate-filings")
def generate_filings_cmd():
    """Generate filing periods for all registered states (current + next year)."""
    from src.calendar.filing_calendar import populate_calendar_for_registered_states
    result = populate_calendar_for_registered_states()
    states = result.get("states_populated", [])
    entries = result.get("entries_created", 0)
    if not states:
        click.echo("No registered states found. Register states on the Registrations page first.")
    else:
        click.echo(f"Generated {entries} filing periods for {len(states)} states: {', '.join(states)}")
        click.echo(f"Years: {result.get('years', [])}")


@cli.command("health-report")
@click.option("--stale-days", default=90, help="Days before a rule is considered stale")
def health_report(stale_days):
    """Generate a Rules Health Report for the intelligence layer."""
    from src.intelligence.health_report import generate_health_report
    click.echo(generate_health_report(stale_threshold_days=stale_days))


@cli.command("query-state")
@click.argument("state_code")
def query_state(state_code):
    """Query the intelligence layer for a state's full nexus profile."""
    from src.intelligence.knowledge_base import query_state_nexus
    result = query_state_nexus(state_code)

    click.echo(f"\n═══ {result['state_name']} ({result['state_code']}) ═══\n")
    click.echo(f"Sales tax: {'Yes' if result['has_sales_tax'] else 'No'}")
    click.echo(f"Contested positions: {'Yes' if result['has_contested_positions'] else 'No'}")
    click.echo(f"Franchise risk: {'Yes' if result['has_franchise_risk'] else 'No'}")

    if result["nexus_rules"]:
        click.echo(f"\nNexus rules ({len(result['nexus_rules'])}):")
        for r in result["nexus_rules"]:
            confidence = r.get("confidence", "?")
            marker = "⚠️" if confidence == "contested" else "✓" if confidence == "high" else "○"
            click.echo(f"  {marker} [{r.get('rule_type', '?')}] {r.get('position_summary', '')[:120]}")
            click.echo(f"    Confidence: {confidence} | Last reviewed: {r.get('last_reviewed', '?')}")
            for src in (r.get("primary_sources") or [])[:3]:
                if isinstance(src, dict):
                    click.echo(f"    Source: {src.get('citation', 'N/A')}")

    if result["franchise_entity_rules"]:
        click.echo(f"\nFranchise/entity rules ({len(result['franchise_entity_rules'])}):")
        for r in result["franchise_entity_rules"]:
            click.echo(f"  🏛️ [{r.get('rule_type', '?')}] {r.get('position_summary', '')[:120]}")
            if r.get("minimum_amount"):
                click.echo(f"    Minimum: ${r['minimum_amount']:,.0f}")
            if r.get("due_date_pattern"):
                click.echo(f"    Due: {r['due_date_pattern'][:80]}")

    if result["court_rulings"]:
        click.echo(f"\nCourt rulings ({len(result['court_rulings'])}):")
        for r in result["court_rulings"]:
            click.echo(f"  📜 {r.get('case_name', '?')} ({r.get('status', '?')})")
            click.echo(f"    {r.get('holding_summary', '')[:120]}")

    if result["admin_rulings"]:
        click.echo(f"\nAdmin rulings ({len(result['admin_rulings'])}):")
        for r in result["admin_rulings"]:
            click.echo(f"  📋 {r.get('title', '?')} ({r.get('status', '?')})")
            click.echo(f"    {r.get('summary', '')[:120]}")

    click.echo(f"\n{result['disclaimer']}")


@cli.command("search-rulings")
@click.argument("query")
@click.option("--type", "ruling_type", type=click.Choice(["all", "court", "admin"]), default="all")
def search_rulings_cmd(query, ruling_type):
    """Search court and administrative rulings by keyword."""
    from src.intelligence.knowledge_base import search_rulings
    results = search_rulings(query, ruling_type)

    if not results:
        click.echo(f"No rulings found matching '{query}'.")
        return

    click.echo(f"\nFound {len(results)} ruling(s) matching '{query}':\n")
    for r in results:
        rtype = r.get("_ruling_type", "?")
        name = r.get("case_name") or r.get("title", "?")
        status = r.get("status", "?")
        click.echo(f"  [{rtype.upper()}] {name} — {status}")
        summary = r.get("holding_summary") or r.get("summary", "")
        click.echo(f"    {summary[:150]}")
        if r.get("relevance_to_fba"):
            click.echo(f"    FBA relevance: {r['relevance_to_fba'][:120]}")
        click.echo()


@cli.command("contested")
def contested_positions():
    """Show all contested positions that need attention."""
    from src.intelligence.knowledge_base import get_all_contested_positions
    positions = get_all_contested_positions()

    if not positions:
        click.echo("No contested positions in the knowledge base.")
        return

    click.echo(f"\n═══ CONTESTED POSITIONS ({len(positions)}) ═══\n")
    for p in positions:
        click.echo(f"  ⚠️  {p['state_code']} — {p['rule_type']}")
        click.echo(f"    {p.get('summary', '')[:120]}")
        if p.get("conservative"):
            click.echo(f"    Conservative: {p['conservative'][:100]}")
        if p.get("aggressive"):
            click.echo(f"    Aggressive: {p['aggressive'][:100]}")
        if p.get("open_questions"):
            click.echo(f"    Open questions: {p['open_questions'][:120]}")
        click.echo()


@cli.command("monitor-sources")
@click.option("--frequency", type=click.Choice(["weekly", "biweekly", "monthly", "quarterly"]),
              default=None, help="Filter by check frequency")
def monitor_sources(frequency):
    """Run change detection on monitored source URLs."""
    from src.intelligence.monitor import run_monitoring_cycle
    click.echo("Running source monitoring cycle...")
    result = run_monitoring_cycle(frequency_filter=frequency)
    click.echo(f"\nSources checked: {result['sources_checked']}")
    click.echo(f"Changes detected: {result['changes_detected']}")
    click.echo(f"Errors: {result['errors']}")

    if result["changes_detected"] > 0:
        click.echo("\nChanges detected in:")
        for d in result["details"]:
            if d.get("change"):
                click.echo(f"  ⚠️  {d.get('state', '?')} — {d.get('title', '?')}")


@cli.command("research-tasks")
@click.option("--status", type=click.Choice(["open", "in_progress", "completed", "all"]), default="open")
def research_tasks_cmd(status):
    """List research tasks from the intelligence layer."""
    from src.db import fetch_all
    tasks = fetch_all("research_tasks")
    if status != "all":
        tasks = [t for t in tasks if t.get("status") == status]

    if not tasks:
        click.echo(f"No {status} research tasks.")
        return

    click.echo(f"\n═══ RESEARCH TASKS ({status.upper()}) ═══\n")
    for t in sorted(tasks, key=lambda x: {"critical": 0, "high": 1, "medium": 2, "low": 3}.get(x.get("priority", "medium"), 2)):
        click.echo(f"  [{t.get('priority', '?').upper()}] {t.get('title', '?')}")
        click.echo(f"    State: {t.get('state_code', '—')} | Type: {t.get('task_type', '?')}")
        click.echo(f"    {t.get('description', '')[:120]}")
        click.echo()


@cli.command("spapi-test")
def spapi_test():
    """Test the SP-API connection (verifies credentials)."""
    from src.config import settings
    if not settings.amazon_sp_enabled:
        click.echo("SP-API not configured.  Set AMAZON_SP_CLIENT_ID, "
                    "AMAZON_SP_CLIENT_SECRET, and AMAZON_SP_REFRESH_TOKEN in .env")
        return
    try:
        from src.amazon_sp.auth import get_access_token
        token = get_access_token()
        click.echo(f"SP-API auth OK.  Access token: {token[:20]}...")
        click.echo(f"Marketplace: {settings.amazon_sp_marketplace_id}")
    except Exception as e:
        click.echo(f"SP-API auth failed: {e}")


@cli.command("spapi-dump-headers")
@click.option("--report", "report_type", required=True,
              help="SP-API report type constant (e.g. GET_FLAT_FILE_ALL_ORDERS_DATA_BY_ORDER_DATE_GENERAL)")
@click.option("--start", "start_str", required=True, help="Start date (YYYY-MM-DD)")
@click.option("--end", "end_str", default=None, help="End date (YYYY-MM-DD)")
def spapi_dump_headers(report_type, start_str, end_str):
    """Download an SP-API report and print the raw header + first 3 data rows."""
    from datetime import date as d
    start = d.fromisoformat(start_str)
    end = d.fromisoformat(end_str) if end_str else d.today()

    click.echo(f"Requesting {report_type}: {start} to {end}")

    def _on_poll(status, elapsed):
        click.echo(f"  [{elapsed}s] {status}")

    from src.amazon_sp.client import request_and_download
    content = request_and_download(report_type, start, end, on_poll=_on_poll)

    lines = content.split("\n")
    click.echo(f"\nTotal lines: {len(lines)}")
    click.echo(f"First line length: {len(lines[0])} chars")
    click.echo(f"\n--- repr of header line ---")
    click.echo(repr(lines[0][:500]))
    click.echo(f"\n--- first 3 data rows (repr) ---")
    for i, line in enumerate(lines[1:4], 1):
        click.echo(f"Row {i}: {repr(line[:500])}")

    # Also detect delimiter and show parsed fieldnames
    from src.amazon_sp.reports import _detect_delimiter, _build_header_lookup
    import csv, io
    delim = _detect_delimiter(lines[0])
    click.echo(f"\n--- delimiter detected: {repr(delim)} ---")
    reader = csv.DictReader(io.StringIO(content), delimiter=delim, quotechar='"')
    if reader.fieldnames:
        click.echo(f"Parsed {len(reader.fieldnames)} headers:")
        for j, fn in enumerate(reader.fieldnames):
            click.echo(f"  [{j:2d}] {fn!r}")
    else:
        click.echo("NO HEADERS DETECTED")


@cli.command("spapi-orders")
@click.option("--start", "start_str", required=True,
              help="Start date (YYYY-MM-DD)")
@click.option("--end", "end_str", default=None,
              help="End date (YYYY-MM-DD, default: today)")
@click.option("--dry-run", is_flag=True,
              help="Parse and display without writing to database")
def spapi_orders(start_str, end_str, dry_run):
    """Fetch Amazon orders via SP-API and ingest into sales_by_state."""
    from datetime import date as d, timedelta
    start = d.fromisoformat(start_str)
    end = d.fromisoformat(end_str) if end_str else d.today()
    # Clamp end to yesterday — SP-API returns empty reports for future dates
    yesterday = d.today() - timedelta(days=1)
    if end > yesterday:
        end = yesterday

    if dry_run:
        click.echo("DRY RUN — no data will be written.\n")

    click.echo(f"Requesting SP-API orders report: {start} to {end}")

    def _on_poll(status, elapsed):
        click.echo(f"  [{elapsed}s] Report status: {status}")

    from src.amazon_sp.reports import fetch_orders
    result = fetch_orders(start, end, dry_run=dry_run, on_poll=_on_poll)
    _print_spapi_result("SP-API Orders", result)


@cli.command("spapi-inventory")
@click.option("--start", "start_str", required=True,
              help="Start date (YYYY-MM-DD)")
@click.option("--end", "end_str", default=None,
              help="End date (YYYY-MM-DD, default: today)")
@click.option("--dry-run", is_flag=True,
              help="Parse and display without writing to database")
def spapi_inventory(start_str, end_str, dry_run):
    """Fetch Inventory Ledger via SP-API and ingest into inventory_events."""
    from datetime import date as d, timedelta
    start = d.fromisoformat(start_str)
    end = d.fromisoformat(end_str) if end_str else d.today()
    yesterday = d.today() - timedelta(days=1)
    if end > yesterday:
        end = yesterday

    if dry_run:
        click.echo("DRY RUN — no data will be written.\n")

    click.echo(f"Requesting SP-API inventory ledger: {start} to {end}")

    def _on_poll(status, elapsed):
        click.echo(f"  [{elapsed}s] Report status: {status}")

    from src.amazon_sp.reports import fetch_inventory
    result = fetch_inventory(start, end, dry_run=dry_run, on_poll=_on_poll)
    _print_spapi_result("SP-API Inventory Ledger", result)


@cli.command("spapi-ledger-summary")
@click.option("--start", "start_str", default=None,
              help="Start date (YYYY-MM-DD). Default: latest complete Amazon day.")
@click.option("--end", "end_str", default=None,
              help="End date (YYYY-MM-DD, default: latest complete Amazon day)")
@click.option("--dry-run", is_flag=True,
              help="Parse and display without writing to database")
def spapi_ledger_summary(start_str, end_str, dry_run):
    """Fetch GET_LEDGER_SUMMARY_VIEW_DATA (DAILY/FC) into tax inventory $."""
    from datetime import date
    from src.rules import amazon_as_of
    from src.amazon_sp.ledger_summary import fetch_ledger_summary

    end = date.fromisoformat(end_str) if end_str else amazon_as_of()
    start = date.fromisoformat(start_str) if start_str else end
    if end > amazon_as_of():
        end = amazon_as_of()
    if start > end:
        start = end

    if dry_run:
        click.echo("DRY RUN — no data will be written.\n")

    click.echo(f"Requesting ledger summary (DAILY/FC): {start} to {end}")

    def _on_poll(status, elapsed):
        click.echo(f"  [{elapsed}s] Report status: {status}")

    result = fetch_ledger_summary(start, end, dry_run=dry_run, on_poll=_on_poll)
    _print_spapi_result("SP-API Ledger Summary", result)


@cli.command("backfill-ledger-summary")
@click.option("--start", "start_str", default="2026-01-01",
              help="Start date (YYYY-MM-DD). Default: 2026-01-01")
@click.option("--end", "end_str", default=None,
              help="End date (YYYY-MM-DD, default: latest complete Amazon day)")
@click.option("--dry-run", is_flag=True,
              help="Parse and display without writing to database")
def backfill_ledger_summary(start_str, end_str, dry_run):
    """Backfill daily ledger-summary COGS, chunked ≤30d like other SP-API reports."""
    from datetime import date
    from src.rules import amazon_as_of
    from src.amazon_sp.ledger_summary import fetch_ledger_summary

    start = date.fromisoformat(start_str)
    end = date.fromisoformat(end_str) if end_str else amazon_as_of()
    if end > amazon_as_of():
        end = amazon_as_of()

    if dry_run:
        click.echo("DRY RUN — no data will be written.\n")

    click.echo(f"Backfilling ledger summary (DAILY/FC): {start} to {end}")

    def _on_poll(status, elapsed):
        click.echo(f"  [{elapsed}s] Report status: {status}")

    result = fetch_ledger_summary(start, end, dry_run=dry_run, on_poll=_on_poll)
    _print_spapi_result("SP-API Ledger Summary backfill", result)


@cli.command("sync-daily")
@click.option("--days", default=7, help="Days back to sync")
def sync_daily_cmd(days):
    """Sync sales_daily from SP-API + Shopify (Amazon LA tz, pending included)."""
    from src.sales_daily import sync_amazon_daily, sync_shopify_daily

    click.echo(f"Syncing Amazon daily sales ({days}d, Pacific timezone, incl. pending)...")
    try:
        result = sync_amazon_daily(days=days)
        click.echo(f"  Amazon: {result.get('rows_upserted', 0)} rows, "
                   f"${result.get('total_gross', 0):,.2f}")
    except Exception as e:
        click.echo(f"  Amazon error: {e}")

    click.echo(f"Syncing Shopify daily sales ({days}d, Eastern timezone)...")
    try:
        result = sync_shopify_daily(days=days)
        click.echo(f"  Shopify: {result.get('rows_upserted', 0)} rows")
    except Exception as e:
        click.echo(f"  Shopify error: {e}")


@cli.command("ads-test")
def ads_test_cmd():
    """Test Amazon Ads API connection and list profiles."""
    from src.config import settings
    if not settings.amazon_ads_enabled:
        click.echo("Amazon Ads not configured. Set AMAZON_ADS_* in .env")
        return
    try:
        from src.amazon_ads.client import get_profiles
        profiles = get_profiles()
        click.echo(f"Connected! {len(profiles)} profile(s):")
        for p in profiles:
            click.echo(f"  ID={p.get('profileId')}  Name={p.get('accountInfo',{}).get('name','')}  "
                       f"Type={p.get('accountInfo',{}).get('type','')}  Marketplace={p.get('accountInfo',{}).get('marketplaceStringId','')}")
    except Exception as e:
        click.echo(f"Error: {e}")


@cli.command("ads-sync")
@click.option("--days", default=14, help="Days of history to sync")
@click.option("--campaigns-only", is_flag=True,
              help="Sync campaigns only — returns without waiting on search terms")
@click.option("--search-terms-only", is_flag=True,
              help="Sync search terms only")
@click.option("--placements-only", is_flag=True,
              help="Sync placement performance only (needs migration_ads_placement.sql)")
@click.option("--with-placements", is_flag=True,
              help="Also sync placement performance on a full sync")
@click.option("--search-term-chunk-days", default=None, type=int,
              help="Chunk size for search-term reports (default 7)")
@click.option("--campaign-chunk-days", default=None, type=int,
              help="Chunk size for campaign reports (default 7, max 30). "
                   "SB/SD stay on 1-day chunks; poll wait is 1800s.")
@click.option("--ad-products", default=None,
              help="Comma-separated ad products for campaigns: SP,SB,SD "
                   "(default: all three). Search terms and placements are "
                   "always Sponsored Products.")
def ads_sync_cmd(days, campaigns_only, search_terms_only, placements_only,
                 with_placements, search_term_chunk_days, campaign_chunk_days,
                 ad_products):
    """Sync Amazon Ads campaigns + search terms (auto-chunked).

    Campaigns chunk at 7 days by default (≤30 max); search terms at 7.
    A single 30-day SB/SD report times out on this account.

    Campaigns cover Sponsored Products, Brands and Display; each row carries
    its campaign_type so the totals match the Ads console while the breakdown
    stays available.
    """
    from src.config import settings
    if not settings.amazon_ads_enabled:
        click.echo("Amazon Ads not configured. Set AMAZON_ADS_* in .env")
        return
    if sum(bool(f) for f in (campaigns_only, search_terms_only, placements_only)) > 1:
        raise click.UsageError("--campaigns-only, --search-terms-only and "
                               "--placements-only are mutually exclusive")
    from src.amazon_ads.reports import sync_ads, SEARCH_TERM_CHUNK_DAYS, AD_PRODUCTS
    from src.db import job_start, job_finish

    products = None
    if ad_products:
        products = tuple(p.strip().upper() for p in ad_products.split(",") if p.strip())
        unknown = [p for p in products if p not in AD_PRODUCTS]
        if unknown:
            raise click.UsageError(
                f"unknown ad product(s) {', '.join(unknown)} — choose from "
                f"{', '.join(AD_PRODUCTS)}")

    job_name = ("ads_campaigns_sync" if campaigns_only else
                "ads_search_terms_sync" if search_terms_only else
                "ads_placements_sync" if placements_only else "ads_sync")
    st_chunk = search_term_chunk_days or SEARCH_TERM_CHUNK_DAYS

    run_id = job_start(job_name)
    scope = ("campaigns only" if campaigns_only else
             "search terms only" if search_terms_only else
             "placements only" if placements_only else
             "campaigns + search terms" + (" + placements" if with_placements else ""))
    from src.rules import ADS_CAMPAIGN_CHUNK_DAYS
    camp_chunk = campaign_chunk_days or ADS_CAMPAIGN_CHUNK_DAYS
    click.echo(f"Syncing last {days} days of Ads data ({scope}; "
               f"campaigns {camp_chunk}d chunks, search terms {st_chunk}d chunks)...")
    result = sync_ads(days=days, campaigns_only=campaigns_only,
                      search_terms_only=search_terms_only,
                      placements_only=placements_only,
                      with_placements=with_placements,
                      search_term_chunk_days=search_term_chunk_days,
                      ad_products=products,
                      campaign_chunk_days=campaign_chunk_days,
                      on_progress=click.echo)
    errors = []

    camp = result.get("campaigns")
    if isinstance(camp, dict) and camp.get("by_type"):
        for t in ("SP", "SB", "SD"):
            v = camp["by_type"].get(t)
            if not v:
                continue
            click.echo(f"  {t}: {v['rows']} rows, ${v['spend']:,.2f} spend, "
                       f"{v['clicks']:,} clicks"
                       + ("" if v["ok"] else f"  ⚠ {len(v['errors'])} error(s)"))

    for key in ["campaigns", "search_terms", "placements"]:
        val = result.get(key)
        if val is None:
            continue  # this half was skipped — don't report it as "0 rows"
        if isinstance(val, dict):
            if val.get("errors"):
                click.echo(f"  {key}: {val.get('rows', 0)} rows, {val.get('inserted', 0)} inserted "
                           f"({val.get('chunks', 0)} chunks, {len(val['errors'])} errors)")
                for err in val["errors"][:3]:
                    click.echo(f"    ⚠ {err}")
                errors.extend(val["errors"])
            elif "error" in val:
                click.echo(f"  {key}: ERROR — {val['error'][:80]}")
                errors.append(val["error"])
            else:
                click.echo(f"  {key}: {val.get('rows', 0)} rows, {val.get('inserted', 0)} inserted "
                           f"({val.get('chunks', 0)} chunks)")
                if val.get("total_spend"):
                    click.echo(f"    Spend: ${val['total_spend']:,.2f}  "
                               f"Dates: {val.get('date_min', '?')} → {val.get('date_max', '?')}")

    click.echo(f"  Range: {result.get('start', '?')} → {result.get('end', '?')}")

    status, message = _ads_sync_outcome(result, days)
    job_finish(run_id, status, message)
    click.echo(f"  Job {job_name}: {status} — {message}")


def _print_search_term_coverage() -> None:
    """min/max/count from ads_search_terms_daily — stored dates only."""
    try:
        from src.db import get_client
        client = get_client()
        lo = (client.table("ads_search_terms_daily").select("date")
              .order("date", desc=False).limit(1).execute())
        hi = (client.table("ads_search_terms_daily").select("date")
              .order("date", desc=True).limit(1).execute())
        n = client.table("ads_search_terms_daily").select("date", count="exact").limit(1).execute()
        min_d = (lo.data or [{}])[0].get("date")
        max_d = (hi.data or [{}])[0].get("date")
        rows = getattr(n, "count", None)
        click.echo(f"  ads_search_terms_daily: min={min_d} max={max_d} rows={rows} "
                   f"(SP-only SUMMARY dates; not padded)")
    except Exception as e:
        click.echo(f"  ads_search_terms_daily coverage lookup failed: {e}")


@cli.command("ads-search-terms-backfill")
@click.option("--days", default=90, show_default=True,
              help="Closed Amazon days to cover (default 90).")
@click.option("--chunk-days", default=7, show_default=True,
              help="Ads reporting chunk size. 7 is what this account completes.")
@click.option("--cancel-report-id", default=None,
              help="Cancel this PENDING report id, then STOP (no fetch).")
def ads_search_terms_backfill_cmd(days, chunk_days, cancel_report_id):
    """One-shot search-terms-only 90d backfill. Never writes to Amazon Ads.

    7-day chunks, upsert each week, skip weeks already stored, newest first.
    If the reporting slot is busy (HTTP 425) or another ads pull holds the
    lock: STOP. Do not retry in a loop. Do not run two of these at once.

    Weekday nightly ingest stays 7d. Sunday 03:30 is the scheduled twin of
    this command. Search-term reports are SP-only.
    """
    from src.config import settings
    if not settings.amazon_ads_enabled:
        click.echo("Amazon Ads not configured. Set AMAZON_ADS_* in .env")
        return
    from src.amazon_ads.client import cancel_report
    from src.amazon_ads.reports import AdsSyncBusy, sync_ads
    from src.db import job_start, job_finish

    if cancel_report_id:
        click.echo(f"Cancelling report {cancel_report_id} (queue cleanup only — "
                   f"not an ads write)...")
        ok = cancel_report(cancel_report_id)
        click.echo("  cancelled" if ok else "  cancel failed or id unknown")
        click.echo("STOP. Do not start a fetch from this command when cancelling.")
        return

    click.echo(f"One-shot search-terms-only backfill: last {days}d in "
               f"{chunk_days}d chunks, skip existing weeks, newest first. "
               f"SP-only. If 425: STOP.")
    _print_search_term_coverage()
    run_id = job_start("ads_search_terms_backfill")
    try:
        result = sync_ads(
            days=days,
            search_terms_only=True,
            search_term_chunk_days=chunk_days,
            skip_existing_search_term_weeks=True,
            newest_first_search_terms=True,
            on_progress=click.echo)
    except AdsSyncBusy as e:
        job_finish(run_id, "skipped", str(e)[:500])
        click.echo(f"STOP: {e}")
        click.echo("Do not start a second search-term CLI. Wait until the slot is free.")
        return
    except Exception as e:
        job_finish(run_id, "fail", str(e)[:500])
        click.echo(f"STOP: {e}")
        return

    st = result.get("search_terms") or {}
    if isinstance(st, dict) and st.get("stopped"):
        click.echo(f"STOP: search-term fetch halted ({st['stopped']}). "
                   f"Weeks already upserted were kept. Do not retry in a loop.")
        for err in (st.get("errors") or [])[:3]:
            click.echo(f"  ⚠ {err}")
    elif isinstance(st, dict):
        click.echo(f"  search_terms: {st.get('rows', 0)} rows, "
                   f"{st.get('inserted', 0)} inserted, "
                   f"{st.get('chunks_skipped_existing', 0)} weeks already stored, "
                   f"{len(st.get('errors') or [])} chunk errors")
        for err in (st.get("errors") or [])[:3]:
            click.echo(f"    ⚠ {err}")
    status, message = _ads_sync_outcome(result, days)
    job_finish(run_id, status, message,
               stats=st if isinstance(st, dict) else None)
    click.echo(f"  Job ads_search_terms_backfill: {status} — {message}")
    _print_search_term_coverage()
    click.echo("Proof SQL: select min(date), max(date), count(*), count(distinct date) "
               "from ads_search_terms_daily;")


def _ads_sync_outcome(result: dict, days: int) -> tuple[str, str]:
    """Classify an ads sync as success / partial / fail.

    Only the halves that actually ran are judged, so `--campaigns-only` is not
    marked partial for the search terms it deliberately skipped. Campaigns
    landing while search terms time out is a *partial*, not a failure — the
    /ppc KPIs and trends are current either way.
    """
    def half(name: str) -> str | None:  # noqa: D401
        val = result.get(name)
        if not isinstance(val, dict):
            return None                       # did not run
        if "error" in val:
            return "fail"
        return "partial" if val.get("errors") else "ok"

    campaigns, search_terms = half("campaigns"), half("search_terms")
    placements = half("placements")

    # Ad-product detail: a whole missing product (SB or SD) is a different kind
    # of partial than a single timed-out chunk, and it is the one that silently
    # under-reports spend against the console. Name it in the message.
    camp = result.get("campaigns")
    product_note = ""
    if isinstance(camp, dict) and camp.get("by_type"):
        ok = camp.get("products_ok") or []
        bad = camp.get("products_failed") or []
        parts = "/".join(f"{t} ${camp['by_type'][t]['spend']:,.2f}"
                         for t in ("SP", "SB", "SD") if t in camp["by_type"])
        product_note = f" [{parts}]"
        if bad:
            product_note += f" — {'+'.join(bad)} FAILED, {'+'.join(ok) or 'nothing'} kept"
    ran = [s for s in (campaigns, search_terms, placements) if s is not None]
    if not ran:
        return "fail", "nothing ran"

    scope = "+".join(result.get("ran", []))
    detail = []
    if campaigns:
        detail.append(f"campaigns {campaigns}")
    if search_terms:
        detail.append(f"search_terms {search_terms}")
    if placements:
        detail.append(f"placements {placements}")
    summary = f"{days}d {scope}: " + ", ".join(detail) + product_note

    if all(s == "ok" for s in ran):
        return "success", summary
    if all(s == "fail" for s in ran):
        return "fail", summary
    return "partial", summary


@cli.command("ads-actions")
@click.option("--target-acos", default=None, type=float,
              help="Target ACOS %. Defaults to the break-even target computed "
                   "from COGS/fee inputs.")
@click.option("--days", default=7, help="Lookback window for search terms")
def ads_actions_cmd(target_acos, days):
    """Generate PPC action recommendations (replaces the open queue).

    With no --target-acos, the break-even target is derived from sku_costs and
    the P&L fee inputs — the same computation the nightly job uses, so a manual
    run and the scheduled run judge terms by the same bar.
    """
    from src.amazon_ads.actions_engine import generate_recommendations
    from src.amazon_ads.strategy import account_target_acos
    from src.db import job_start, job_finish

    if target_acos is None:
        target_acos, target_basis = account_target_acos()
    else:
        target_basis = "explicit"

    run_id = job_start("ads_actions")
    try:
        recs = generate_recommendations(target_acos=target_acos, lookback_days=days)
    except Exception as e:
        job_finish(run_id, "fail", str(e)[:400])
        raise click.ClickException(f"Recommendation generation failed: {e}")

    if not recs:
        msg = (f"No recommendations — no search term data in the last {days} days, "
               f"or everything is within the {target_acos:.1f}% target")
        click.echo(msg)
        job_finish(run_id, "success", msg,
                   stats={"count": 0, "target_acos": target_acos,
                          "target_basis": target_basis, "days": days})
        return

    click.echo(f"{len(recs)} recommendations ({days}d window, "
               f"target ACOS {target_acos:.1f}% [{target_basis}]):")
    for r in recs[:15]:
        click.echo(f"  [{r['priority']}] {r['type']}: {r.get('entity_name','')[:40]}")
        click.echo(f"       Impact: ${r['impact_estimate']:.2f}  {r['suggested_action'][:80]}")
    job_finish(run_id, "success",
               f"{len(recs)} recs ({days}d, target {target_acos:.1f}%)",
               stats={"count": len(recs), "target_acos": target_acos,
                      "target_basis": target_basis, "days": days})


@cli.command("ads-spend-audit")
@click.option("--days", default=7, show_default=True, help="Closed days to audit")
@click.option("--expect-spend", default=None, type=float,
              help="Seller Central console total for the same window")
def ads_spend_audit_cmd(days, expect_spend):
    """Reconcile 7D spend: by day, by ad product, and placement vs total.

    Three numbers should agree: the /ppc header, the placement panel, and the
    console. This prints all three inputs so a gap is attributable instead of
    mysterious.
    """
    from collections import defaultdict

    from src.db import get_client
    from src.rules import amazon_as_of, window_start

    client = get_client()
    asof = amazon_as_of()
    start = window_start(asof, days)

    def page(table, cols):
        rows, offset = [], 0
        while True:
            p = (client.table(table).select(cols)
                 .gte("date", str(start)).lte("date", str(asof))
                 .order("date").range(offset, offset + 999).execute().data) or []
            rows.extend(p)
            if len(p) < 1000:
                break
            offset += 1000
        return rows

    click.echo(f"Ads spend audit — {days} closed day(s) {start} → {asof}")
    click.echo("  as-of rule: yesterday in America/Los_Angeles. Today is still "
               "accruing and is deliberately excluded from every window.")

    camp = page("ads_campaigns_daily", "date,campaign_type,spend,clicks")
    total = sum(float(r["spend"] or 0) for r in camp)

    by_type = defaultdict(float)
    for r in camp:
        by_type[(r.get("campaign_type") or "SP").upper()] += float(r["spend"] or 0)
    by_day = defaultdict(float)
    for r in camp:
        by_day[r["date"]] += float(r["spend"] or 0)

    click.echo(f"\n  ads_campaigns_daily TOTAL: ${total:,.2f}  ({len(camp)} rows)")
    click.echo("  by day:")
    for d in sorted(by_day):
        click.echo(f"    {d}  ${by_day[d]:>9,.2f}")
    click.echo("  by ad product:")
    for t in ("SP", "SB", "SD"):
        v = by_type.get(t)
        click.echo(f"    {t}  " + (f"${v:>9,.2f}" if v is not None
                                   else "        —   (no rows stored)"))

    try:
        pl = page("ads_placement_daily", "date,placement,spend")
        pl_total = sum(float(r["spend"] or 0) for r in pl)
        click.echo(f"\n  ads_placement_daily TOTAL: ${pl_total:,.2f}  ({len(pl)} rows)")
        by_pl = defaultdict(float)
        for r in pl:
            by_pl[r.get("placement") or "Unknown"] += float(r["spend"] or 0)
        for k, v in sorted(by_pl.items(), key=lambda x: -x[1]):
            click.echo(f"    {k:<28} ${v:>9,.2f}")
        diff = total - pl_total
        if abs(diff) < 0.01:
            click.echo("    placement sum == campaign total ✓")
        else:
            click.echo(f"    UNALLOCATED: ${diff:,.2f} — placements do not sum to "
                       f"the campaign total. Placement data is Sponsored Products "
                       f"only, so SB/SD spend can never appear here.")
    except Exception as e:
        click.echo(f"\n  placement table unavailable: {str(e)[:120]}")

    if expect_spend is not None:
        gap = total - float(expect_spend)
        pct = abs(gap) / float(expect_spend) * 100 if expect_spend else 0
        click.echo(f"\n  console  : ${float(expect_spend):,.2f}")
        click.echo(f"  agent    : ${total:,.2f}")
        click.echo(f"  gap      : ${gap:+,.2f}  ({pct:.1f}%)")
        missing = [t for t in ("SB", "SD") if t not in by_type]
        if missing and gap < 0:
            click.echo(f"  LIKELY CAUSE: no {'/'.join(missing)} rows stored. The "
                       f"console total spans Sponsored Products + Brands + "
                       f"Display; this agent has only "
                       f"{'/'.join(sorted(by_type))}. Run "
                       f"`ads-sync --days {days} --campaigns-only` and check "
                       f"whether the SB/SD reports complete.")

    click.echo("\n  Monitoring aid — Amazon remains the source of truth for spend.")


@cli.command("ads-reconcile")
@click.option("--date", "target_date", default=None,
              help="Day to reconcile (YYYY-MM-DD). Defaults to the LA as-of day.")
@click.option("--expect-spend", default=None, type=float, help="Amazon Ads console spend")
@click.option("--expect-clicks", default=None, type=int, help="Amazon Ads console clicks")
@click.option("--tolerance-pct", default=2.0, help="Gap that counts as a mismatch")
def ads_reconcile_cmd(target_date, expect_spend, expect_clicks, tolerance_pct):
    """Compare stored ads totals for a day against the Amazon Ads console.

    Amazon is the source of truth for ad spend and clicks. This prints what the
    agent has stored, using the ACTUAL column names on ads_campaigns_daily:

        spend        -> spend
        clicks       -> clicks
        attributed   -> sales_14d     (there is no `sales` column)
        orders       -> orders_14d    (there is no `orders`/`purchases` column)
        impressions  -> impressions
    """
    from datetime import date as _date
    from src.db import get_client
    from src.rules import amazon_as_of

    day = target_date or amazon_as_of().isoformat()
    try:
        _date.fromisoformat(day)
    except ValueError:
        raise click.UsageError(f"--date must be YYYY-MM-DD, got {day!r}")

    client = get_client()
    rows, offset = [], 0
    while True:
        page = (client.table("ads_campaigns_daily")
                .select("campaign_id,campaign_name,campaign_status,campaign_type,"
                        "spend,clicks,impressions,sales_14d,orders_14d")
                .eq("date", day).range(offset, offset + 999).execute().data) or []
        rows.extend(page)
        if len(page) < 1000:
            break
        offset += 1000

    spend = sum(float(r.get("spend") or 0) for r in rows)
    clicks = sum(int(r.get("clicks") or 0) for r in rows)
    impressions = sum(int(r.get("impressions") or 0) for r in rows)
    sales = sum(float(r.get("sales_14d") or 0) for r in rows)
    orders = sum(int(r.get("orders_14d") or 0) for r in rows)

    click.echo(f"ads_campaigns_daily — {day}  (America/Los_Angeles reporting day)")
    click.echo(f"  campaign rows : {len(rows)}")
    click.echo(f"  spend         : ${spend:,.2f}   [column: spend]")
    click.echo(f"  clicks        : {clicks:,}        [column: clicks]")
    click.echo(f"  impressions   : {impressions:,}")
    click.echo(f"  CPC           : ${spend / clicks:,.2f}" if clicks else "  CPC           : —")
    click.echo(f"  CTR           : {clicks / impressions * 100:.2f}%" if impressions else "  CTR           : —")
    click.echo(f"  attributed sales: ${sales:,.2f}   [column: sales_14d, 14-day attribution]")
    click.echo(f"  orders        : {orders:,}        [column: orders_14d]")
    click.echo(f"  ROAS          : {sales / spend:.2f}" if spend else "  ROAS          : —")
    click.echo(f"  ACOS          : {spend / sales * 100:.1f}%" if sales else "  ACOS          : —")

    active = sum(1 for r in rows if str(r.get("campaign_status", "")).upper() == "ENABLED")
    click.echo(f"  campaigns     : {active} enabled / {len(rows)} rows")

    # By ad product. The console's daily total spans Sponsored Products,
    # Brands and Display, so a per-type breakdown is what makes an agent-vs-
    # console gap diagnosable rather than just visible.
    by_type: dict[str, dict] = {}
    for r in rows:
        t = (r.get("campaign_type") or "SP").upper()
        b = by_type.setdefault(t, {"rows": 0, "spend": 0.0, "clicks": 0,
                                   "impressions": 0, "sales": 0.0})
        b["rows"] += 1
        b["spend"] += float(r.get("spend") or 0)
        b["clicks"] += int(r.get("clicks") or 0)
        b["impressions"] += int(r.get("impressions") or 0)
        b["sales"] += float(r.get("sales_14d") or 0)

    click.echo("")
    click.echo("  by campaign_type:")
    click.echo(f"    {'type':<6} {'rows':>5} {'spend':>11} {'clicks':>8} "
               f"{'share':>7} {'CPC':>7}")
    for t in ("SP", "SB", "SD"):
        b = by_type.get(t)
        if not b:
            click.echo(f"    {t:<6} {'—':>5} {'—':>11} {'—':>8} {'—':>7} {'—':>7}"
                       "   (no rows stored for this day)")
            continue
        share = (b["spend"] / spend * 100) if spend else 0.0
        cpc = (b["spend"] / b["clicks"]) if b["clicks"] else 0.0
        click.echo(f"    {t:<6} {b['rows']:>5} ${b['spend']:>10,.2f} "
                   f"{b['clicks']:>8,} {share:>6.1f}% "
                   f"{('$%.2f' % cpc) if b['clicks'] else '—':>7}")
    other = {t: b for t, b in by_type.items() if t not in ("SP", "SB", "SD")}
    for t, b in sorted(other.items()):
        share = (b["spend"] / spend * 100) if spend else 0.0
        click.echo(f"    {t:<6} {b['rows']:>5} ${b['spend']:>10,.2f} "
                   f"{b['clicks']:>8,} {share:>6.1f}%")
    click.echo(f"    {'TOTAL':<6} {len(rows):>5} ${spend:>10,.2f} {clicks:>8,} "
               f"{100.0 if spend else 0.0:>6.1f}%")
    missing = [t for t in ("SB", "SD") if t not in by_type]
    if missing:
        click.echo(f"    note: no {'/'.join(missing)} rows — if the console shows "
                   f"spend there, run: ads-sync --days N --campaigns-only")
    click.echo("")

    mismatch = False
    for label, got, want in (("spend", spend, expect_spend), ("clicks", float(clicks), expect_clicks)):
        if want is None:
            continue
        gap = got - float(want)
        pct = (abs(gap) / float(want) * 100) if want else 0.0
        ok = pct <= tolerance_pct
        mismatch = mismatch or not ok
        arrow = "OK " if ok else "GAP"
        fmt_got = f"${got:,.2f}" if label == "spend" else f"{got:,.0f}"
        fmt_want = f"${float(want):,.2f}" if label == "spend" else f"{float(want):,.0f}"
        click.echo(f"  [{arrow}] {label}: agent {fmt_got} vs console {fmt_want}  "
                   f"({gap:+,.2f}, {pct:.2f}%)")
    if (expect_spend or expect_clicks) and not mismatch:
        click.echo("  Within tolerance — Amazon remains the source of truth for spend/clicks.")


@cli.command("ads-heal")
@click.option("--days", default=None, type=int,
              help="Lookback closed days (default: ads.campaign_sb_sd_daily_days)")
@click.option("--dry-run", is_flag=True,
              help="Only list gaps — do not call the Ads API")
def ads_heal_cmd(days, dry_run):
    """Backfill missing Sponsored Brands / Display days automatically.

    Finds closed LA days that have Sponsored Products rows but no SB/SD, then
    re-pulls only those products in 1-day chunks. The midday scheduler and the
    post-partial retry after a nightly SB/SD failure both use this path so
    console gaps do not wait for a human.
    """
    from src.amazon_ads.heal import missing_product_days, sync_missing_sb_sd
    from src.db import get_client
    from src.rules import ADS_SB_SD_DAILY_DAYS, amazon_as_of, window_start

    lookback = days or ADS_SB_SD_DAILY_DAYS
    as_of = amazon_as_of()
    start = window_start(as_of, lookback)

    if dry_run:
        client = get_client()
        rows, offset = [], 0
        while True:
            page = (client.table("ads_campaigns_daily")
                    .select("date,campaign_type")
                    .gte("date", start.isoformat())
                    .lte("date", as_of.isoformat())
                    .range(offset, offset + 999).execute().data) or []
            rows.extend(page)
            if len(page) < 1000:
                break
            offset += 1000
        missing = missing_product_days(rows, as_of=as_of, lookback_days=lookback)
        if not missing:
            click.echo(f"No SB/SD gaps in {start} → {as_of}")
            return
        click.echo(f"Gaps in {start} → {as_of} (dry-run, no API call):")
        for product, ds in missing.items():
            click.echo(f"  {product}: {', '.join(d.isoformat() for d in ds)}")
        return

    from src.db import job_finish, job_start
    run_id = job_start("ads_sb_sd_heal")
    try:
        out = sync_missing_sb_sd(lookback_days=lookback, on_progress=click.echo)
    except Exception as e:
        job_finish(run_id, "fail", str(e)[:500])
        raise
    if not out.get("healed"):
        job_finish(run_id, "success", out.get("reason") or "no gaps")
        click.echo(f"Ads heal: {out.get('reason') or 'no gaps'}")
        return
    camp = (out.get("result") or {}).get("campaigns") or {}
    bad = camp.get("products_failed") or []
    status = "partial" if bad else "success"
    msg = (f"healed {','.join(out.get('products') or [])} "
           f"{out['window']['start']}→{out['as_of']}")
    if bad:
        msg += f" — {'+'.join(bad)} still missing"
    job_finish(run_id, status, msg, stats=camp if isinstance(camp, dict) else None)
    click.echo(f"Ads heal {status}: {msg}")


@cli.command("ads-outcomes")
@click.option("--dry-run", is_flag=True, help="Show what would be written, write nothing")
@click.option("--as-of", default=None, help="Override the LA as-of date (YYYY-MM-DD)")
def ads_outcomes_cmd(dry_run, as_of):
    """Snapshot outcomes for applied/dismissed actions whose window has closed.

    Idempotent — a horizon already recorded is skipped, and a horizon whose
    window has not closed yet is not written at all.
    """
    from datetime import date as _date
    from src.amazon_ads.learning import snapshot_outcomes

    target = _date.fromisoformat(as_of) if as_of else None
    r = snapshot_outcomes(as_of=target, dry_run=dry_run)
    if r.get("skipped"):
        click.echo(f"Skipped: {r['skipped']}")
        return
    if r.get("error"):
        raise click.ClickException(r["error"])
    click.echo(f"as-of {r['as_of']}  horizons {r['offsets']}")
    click.echo(f"  decisions considered : {r['decisions_considered']}")
    click.echo(f"  due this run         : {r.get('due', 0)}")
    click.echo(f"  written              : {r['written']}{' (dry run)' if dry_run else ''}")
    click.echo(f"  not due yet          : {r['skipped_not_due']}")
    click.echo(f"  already recorded     : {r['already_present']}")
    for sample in (r.get("sample") or [])[:3]:
        click.echo(f"    +{sample['horizon_days']}d {sample['window_start']}→{sample['window_end']}"
                   f"  spend ${sample['spend']:.2f} sales ${sample['ad_sales']:.2f}")


@cli.command("ads-impact")
def ads_impact_cmd():
    """Observational summary of action decisions and their outcomes."""
    from src.amazon_ads.learning import impact_summary
    r = impact_summary()
    if not r.get("available"):
        click.echo(r.get("note") or r.get("error") or "Learning tables unavailable")
        return
    t = r["totals"]
    click.echo(f"decisions {t['decisions']}  (applied {t['applied']}, dismissed {t['dismissed']}, "
               f"open {t['open']})   outcome rows {t['outcomes']}")
    click.echo(f"\n{'action type':22s}{'total':>7s}{'open':>7s}{'applied':>9s}{'dismissed':>11s}")
    for k, v in sorted(r["by_type"].items(), key=lambda kv: -kv[1]["total"]):
        click.echo(f"{k:22s}{v['total']:7d}{v['open']:7d}{v['applied']:9d}{v['dismissed']:11d}")
    for h, statuses in sorted(r["horizons"].items()):
        click.echo(f"\n  {h} post-decision window:")
        for st, v in statuses.items():
            click.echo(f"    {st:10s} n={v['n']:3d}  spend ${v['spend']:>9,.2f}  "
                       f"ad sales ${v['ad_sales']:>9,.2f}  ACOS {v['acos']}%")
    click.echo(f"\n{r['caveat']}")


@cli.command("jobs")
@click.option("--limit", default=20, help="How many runs to show")
@click.option("--job", default=None, help="Filter to one job_name")
@click.option("--failures", is_flag=True, help="Show only fail/partial runs")
def jobs_cmd(limit, job, failures):
    """Show recent scheduled-job runs from job_runs (what the agent has done)."""
    from src.db import get_client
    from datetime import datetime, timezone

    q = get_client().table("job_runs").select("*").order("started_at", desc=True)
    if job:
        q = q.eq("job_name", job)
    if failures:
        q = q.in_("status", ["fail", "partial"])
    rows = (q.limit(limit).execute().data) or []

    if not rows:
        click.echo("No job runs recorded yet.")
        return

    now = datetime.now(timezone.utc)
    click.echo(f"{'STARTED (local)':<20s} {'JOB':<24s} {'STATUS':<9s} {'TOOK':>7s}  MESSAGE")
    click.echo("─" * 110)
    for r in rows:
        started_raw = str(r.get("started_at") or "")
        try:
            started = datetime.fromisoformat(started_raw.replace("Z", "+00:00"))
            started_local = started.astimezone().strftime("%Y-%m-%d %H:%M:%S")
        except Exception:
            started, started_local = None, started_raw[:19]

        took = "—"
        if started:
            finished_raw = r.get("finished_at")
            if finished_raw:
                try:
                    finished = datetime.fromisoformat(str(finished_raw).replace("Z", "+00:00"))
                    took = f"{(finished - started).total_seconds():.0f}s"
                except Exception:
                    pass
            elif r.get("status") == "running":
                took = f"{(now - started).total_seconds() / 60:.0f}m…"

        status = str(r.get("status") or "?")
        colour = {"success": "green", "partial": "yellow",
                  "fail": "red", "running": "cyan"}.get(status)
        click.echo(f"{started_local:<20s} {str(r.get('job_name'))[:24]:<24s} "
                   f"{click.style(f'{status:<9s}', fg=colour)} {took:>7s}  "
                   f"{str(r.get('message') or '')[:50]}")


@cli.command("ads-waste")
@click.option("--days", default=14, help="Lookback days")
def ads_waste_cmd(days):
    """Show top wasted ad spend."""
    from src.db import fetch_all
    try:
        search_terms = fetch_all("ads_search_terms_daily")
    except Exception:
        click.echo("No search term data. Run ads-sync first.")
        return
    # Group zero-order spend by search term
    waste = {}
    for st in search_terms:
        orders = int(st.get("orders_14d", 0) or 0)
        if orders == 0:
            term = st.get("search_term", "?")
            waste[term] = waste.get(term, 0) + float(st.get("spend", 0) or 0)
    sorted_waste = sorted(waste.items(), key=lambda x: -x[1])
    total = sum(v for _, v in sorted_waste)
    click.echo(f"Total wasted spend (0 orders): ${total:,.2f} across {len(sorted_waste)} terms")
    for term, amount in sorted_waste[:20]:
        click.echo(f"  ${amount:>8.2f}  {term[:50]}")


@cli.command("economics-sync")
@click.option("--days", default=30, help="Days to fetch")
def economics_sync_cmd(days):
    """Fetch Amazon financial transactions → reconciled net proceeds."""
    from src.amazon_sp.economics import sync_economics

    click.echo(f"Fetching Amazon financial transactions ({days}d)...")
    result = sync_economics(days=days)
    click.echo(f"Transactions: {result['transactions']}")
    click.echo(f"Days with data: {result['days']}")
    click.echo(f"Rows upserted: {result['inserted']}")
    click.echo(f"Sales:         ${result.get('total_sales', 0):,.2f}")
    click.echo(f"Fees:          ${result.get('total_fees', 0):,.2f}")
    click.echo(f"Ad spend:      ${result.get('total_ad_spend', 0):,.2f}")
    click.echo(f"COGS:          ${result.get('total_cogs', 0):,.2f}")
    click.echo(f"Contribution:  ${result.get('total_contribution', 0):,.2f}"
               f"  (sales - fees - ads - COGS)")
    click.echo(f"Settled days:  {result.get('settled_days', 0)} of {result.get('days', 0)}")
    click.echo(f"Amazon payout: ${result.get('total_payout', 0):,.2f}"
               f"  — settlement reconciliation only, ~2x/month, postedDate basis")


@cli.command("costs-import")
@click.argument("file_path")
def costs_import_cmd(file_path):
    """Import COGS from xlsx or csv into sku_costs."""
    from pathlib import Path
    from src.db import upsert_rows

    path = Path(file_path)
    if not path.exists():
        click.echo(f"File not found: {path}")
        return

    rows: list[dict] = []

    if path.suffix.lower() in (".xlsx", ".xls"):
        import openpyxl
        wb = openpyxl.load_workbook(str(path), data_only=True)
        ws = wb.active
        headers = [str(c.value or "").strip().lower() for c in next(ws.iter_rows(min_row=1, max_row=1))]

        sku_col = next((i for i, h in enumerate(headers) if "sku" in h), 0)
        name_col = next((i for i, h in enumerate(headers) if "name" in h or "product" in h), 1)
        cost_col = next((i for i, h in enumerate(headers) if "cost" in h or "cogs" in h), 2)

        for row in ws.iter_rows(min_row=2, values_only=True):
            sku = str(row[sku_col] or "").strip()
            name = str(row[name_col] or "").strip() if name_col < len(row) else ""
            cost = float(row[cost_col] or 0) if cost_col < len(row) else 0
            if sku and cost >= 0:
                rows.append({"sku": sku, "product_name": name or None,
                            "cogs_per_unit": round(cost, 4), "source": path.name})
    else:
        import csv
        with open(path) as f:
            reader = csv.DictReader(f)
            for r in reader:
                sku = (r.get("SKU") or r.get("sku") or "").strip()
                name = (r.get("Product Name") or r.get("product_name") or "").strip()
                cost = float(r.get("Cost Per Unit") or r.get("cogs_per_unit") or 0)
                if sku:
                    rows.append({"sku": sku, "product_name": name or None,
                                "cogs_per_unit": round(cost, 4), "source": path.name})

    # Handle known typo: DDPE00019Shop ↔ DDPE0019Shop
    extra: list[dict] = []
    for r in rows:
        if r["sku"] == "DDPE00019Shop":
            extra.append({**r, "sku": "DDPE0019Shop"})
        elif r["sku"] == "DDPE0019Shop":
            extra.append({**r, "sku": "DDPE00019Shop"})
    rows.extend(extra)

    if not rows:
        click.echo("No rows found.")
        return

    try:
        inserted = upsert_rows("sku_costs", rows, on_conflict="sku")
    except Exception:
        # product_name column may not exist — retry without it
        for r in rows:
            r.pop("product_name", None)
        inserted = upsert_rows("sku_costs", rows, on_conflict="sku")
    click.echo(f"Imported {len(rows)} SKU costs ({inserted} upserted) from {path.name}")
    for r in sorted(rows, key=lambda x: x["sku"])[:5]:
        click.echo(f"  {r['sku']:<18} ${r['cogs_per_unit']:>6.2f}  {(r.get('product_name') or '')[:40]}")
    if len(rows) > 5:
        click.echo(f"  ... and {len(rows)-5} more")


@cli.command("pnl-validate")
@click.option("--date", "target_date", required=True, help="Date to validate YYYY-MM-DD")
def pnl_validate_cmd(target_date):
    """Print detailed P&L breakdown for a single day (for Seller Central comparison)."""
    from src.amazon_sp.economics import validate_day

    click.echo(f"Validating P&L for {target_date}...")
    v = validate_day(target_date)
    click.echo(f"{'='*55}")
    click.echo(f"  Date: {v['date']} ({v['date_basis']})")
    click.echo(f"  Transactions: {v['transaction_types']}")
    click.echo(f"  Units shipped: {v['units_shipped']}")
    click.echo(f"{'='*55}")
    click.echo(f"  Product charges:    ${v['product_charges']:>10,.2f}")
    click.echo(f"  Refund charges:     ${v['refund_charges']:>10,.2f}")
    click.echo(f"  Amazon fees:        ${v['amazon_fees_display']:>10,.2f}  (referral + FBA)")
    click.echo(f"  Other adjustments:  ${v['other_adjustments']:>10,.2f}")
    click.echo(f"  {'─'*40}")
    click.echo(f"  Amazon payout:      ${v['amazon_payout']:>10,.2f}")
    click.echo(f"{'='*55}")
    click.echo(f"  Formula: {v['formula']}")
    click.echo(f"  Compare to: {v['compare_to']}")


def _entity_line(o) -> str:
    due = o.due_date.isoformat() if o.due_date else "no date"
    if o.days_overdue:
        when = f"OVERDUE {o.days_overdue}d"
    elif o.days_until_due is not None:
        when = f"in {o.days_until_due}d"
    else:
        when = "—"
    amt = f" ~${o.amount_estimate:,.0f}" if o.amount_estimate else ""
    return (f"    {o.state_code} {o.form_code:<24} {o.period_label}  "
            f"due {due:<12} {when:<14} [{o.confidence}]{amt}")


@cli.command("inventory-health")
@click.option("--unknown-limit", default=20, help="How many unknown FC codes to list")
def inventory_health_cmd(unknown_limit):
    """Freshness and coverage of the FBA inventory ledger.

    Physical nexus — and therefore which states to register in — is driven by
    where inventory has been held. A ledger that stopped updating, or an FC
    code with no state mapping, both look like "no nexus" rather than an error.
    """
    from datetime import datetime, timezone
    from src.db import fetch_all, get_client
    from src.inventory.ledger_health import build_health

    client = get_client()
    events, offset = [], 0
    while True:
        page = (client.table("inventory_events")
                .select("event_date,fc_code,state_code,source_file")
                .range(offset, offset + 999).execute().data) or []
        events.extend(page)
        if len(page) < 1000:
            break
        offset += 1000

    h = build_health(events, fetch_all("job_runs"), datetime.now(timezone.utc))

    icon = {"ok": "OK", "stale": "STALE", "critical": "CRITICAL"}[h.status]
    click.echo(f"FBA inventory ledger — {icon}")
    if h.last_success_at:
        click.echo(f"  last successful sync : {h.last_success_at[:19]} "
                   f"({h.hours_since_success:.1f}h ago, {h.last_success_status})")
    else:
        click.echo("  last successful sync : none on record")
    click.echo(f"  latest event_date    : {h.date_max}")
    click.echo(f"  earliest event_date  : {h.date_min}")
    click.echo(f"  total events         : {h.total_events:,}")
    click.echo(f"  distinct states      : {h.distinct_states}")
    click.echo(f"  events by year       : "
               + ", ".join(f"{y} {n:,}" for y, n in h.events_by_year.items()))
    click.echo(f"  sources              : "
               + ", ".join(f"{k} {v:,}" for k, v in list(h.sources.items())[:3]))
    click.echo(f"  states               : {', '.join(h.states)}")

    if h.unknown_fcs:
        click.echo("")
        click.echo(f"  UNMAPPED FC CODES: {len(h.unknown_fcs)} distinct, "
                   f"{h.unknown_event_count:,} event(s)")
        click.echo("  These have NO state, so the physical-nexus engine cannot see them.")
        click.echo("  Look each code up in Seller Central (Inventory > Shipments shows the")
        click.echo("  destination address) and add it to config/fc_codes.json, then run:")
        click.echo("    python -m src.main inventory-remap-fc --apply")
        click.echo("")
        click.echo(f"    {'code':<9}{'events':>7}  {'first seen':<12}{'last seen':<12}")
        for u in h.unknown_fcs[:unknown_limit]:
            click.echo(f"    {u.fc_code:<9}{u.events:>7}  {u.first_seen:<12}{u.last_seen:<12}")
        if len(h.unknown_fcs) > unknown_limit:
            click.echo(f"    +{len(h.unknown_fcs) - unknown_limit} more")

    # States with inventory that the nexus engine does NOT count. The user's
    # registration criterion is "states with Amazon inventory since 2024", so
    # the difference between those two sets is exactly what needs to be visible
    # — not silently reconciled. The reason comes from config/state_rules.json;
    # nothing here decides it.
    try:
        from src.config import load_state_rules
        from src.engines.physical_nexus import evaluate_physical_nexus

        rules = load_state_rules().get("states", {})
        nexus_states = set(evaluate_physical_nexus().get("nexus_states") or [])
        gap = [st for st in h.states if st not in nexus_states]
        if gap:
            click.echo("")
            click.echo(f"  INVENTORY BUT NOT COUNTED AS PHYSICAL NEXUS: {len(gap)} state(s)")
            click.echo("  You have stored inventory here, but state_rules says FBA stock")
            click.echo("  alone does not create nexus (or the state has no sales tax).")
            click.echo("  Worth a CPA conversation if you are registering on inventory history.")
            for st in gap:
                r = rules.get(st, {})
                why = ("no state sales tax" if not r.get("has_sales_tax", True)
                       else f"fba_inventory_creates_nexus = {r.get('fba_inventory_creates_nexus')}")
                click.echo(f"    {st}: {why}")
    except Exception as e:
        click.echo(f"  (could not compare against physical nexus: {str(e)[:80]})")

    if h.is_stale:
        click.echo("")
        click.echo("  ⚠ The ledger is stale. Check that the scheduler is running:")
        click.echo("    launchctl list | grep tallowbourn      # launchd agent")
        click.echo("    python -m src.main jobs                 # recent job outcomes")


@cli.command("inventory-remap-fc")
@click.option("--dry-run/--apply", default=True, help="Dry run by default")
@click.option("--recheck", is_flag=True,
              help="Also CORRECT rows whose stored state disagrees with the "
                   "current map. Use after fixing a wrong mapping — the default "
                   "pass only fills in states that are missing.")
def inventory_remap_fc_cmd(dry_run, recheck):
    """Re-resolve state_code on stored events after updating fc_codes.json.

    state_code is written at parse time, so adding a code to config/fc_codes.json
    does NOT retroactively fix events already stored — they keep a null state and
    stay invisible to physical nexus. This backfills them.

    Only fills in states that are currently missing; it never overwrites or
    deletes an existing mapping.

    /tax-inventory unknown_fcs reads inventory_ledger_summary_daily (state_code
    written at parse time). After adding codes, re-run the existing
    `spapi-ledger-summary` / `backfill-ledger-summary` upsert so those rows
    pick up the new map — do not invent a second remap command.
    """
    from src.db import get_client, log_audit
    from src.mappers.fc_to_state import fc_to_state

    client = get_client()
    rows, offset = [], 0
    while True:
        page = (client.table("inventory_events").select("id,fc_code,state_code")
                .is_("state_code", "null")
                .range(offset, offset + 999).execute().data) or []
        rows.extend(page)
        if len(page) < 1000:
            break
        offset += 1000

    fixable, still_unknown = {}, {}
    for r in rows:
        fc = r.get("fc_code")
        if not fc:
            continue
        state = fc_to_state(fc)
        if state:
            fixable.setdefault(fc, []).append(r["id"])
        else:
            still_unknown[fc] = still_unknown.get(fc, 0) + 1

    # Rows whose stored state contradicts the map. These exist when a mapping
    # was wrong and has since been corrected — the fill-in pass cannot see them
    # because they already have a (wrong) state.
    mismatched: dict[str, list] = {}
    if recheck:
        allrows, offset = [], 0
        while True:
            page = (client.table("inventory_events").select("id,fc_code,state_code")
                    .not_.is_("state_code", "null")
                    .range(offset, offset + 999).execute().data) or []
            allrows.extend(page)
            if len(page) < 1000:
                break
            offset += 1000
        for r in allrows:
            fc, cur = r.get("fc_code"), r.get("state_code")
            if not fc:
                continue
            want = fc_to_state(fc)
            if want and cur and want != cur:
                mismatched.setdefault(f"{fc}:{cur}->{want}", []).append(r["id"])

    total = sum(len(v) for v in fixable.values())
    click.echo(f"{'DRY RUN — ' if dry_run else ''}events with no state: {len(rows):,}")
    click.echo(f"  now resolvable from fc_codes.json: {total:,} "
               f"across {len(fixable)} code(s)")
    for fc, ids in sorted(fixable.items(), key=lambda kv: -len(kv[1])):
        click.echo(f"    {fc} → {fc_to_state(fc)}  ({len(ids):,} events)")
    if still_unknown:
        click.echo(f"  still unmapped: {sum(still_unknown.values()):,} event(s) "
                   f"across {len(still_unknown)} code(s): "
                   f"{', '.join(sorted(still_unknown)[:12])}")

    if mismatched:
        n = sum(len(v) for v in mismatched.values())
        click.echo(f"  MISMATCHED against the current map: {n:,} event(s)")
        for key, ids in sorted(mismatched.items(), key=lambda kv: -len(kv[1])):
            click.echo(f"    {key}  ({len(ids):,} events)")

    if dry_run or not (total or mismatched):
        if dry_run and (total or mismatched):
            click.echo("\n  Re-run with --apply to write these.")
        return

    updated = 0
    for fc, ids in fixable.items():
        state = fc_to_state(fc)
        for i in range(0, len(ids), 200):
            batch = ids[i:i + 200]
            client.table("inventory_events").update(
                {"state_code": state}).in_("id", batch).execute()
            updated += len(batch)
    corrected = 0
    for key, ids in mismatched.items():
        want = key.split("->")[-1]
        for i in range(0, len(ids), 200):
            batch = ids[i:i + 200]
            client.table("inventory_events").update(
                {"state_code": want}).in_("id", batch).execute()
            corrected += len(batch)
    if corrected:
        click.echo(f"  Corrected {corrected:,} mismatched event(s).")

    log_audit(action="inventory_remap_fc", category="ingestion",
              details={"events_updated": updated, "events_corrected": corrected,
                       "codes": sorted(fixable),
                       "mismatched": sorted(mismatched)})
    click.echo(f"\n  Updated {updated:,} event(s). Re-run `analyze` so physical "
               f"nexus picks up any newly-covered states.")


@cli.command("brand-reclassify")
@click.option("--dry-run/--apply", default=True, help="Dry run by default")
def brand_reclassify_cmd(dry_run):
    """Re-apply brand rules to stored SQP rows after editing brand_terms.json.

    `is_branded` is written at ingest time, so changing the term list does not
    retroactively fix history — the 2025-10-31 rename in particular would leave
    legacy 'Dr. Dave's Primal Essence' weeks misfiled as non-brand. This
    recomputes stored rows; it never re-fetches from Amazon, so it costs no quota.
    """
    from src.amazon_ads.brand_terms import classify
    from src.db import get_client

    client = get_client()
    rows, offset = [], 0
    try:
        while True:
            page = (client.table("sqp_weekly")
                    .select("id,query_normalized,is_branded,brand_rule,week_start")
                    .order("week_start").order("query_normalized")
                    .range(offset, offset + 999).execute().data) or []
            rows.extend(page)
            if len(page) < 1000:
                break
            offset += 1000
    except Exception as e:
        raise click.ClickException(f"sqp_weekly unavailable: {str(e)[:140]}")

    changes = []
    for r in rows:
        c = classify(r.get("query_normalized") or "")
        if bool(r.get("is_branded")) != c["branded"] or \
                (r.get("brand_rule") or None) != c["matched_rule"]:
            changes.append({"id": r["id"], "query": r.get("query_normalized"),
                            "was": bool(r.get("is_branded")),
                            "now": c["branded"], "rule": c["matched_rule"],
                            "week": r.get("week_start")})

    click.echo(f"{'DRY RUN — ' if dry_run else ''}brand reclassification")
    click.echo(f"  rows scanned : {len(rows):,}")
    click.echo(f"  changes      : {len(changes):,}")
    flips = [c for c in changes if c["was"] != c["now"]]
    gained = [c for c in flips if c["now"]]
    lost = [c for c in flips if not c["now"]]
    click.echo(f"    now branded    : {len(gained)}")
    click.echo(f"    now non-brand  : {len(lost)}")
    for c in (gained + lost)[:12]:
        click.echo(f"    {c['week']}  {str(c['query'])[:40]:<41} "
                   f"{c['was']} -> {c['now']}  ({c['rule']})")

    if dry_run or not changes:
        if dry_run and changes:
            click.echo("\n  Re-run with --apply to write, then `brand-share`.")
        return

    for c in changes:
        client.table("sqp_weekly").update(
            {"is_branded": c["now"], "brand_rule": c["rule"]}
        ).eq("id", c["id"]).execute()
    click.echo(f"\n  Updated {len(changes):,} row(s). Run `brand-share` to see "
               f"the corrected mix, and `ads-actions` so bid caps pick up any "
               f"newly-branded terms.")


@cli.command("brand-share")
@click.option("--weeks", default=15, help="Weeks of trend to show")
@click.option("--opportunities", default=15, help="Top non-brand opportunities")
def brand_share_cmd(weeks, opportunities):
    """Branded vs non-branded market share — the automated tracker DASHBOARD.

    Reads sqp_weekly (populated by sqp-sync). SP-API SQP is the ASIN view, so
    denominators cover queries our ASINs appeared in — not full Brand View
    category coverage.
    """
    from src.amazon_ads.brand_rollup import callouts, rollup_weeks, top_opportunities
    from src.amazon_ads.organic_rank import fetch_ranks
    from src.db import get_client

    client = get_client()
    try:
        # ORDER BY reaches `id`: hundreds of rows share each week_start, so
        # ordering by the date alone leaves page boundaries undefined and rows
        # silently drop and duplicate across the 17 pages this table needs.
        # Ascending with an `id` tiebreak. Two bugs lived in the old version:
        # ordering by week_start alone left page boundaries undefined across 25
        # pages, and a `len(rows) > 20000` cap combined with DESC ordering
        # silently dropped the OLDEST rows once the table outgrew it. The SQP
        # backfill pushed it past 24,000 rows and the CLI quietly lost the three
        # earliest weeks — reporting 30 while the dashboard and brief said 33.
        MAX_ROWS = 200_000
        rows, offset, truncated = [], 0, False
        while True:
            page = (client.table("sqp_weekly").select("*")
                    .order("week_start").order("id")
                    .range(offset, offset + 999).execute().data) or []
            rows.extend(page)
            if len(page) < 1000:
                break
            if len(rows) >= MAX_ROWS:
                truncated = True     # never silent — see the warning below
                break
            offset += 1000
    except Exception as e:
        if "sqp_weekly" in str(e):
            raise click.ClickException(
                "sqp_weekly table missing — run supabase/migration_sqp_weekly.sql, "
                "then `sqp-sync --apply`.")
        raise

    if not rows:
        click.echo("No SQP weeks stored yet. Run `sqp-sync --apply` "
                   "(and apply migration_sqp_weekly.sql).")
        return

    all_weeks = rollup_weeks(rows)
    series = all_weeks[-weeks:]
    click.echo("Branded market share — automated tracker")
    if truncated:
        click.echo(f"  WARNING: stopped at {len(rows):,} rows — the history "
                   f"below is incomplete.")
    # "weeks stored" used to print len(series) — the --weeks DISPLAY cap, not
    # the stored count. It read 15 while the dashboard and the PPC brief both
    # reported 33 from the same table, so three surfaces disagreed about the
    # size of the history. Show both, and say which is which.
    click.echo(f"  weeks stored : {len(all_weeks)}  "
               f"({all_weeks[0].week_start} → {all_weeks[-1].week_start})")
    if len(series) < len(all_weeks):
        click.echo(f"  showing      : last {len(series)} "
                   f"(--weeks {len(all_weeks)} for all)")
    click.echo("")
    click.echo(f"  {'week':<12}{'branded':>9}{'non-brand':>11}{'mix':>7}"
               f"{'nb.share*':>11}{'nb.share(all)':>15}")
    for w in series:
        mix = f"{w.branded_mix:.0%}" if w.branded_mix is not None else "—"
        nsp = (f"{w.non_branded.share_present:.2%}"
               if w.non_branded.share_present is not None else "—")
        ns = f"{w.non_branded.share:.2%}" if w.non_branded.share is not None else "—"
        click.echo(f"  {w.week_start:<12}{w.branded.purchases:>9,}"
                   f"{w.non_branded.purchases:>11,}{mix:>7}{nsp:>11}{ns:>15}")
    click.echo("    * share on queries where we actually appeared "
               "(impression share > 0). The (all) column includes every query "
               "our ASINs touched — dominated by huge terms we have no presence "
               "in, so it understates performance.")

    click.echo("")
    for c in callouts(series):
        click.echo(f"  • {c}")

    ranks = {}
    try:
        ranks = {k[1]: v for k, v in fetch_ranks().items()}
    except Exception:
        pass
    opps = top_opportunities(rows, ranks=ranks, limit=opportunities)
    if opps:
        click.echo("")
        click.echo("  Top non-brand opportunities (market demand we are not capturing):")
        click.echo(f"    {'query':<44}{'market':>8}{'ours':>7}{'share':>8}{'band':>6}")
        for o in opps:
            band = str(o.rank_band) if o.rank_band is not None else "-"
            click.echo(f"    {o.query[:43]:<44}{o.market_purchases:>8,}"
                       f"{o.our_purchases:>7,}{o.share:>7.1%}{band:>6}")

    click.echo("\n  Scope: SP-API SQP is the ASIN view — denominators cover queries "
               "our ASINs appeared in, not the full category. Trend is reliable; "
               "the absolute level is not Brand View parity.")


@cli.command("sqp-sync")
@click.option("--period", type=click.Choice(["WEEK", "MONTH", "QUARTER"]), default=None)
@click.option("--asin", "asins", multiple=True, help="Override configured ASINs")
@click.option("--ref", "ref_date", default=None,
              help="Reference date YYYY-MM-DD — pulls the complete week before this day")
@click.option("--dry-run/--apply", default=True, help="Dry run by default")
@click.option("--no-previous-retry", is_flag=True,
              help="Do not fall back to the prior week on empty/FATAL")
def sqp_sync_cmd(period, asins, ref_date, dry_run, no_previous_retry):
    """Pull Brand Analytics SQP via SP-API into keyword_organic_rank.

    Debugging entry point — the scheduled weekly job is the primary path.
    Requires Brand Registry + the Brand Analytics role on the SP-API app.
    """
    from datetime import date as _date

    from src.amazon_sp.sqp import BrandAnalyticsRoleError, sync_sqp

    ref = None
    if ref_date:
        try:
            ref = _date.fromisoformat(ref_date)
        except ValueError as e:
            raise click.UsageError(f"--ref must be YYYY-MM-DD: {e}") from e

    try:
        r = sync_sqp(asins=list(asins) or None, period=period, ref=ref,
                     dry_run=dry_run, retry_previous=not no_previous_retry)
    except BrandAnalyticsRoleError as e:
        raise click.ClickException(str(e))

    click.echo(f"{'DRY RUN — ' if dry_run else ''}SQP sync via SP-API")
    click.echo(f"  report type : GET_BRAND_ANALYTICS_SEARCH_QUERY_PERFORMANCE_REPORT")
    click.echo(f"  period      : {r['period']}  {r['start']} → {r['end']}")
    if r.get("retried_previous_period"):
        fa = r.get("first_attempt") or {}
        click.echo(f"  retried     : latest period {fa.get('start')}→{fa.get('end')} "
                   f"was empty/FATAL (Amazon publishes ~24-48h after close)")
        for e in (fa.get("errors") or [])[:3]:
            click.echo(f"               first attempt: {e[:120]}")
    res = r.get("asin_resolution") or {}
    click.echo(f"  asins       : {len(r.get('asins') or [])} in {r['batches']} "
               f"batch(es) [{res.get('basis', '?')}]")
    if res.get("note") and res.get("basis") != "config":
        click.echo(f"    ⚠ {res['note']}")
    click.echo(f"    requested : {', '.join((r.get('asins') or [])[:8])}"
               + (" …" if len(r.get('asins') or []) > 8 else ""))
    for d in r.get("diagnostics") or []:
        click.echo(f"    batch {d['batch']}: report {d.get('report_id')} "
                   f"{d.get('doc_bytes', 0):,}B keys={d.get('top_level_keys')} "
                   f"records_key={d.get('had_records_key')} "
                   f"parsed={d.get('records_parsed', 0)}")
    click.echo(f"  parsed      : {r.get('parsed', 0)}  skipped: {r.get('skipped', 0)}")
    click.echo(f"  unique keys : {len(r['rows'])}")
    if not dry_run:
        click.echo(f"  written     : {r.get('written', 0)}")
        if r.get("weekly_written") is not None:
            click.echo(f"  weekly rows : {r.get('weekly_written', 0)}")
    for w in r.get("warnings", []):
        click.echo(f"  ⚠ {w}")
    for e in r.get("errors", []):
        click.echo(f"  ✗ {e}")
    for row in r["rows"][:10]:
        cs = (row.get("raw") or {}).get("click_share")
        click.echo(f"    {row['keyword_normalized'][:44]:<45} band={row['organic_rank']}"
                   + (f"  click_share={cs:.0%}" if cs is not None else ""))
    if dry_run and r["rows"]:
        click.echo("\n  Re-run with --apply to write.")


@cli.command("sqp-backfill")
@click.option("--max-weeks", default=4, show_default=True,
              help="Weeks per invocation. Deliberately small — SQP quota is tight.")
@click.option("--from", "from_date", default=None, help="Earliest week end (YYYY-MM-DD)")
@click.option("--to", "to_date", default=None,
              help="Latest week end (YYYY-MM-DD). Clamped to the last COMPLETE week.")
@click.option("--sleep", "sleep_secs", default=90, show_default=True,
              help="Seconds between report requests")
@click.option("--resume/--no-resume", default=True, show_default=True,
              help="Skip weeks that already have rows")
@click.option("--dry-run/--apply", default=True, help="Dry run by default")
@click.option("--refresh-actions", is_flag=True,
              help="Rebuild ads recommendations once at the end (not per week)")
def sqp_backfill_cmd(max_weeks, from_date, to_date, sleep_secs, resume,
                     dry_run, refresh_actions):
    """Backfill completed SQP weeks so brand-share trends can form.

    Walks Sunday-Saturday weeks backward, ONE report request at a time, with a
    sleep between them. Brand Analytics quota is tight: this is a
    few-weeks-per-day job, not an instant history load.
    """
    from datetime import date as _date

    from src.amazon_sp.sqp import BrandAnalyticsRoleError, backfill

    lo = _date.fromisoformat(from_date) if from_date else None
    hi = _date.fromisoformat(to_date) if to_date else None

    try:
        r = backfill(max_weeks=max_weeks, from_date=lo, to_date=hi,
                     sleep_secs=sleep_secs, resume=resume, dry_run=dry_run,
                     on_progress=click.echo)
    except BrandAnalyticsRoleError as e:
        raise click.ClickException(str(e))

    res = r["asin_resolution"]
    click.echo(f"\n{'DRY RUN — ' if dry_run else ''}SQP backfill")
    click.echo(f"  asins           : {len(r['asins'])} [{res.get('basis')}]")
    if res.get("basis") != "config":
        click.echo(f"    ⚠ {res.get('note')}")
    click.echo(f"  already stored  : {len(r['skipped_existing'])} week(s) skipped "
               f"{sorted(r['skipped_existing'])[-3:]}")
    click.echo(f"  planned         : {len(r['planned'])} week(s)")
    for ws, we in r["planned"]:
        click.echo(f"    {ws} → {we}")

    if not dry_run:
        click.echo(f"  written         : {len(r['weeks_done'])} week(s), "
                   f"{r['rows_written']} rank row(s), "
                   f"{r['weekly_written']} weekly row(s)")
        for w in r["weeks_done"]:
            click.echo(f"    {w['week_start']} → {w['week_end']}: "
                       f"{w['rows']} ranks, {w['weekly']} weekly")
        for w in r["weeks_failed"]:
            click.echo(f"    ✗ {w['week_end']}: {w['error'][:120]}")
        if r["quota_exceeded"]:
            click.echo(f"\n  ⚠ {r['resume_hint']}")

    if dry_run and r["planned"]:
        click.echo(f"\n  Re-run with --apply to fetch these "
                   f"({sleep_secs}s between requests, "
                   f"~{len(r['planned']) * sleep_secs // 60}min).")
        click.echo("  Recommended cadence: 4 weeks/day until caught up.")

    if refresh_actions and not dry_run and r["weeks_done"]:
        click.echo("\n  Rebuilding ads recommendations so the gate sees new ranks…")
        try:
            _run_ads_actions()
        except Exception as e:
            click.echo(f"  refresh failed: {str(e)[:160]}")


@cli.command("sqp-status")
def sqp_status_cmd():
    """Freshness of the organic-rank data the PPC gate reads."""
    from collections import Counter
    from datetime import date as _date

    from src.amazon_ads.organic_rank import fetch_ranks, load_config

    cfg = load_config()
    auto = cfg.get("sqp_auto") or {}
    ranks = fetch_ranks()

    click.echo("SQP / organic-rank status")
    click.echo(f"  auto-sync enabled : {auto.get('enabled', False)}")
    click.echo(f"  configured ASINs  : {', '.join(auto.get('asins') or []) or '(none)'}")
    click.echo(f"  period            : {auto.get('report_period', 'WEEK')}")
    sched = auto.get("schedule") or {}
    if sched:
        click.echo(f"  schedule          : {sched.get('day_of_week')} "
                   f"{sched.get('hour', 0):02d}:{sched.get('minute', 0):02d} "
                   f"{sched.get('timezone')}")
    click.echo(f"  rank rows stored  : {len(ranks)}")
    if not ranks:
        click.echo("    (none — run `sqp-sync --apply`, or the migration is pending)")
        return

    by_source = Counter(str(r.get("source")) for r in ranks.values())
    click.echo(f"  by source         : {dict(by_source)}")
    dates = sorted(str(r.get("as_of") or "") for r in ranks.values() if r.get("as_of"))
    if dates:
        newest = dates[-1]
        age = (_date.today() - _date.fromisoformat(newest)).days
        stale_after = cfg["stale_after_days"]
        flag = "STALE — gates as unknown" if age > stale_after else "fresh"
        click.echo(f"  newest as_of      : {newest} ({age}d ago, {flag})")
        click.echo(f"  oldest as_of      : {dates[0]}")


@cli.command("sqp-import")
@click.argument("path", type=click.Path(exists=True))
@click.option("--asin", default=None, help="ASIN when the export has no ASIN column")
@click.option("--as-of", default=None, help="Reporting date (YYYY-MM-DD)")
@click.option("--dry-run/--apply", default=True, help="Dry run by default")
def sqp_import_cmd(path, asin, as_of, dry_run):
    """Import Brand Analytics Search Query Performance for organic-rank gating.

    SQP is the official Amazon signal. It reports click SHARE, not SERP
    position, so rows without an explicit rank column get a coarse rank BAND
    derived from share — recorded as a band, never as a measured position.
    """
    from datetime import date as _date

    from src.amazon_ads.organic_rank import load_config
    from src.amazon_ads.sqp_import import import_sqp

    cfg = load_config()
    ref = _date.fromisoformat(as_of) if as_of else None
    r = import_sqp(path, default_asin=asin or cfg["default_asin"],
                   as_of=ref, dry_run=dry_run)

    click.echo(f"{'DRY RUN — ' if dry_run else ''}SQP import: {path}")
    click.echo(f"  columns matched : {r.get('columns')}")
    click.echo(f"  parsed          : {r['parsed']}")
    click.echo(f"  skipped         : {r['skipped']} (no query or no rank evidence)")
    click.echo(f"  unique keywords : {len(r['rows'])}")
    if not dry_run:
        click.echo(f"  written         : {r.get('written', 0)}")
    for w in r.get("warnings", []):
        click.echo(f"  ⚠ {w}")
    for row in r["rows"][:10]:
        share = row.get("impression_share_organic")
        click.echo(f"    {row['keyword_normalized'][:44]:<45} rank={row['organic_rank']}"
                   + (f"  share={share:.0%}" if share is not None else ""))
    if dry_run and r["rows"]:
        click.echo("\n  Re-run with --apply to write.")


@cli.command("rank-set")
@click.argument("keyword")
@click.argument("rank", type=int)
@click.option("--asin", default=None, help="Defaults to config default_asin")
@click.option("--as-of", default=None, help="YYYY-MM-DD (default today)")
@click.option("--page", default=None, type=int)
def rank_set_cmd(keyword, rank, asin, as_of, page):
    """Manually record our organic rank for a query.

    A manual row and a weekly SQP row can coexist; the freshest wins at read
    time. Use this when you have checked the SERP yourself.
    """
    from datetime import date as _date

    from src.amazon_ads.organic_rank import (
        load_config, normalize_keyword, upsert_ranks,
    )

    cfg = load_config()
    row = {
        "asin": asin or cfg["default_asin"] or "",
        "keyword_normalized": normalize_keyword(keyword),
        "keyword_raw": keyword,
        "organic_rank": rank,
        "page": page if page is not None else (1 if rank <= 48 else 2),
        "source": "manual",
        "as_of": as_of or _date.today().isoformat(),
    }
    n = upsert_ranks([row])
    click.echo(f"recorded: {row['asin']} \"{row['keyword_normalized']}\" "
               f"rank {rank} as_of {row['as_of']} ({n} row)")


@cli.command("ads-mark")
@click.option("--id", "rec_ids", multiple=True, help="Recommendation id(s)")
@click.option("--priority", default=None, help="Mark every open action at this priority (P0/P1/...)")
@click.option("--type", "rec_type", default=None, help="Mark every open action of this type")
@click.option("--status", type=click.Choice(["applied", "dismissed"]), default="applied")
@click.option("--dry-run/--apply", default=True, help="Dry run by default")
def ads_mark_cmd(rec_ids, priority, rec_type, status, dry_run):
    """Record which recommendations you actioned — this is what closes the loop.

    Until an action is marked, the outcome snapshot has no before/after to
    measure, so the system stays rules-based and never learns which advice
    worked. Marking a batch after applying it in Seller Central is the whole
    cost of making next week's brief evidence-weighted.
    """
    from src.amazon_ads.learning import mark_recommendation
    from src.db import fetch_all

    recs = [r for r in fetch_all("ads_recommendations")
            if str(r.get("status") or "open") == "open"]
    if rec_ids:
        chosen = [r for r in recs if str(r.get("id")) in set(rec_ids)]
    elif priority:
        chosen = [r for r in recs
                  if str(r.get("priority") or "").upper() == priority.upper()]
    elif rec_type:
        chosen = [r for r in recs
                  if str(r.get("type") or "").upper() == rec_type.upper()]
    else:
        raise click.UsageError("Pass --id, --priority or --type to select actions.")

    if not chosen:
        click.echo("No open recommendations match that selection.")
        return

    click.echo(f"{'DRY RUN — ' if dry_run else ''}marking {len(chosen)} "
               f"recommendation(s) as {status}:")
    for r in chosen[:25]:
        click.echo(f"  [{r.get('priority')}] {r.get('type'):<22} "
                   f"{str(r.get('entity_name'))[:44]}")
    if len(chosen) > 25:
        click.echo(f"  +{len(chosen) - 25} more")

    if dry_run:
        click.echo("\n  Re-run with --apply to record them.")
        return

    marked = linked = 0
    for r in chosen:
        result = mark_recommendation(r, status)
        if result.get("ok"):
            marked += 1
            if result.get("decisionLogged"):
                linked += 1

    click.echo(f"\n  Marked {marked} recommendation(s); {linked} mirrored onto "
               f"the decision log.")
    click.echo("  Outcomes are snapshotted at the 7/14/30-day horizons by the "
               "nightly `ads-outcomes` job — nothing further to do now.")
    click.echo("  Next week's `ppc-export` will include what actually happened.")


@cli.command("shopify-backfill")
@click.option("--since", default=None, help="YYYY-MM-DD; omit for full history")
@click.option("--with-email", is_flag=True,
              help="Also store the raw email (default: hash only)")
@click.option("--max-pages", default=None, type=int, help="Stop early (testing)")
def shopify_backfill_cmd(since, with_email, max_pages):
    """Store every Shopify order so AOV / LTV / repeat / cohorts become possible.

    Idempotent: each page is upserted on order_id as it arrives, so re-running
    is safe and a interrupted run resumes by simply being run again.
    """
    from datetime import date as _date

    from src.shopify_backfill import backfill

    since_d = _date.fromisoformat(since) if since else None
    click.echo(f"Shopify order backfill — {'since ' + since if since else 'FULL history'}")
    if with_email:
        click.echo("  --with-email: raw addresses WILL be stored. No metric here "
                   "needs them; the hash alone answers repeat/LTV/cohort.")
    r = backfill(since=since_d, with_email=with_email, max_pages=max_pages,
                 progress=click.echo)
    if r.get("error"):
        raise click.ClickException(r["error"])
    click.echo(f"\n  {r['written']:,} order(s) upserted over {r['pages']} page(s)")
    if r.get("first_date"):
        click.echo(f"  range: {r['first_date']} → {r['last_date']}")
    if r.get("skipped"):
        click.echo(f"  skipped (no id or no date): {r['skipped']}")
    click.echo("  Next: python -m src.main shopify-metrics")


@cli.command("shopify-metrics")
@click.option("--months", default=12, help="Cohort months to show")
@click.option("--since", default=None,
              help="YYYY-MM-DD; scope every figure to orders on/after this date")
def shopify_metrics_cmd(months, since):
    """AOV, LTV, repeat rate and cohorts. Shopify only — see the Amazon note."""
    from src import shopify_metrics as M
    from src.shopify_backfill import load_orders

    try:
        rows = load_orders()
    except Exception as e:
        if "shopify_orders" in str(e):
            raise click.ClickException(
                "Table shopify_orders is missing — run "
                "supabase/migration_shopify_orders.sql, then `shopify-backfill`.")
        raise
    if not rows:
        click.echo("No stored Shopify orders. Run `shopify-backfill` first.")
        return

    scope = ""
    if since:
        rows = [r for r in rows if str(r["order_date"]) >= since]
        scope = f"  (scoped to orders on/after {since})"
        if not rows:
            click.echo(f"No orders on/after {since}.")
            return

    s = M.summarise(rows)
    click.echo("Shopify customer metrics" + scope)
    click.echo(f"  history {s.first_order_date} → {s.last_order_date}   "
               f"{s.orders:,} orders · {s.customers:,} customers · "
               f"${s.revenue:,.0f} net revenue")
    click.echo("")
    click.echo(f"  \033[1m{s.interpretation()}\033[0m")
    click.echo("")

    click.echo("  ORDER VALUE                    mean      median")
    click.echo(f"    net merchandise        {s.aov:>10,.2f}  {s.aov_median:>10,.2f}   "
               f"(after discounts, before tax/shipping, less refunds)")
    if s.aov_paid is not None:
        click.echo(f"    total paid             {s.aov_paid:>10,.2f}  "
                   f"{s.aov_paid_median:>10,.2f}   "
                   f"(incl. tax + shipping — matches Shopify Admin AOV)")
    click.echo("")

    click.echo("  LIFETIME VALUE                 mean      median")
    click.echo(f"    all customers          {s.ltv_mean:>10,.2f}  {s.ltv_median:>10,.2f}   "
               f"median = typical customer ever (most buy once)")
    if s.ltv_mean_repeat is not None:
        click.echo(f"    repeaters (2+ orders)  {s.ltv_mean_repeat:>10,.2f}  "
                   f"{s.ltv_median_repeat:>10,.2f}   "
                   f"the customers worth acquiring more of")
    click.echo("")

    click.echo("  REPEAT BEHAVIOUR")
    click.echo(f"    repeat rate            {s.repeat_rate:>10.1%}   "
               f"({s.repeat_customers:,} of {s.customers:,})")
    if s.orders_per_repeater is not None:
        click.echo(f"    orders per repeater    {s.orders_per_repeater:>10.2f}")
        click.echo(f"    revenue from repeaters {s.revenue_from_repeaters_pct:>9.1f}%   "
                   f"(${s.revenue_from_repeaters:,.0f})")
    click.echo(f"    orders per customer    {s.orders_per_customer:>10.2f}")
    click.echo("")

    click.echo("  ORDER VALUE BY YEAR")
    click.echo(f"    {'year':<7}{'orders':>8}{'net AOV':>10}{'paid AOV':>10}"
               f"{'net median':>12}{'revenue':>12}")
    for y in M.by_year(rows):
        click.echo(f"    {y['year']:<7}{y['orders']:>8,}{y['aov']:>10,.2f}"
                   f"{y['aovPaid']:>10,.2f}{y['aovMedian']:>12,.2f}"
                   f"{y['revenue']:>12,.0f}")
    click.echo("")

    click.echo("  LTV BY ACQUISITION COHORT (revenue per customer within N days "
               "of their own first order)")
    click.echo(f"    {'cohort':<9}{'cust':>6}{'day 90':>10}{'day 365':>10}")
    for c in M.cohort_ltv_at_days(rows)[-months:]:
        d90 = f"{c['day90']:>10,.2f}" if c["day90"] is not None else f"{'—':>10}"
        d365 = f"{c['day365']:>10,.2f}" if c["day365"] is not None else f"{'—':>10}"
        click.echo(f"    {c['cohort']:<9}{c['customers']:>6}{d90}{d365}")
    click.echo("    '—' = that window has not fully elapsed for this cohort yet.")
    click.echo("")

    note = M.amazon_identity_note()
    click.echo("\n  Amazon: person-level metrics are NOT available.")
    for w in note["why"]:
        click.echo(f"    - {w}")
    click.echo(f"    {note['doNotDo']}")


@cli.command("health-ping")
@click.option("--send", is_flag=True, help="Actually deliver to Telegram")
@click.option("--dry-run", "dry", is_flag=True, default=False,
              help="Print the message and the send decision, deliver nothing")
def health_ping_cmd(send, dry):
    """Daily agent health check-in — one short message, or silence.

    Debug aid only; the scheduler runs this itself once a day. --dry-run never
    touches Telegram, so it works with TELEGRAM_* unset.
    """
    from src.health import run_health_ping

    if send and dry:
        raise click.UsageError("--send and --dry-run are mutually exclusive.")
    r = run_health_ping(send=send)

    click.echo(r["message"])
    click.echo("")
    click.echo(f"  severity:   {r['severity']}")
    click.echo(f"  would send: {r['would_send']} ({r['reason']})")
    if r["muted"]:
        click.echo("  MUTED by HEALTH_TELEGRAM=0 — nothing will be delivered.")
    if send:
        click.echo(f"  sent:       {r['sent']}"
                   + (f" — {r['error']}" if r.get("error") else ""))
    elif not dry:
        click.echo("  (nothing delivered; pass --send to deliver)")


@cli.command("ppc-export")
@click.option("--days", default=7, show_default=True, help="Closed days to cover")
@click.option("--out", "out_path", default=None, help="Write to a file")
@click.option("--brief-only", is_flag=True, help="Omit the AI instruction wrapper")
@click.option("--publish", is_flag=True,
              help="Also store it so the dashboard button works without Python")
@click.option("--emit-json", "emit_json", is_flag=True,
              help="Print a JSON envelope (metadata + markdown) instead of the brief")
def ppc_export_cmd(days, out_path, brief_only, publish, emit_json):
    """Full paste-ready PPC Command Brief, with a graded prior window.

    Everything the system knows, with provenance and explicit gaps, wrapped in
    hard rules that keep the receiving model inside the evidence.
    """
    from src.amazon_ads.export_brief import (build_brief, build_prompt, gather,
                                             grade_window, publish_brief)

    d = gather(days=days)
    brief = build_brief(d)
    prompt = build_prompt(brief)
    text = brief if brief_only else prompt
    grade = grade_window(d)

    if emit_json:
        # Structured envelope for the dashboard. The alternative was having the
        # API route regex the as-of date back out of the rendered markdown to
        # name the download file — parsing prose that this same command wrote
        # is a self-inflicted coupling, and it breaks the moment a heading is
        # reworded.
        import json as _json
        click.echo(_json.dumps({
            "as_of": d["as_of"], "window_start": d["start"], "days": d["days"],
            "prior_start": d.get("prior_start"), "prior_end": d.get("prior_end"),
            "score": round(grade.score, 1), "letter": grade.letter,
            "formula_version": grade.formula_version,
            "brief_md": brief, "prompt_md": prompt, "chars": len(prompt),
        }))
        if publish:
            res = publish_brief(d, brief, prompt)
            if not res.get("published"):
                click.echo(f"NOT published: {res.get('error')}", err=True)
        return

    if out_path:
        with open(out_path, "w") as f:
            f.write(text + "\n")
        click.echo(f"Wrote {len(text):,} chars to {out_path}")
    elif not publish:
        click.echo(text)

    if out_path or publish:
        click.echo(f"  grade: {grade.score:.0f}/100 ({grade.letter}) "
                   f"formula {grade.formula_version}"
                   + (f", dropped {', '.join(grade.dropped)}" if grade.dropped else ""))
        click.echo(f"  window: {d['start']} → {d['as_of']} vs "
                   f"{d.get('prior_start')} → {d.get('prior_end')}")
        click.echo(f"  {len(d['recs'])} recommendation(s), {len(d['gaps'])} known gap(s)")

    if publish:
        res = publish_brief(d, brief, prompt)
        if res.get("published"):
            click.echo(f"  published to ppc_briefs — the dashboard 'Copy full AI "
                       f"brief' button can now serve it without Python.")
        else:
            click.echo(f"  NOT published: {res.get('error')}", err=True)


@cli.command("ppc-playbook")
@click.option("--range", "rng", type=click.Choice(["7d", "14d", "30d"]), default="7d")
def ppc_playbook_cmd(rng):
    """What to do this week, in order — waste first, growth last.

    Reads live KPIs. Every figure is queried at call time; nothing is hardcoded.
    """
    import json as _json
    from collections import defaultdict

    from src.amazon_ads.playbook import WEEKLY_CADENCE, build_playbook
    from src.db import fetch_all, get_client
    from src.rules import amazon_as_of, window_start

    days = int(rng.rstrip("d"))
    asof = amazon_as_of()
    start = window_start(asof, days)
    client = get_client()

    try:
        from src.amazon_ads.strategy import account_target_acos
        target_acos, _basis = account_target_acos()
        target_acos = float(target_acos)
    except Exception:
        target_acos = 30.0

    # Queued recommendations (rank-gate fields live in evidence).
    recs = []
    for r in fetch_all("ads_recommendations"):
        ev = r.get("evidence")
        if isinstance(ev, str):
            try:
                ev = _json.loads(ev)
            except ValueError:
                ev = {}
        recs.append({**r, "evidence": ev or {}})

    # Placement rollup for the same closed-day window as the KPIs.
    pl_agg: dict[str, dict] = defaultdict(lambda: {"spend": 0.0, "sales": 0.0})
    try:
        offset = 0
        while True:
            page = (client.table("ads_placement_daily")
                    .select("date,campaign_id,placement,spend,sales_14d")
                    .gte("date", str(start)).lte("date", str(asof))
                    .order("date").order("campaign_id").order("placement")
                    .range(offset, offset + 999).execute().data) or []
            for row in page:
                e = pl_agg[str(row.get("placement") or "Unknown")]
                e["spend"] += float(row.get("spend") or 0)
                e["sales"] += float(row.get("sales_14d") or 0)
            if len(page) < 1000:
                break
            offset += 1000
    except Exception as e:
        click.echo(f"  (placement data unavailable: {str(e)[:90]})")
    placements = [{"placement": k, **v} for k, v in pl_agg.items()]

    # Role shares come from the same strategy config the dashboard reads.
    roles = []
    try:
        from src.amazon_ads.strategy import role_rollup

        camp_rows, offset = [], 0
        while True:
            page = (client.table("ads_campaigns_daily")
                    .select("date,campaign_id,campaign_name,spend,sales_14d,"
                            "orders_14d,clicks")
                    .gte("date", str(start)).lte("date", str(asof))
                    .order("date").order("campaign_id")
                    .range(offset, offset + 999).execute().data) or []
            camp_rows.extend(page)
            if len(page) < 1000:
                break
            offset += 1000
        roles = role_rollup(camp_rows)
    except Exception as e:
        log.debug("role rollup unavailable: %s", e)

    actions = build_playbook(target_acos, recs, placements, roles)

    click.echo(f"PPC playbook — {days} closed day(s) {start} → {asof}")
    click.echo(f"  break-even target ACOS: {target_acos:.1f}%")
    click.echo(f"  queued recommendations: {len(recs)}")
    if not actions:
        click.echo("\n  Nothing actionable right now.")
    for a in actions:
        click.echo("")
        click.echo(f"  [{a.priority}] {a.title}"
                   + (f"   (~{a.impact:,.0f} at stake)" if a.impact else ""))
        click.echo(f"       WHY : {a.why}")
        click.echo(f"       DO  : {a.do}")

    click.echo("\n  Weekly cadence:")
    for line in WEEKLY_CADENCE:
        click.echo(f"    • {line}")
    click.echo("\n  SQP note: multi-week ASIN-view SQP drives TREND and the rank "
               "gate. It is not Brand View / full category parity.")


@cli.command("ppc-plan")
@click.option("--only-needs-rank-check", is_flag=True,
              help="Show only actions held for a manual organic-rank check")
def ppc_plan_cmd(only_needs_rank_check):
    """PPC action plan with organic-rank gating applied to bid increases.

    Rank comes from Brand Analytics SQP or manual entry — never from the Ads
    API, which does not expose it. This is a cannibalization guard, not an
    incrementality measurement.
    """
    import json as _json

    from src.amazon_ads.organic_rank import load_config
    from src.db import fetch_all

    cfg = load_config()
    # Read the queue rather than regenerating it: generate_recommendations()
    # replaces ads_recommendations as a side effect, and a command that only
    # prints the plan must not rewrite it. Run `ads-actions` to refresh.
    raw = fetch_all("ads_recommendations")
    recs = []
    for r in raw:
        ev = r.get("evidence")
        if isinstance(ev, str):
            try:
                ev = _json.loads(ev)
            except ValueError:
                ev = {}
        recs.append({**r, "evidence": ev or {}})
    bids = [r for r in recs if (r.get("evidence") or {}).get("action_type") == "increase_bid"]

    summary = {"raises_full": 0, "raises_capped": 0, "holds_rank": 0,
               "needs_rank_check": 0, "rank_unknown_allowed": 0, "gating_disabled": 0}
    for r in bids:
        pol = (r.get("evidence") or {}).get("rank_policy_applied") or ""
        key = {"full_increase": "raises_full", "capped": "raises_capped",
               "hold": "holds_rank", "needs_rank_check": "needs_rank_check",
               "rank_unknown_allowed": "rank_unknown_allowed",
               "gating_disabled": "gating_disabled"}.get(pol)
        if key:
            summary[key] += 1

    click.echo("PPC plan — organic-rank gated (queued actions; "
               "run `ads-actions` to regenerate)")
    click.echo(f"  gating enabled   : {cfg['enabled']} "
               f"(high-bid threshold {_usd_or(cfg['high_bid_threshold'])}, "
               f"stale after {cfg['stale_after_days']}d)")
    click.echo(f"  total actions    : {len(recs)}")
    click.echo(f"  bid increases    : {len(bids)}")
    for k in ("raises_full", "raises_capped", "holds_rank", "needs_rank_check",
              "rank_unknown_allowed", "gating_disabled"):
        if summary[k]:
            click.echo(f"    {k:<22}: {summary[k]}")

    shown = [r for r in bids
             if not only_needs_rank_check
             or (r.get("evidence") or {}).get("needs_rank_check")]
    if shown:
        click.echo("")
        click.echo(f"    {'keyword':<40}{'cur':>7}{'plan':>8}{'final':>8}"
                   f"{'rank':>6}  {'risk':<8}policy")
        for r in sorted(shown, key=lambda x: -float(x.get("impact") or 0))[:40]:
            e = r.get("evidence") or {}
            rank = e.get("organic_rank")
            rank_s = str(rank) if rank is not None else ("brand" if e.get("rank_branded") else "?")
            click.echo(f"    {str(r.get('entity_name'))[:39]:<40}"
                       f"{float(e.get('cpc') or 0):>7.2f}"
                       f"{float(e.get('proposed_bid_before_rank_gate') or 0):>8.2f}"
                       f"{float(e.get('suggested_bid') or 0):>8.2f}"
                       f"{rank_s:>6}  {str(e.get('cannibalization_risk')):<8}"
                       f"{e.get('rank_policy_applied')}")

    click.echo("\n  Rank source: Brand Analytics SQP / manual. The Ads API does "
               "not provide organic rank.")
    click.echo("  This is a cannibalization guard, not proof of incrementality.")
    click.echo("  needs_rank_check actions are never auto-applied.")


def _usd_or(v) -> str:
    try:
        return f"${float(v):,.2f}"
    except (TypeError, ValueError):
        return str(v)


@cli.command("registration-plan")
@click.option("--csv", "csv_path", default=None, help="Write the full plan to CSV")
@click.option("--action", type=click.Choice(["all", "register_now", "review_contested",
              "monitor", "already_registered", "no_sales_tax"]), default="all")
def registration_plan_cmd(csv_path, action):
    """Ranked, auditable sales-tax registration plan from live data.

    Sales tax only. Entity and business-activity taxes (CA $800, WA B&O, OR CAT)
    never drive a registration recommendation here — see `entity-matrix`.
    """
    import csv as _csv

    from src.exports.registration_plan import (
        build_rows, counts_by_action, to_csv_rows,
    )

    rows = build_rows()
    counts = counts_by_action(rows)

    click.echo("Sales-tax registration plan — live data")
    for a in ("register_now", "review_contested", "monitor",
              "already_registered", "no_sales_tax"):
        click.echo(f"  {a:<20}: {counts.get(a, 0)}")
    click.echo(f"  {'TOTAL':<20}: {len(rows)}")

    residual = next((r.residual_risk for r in rows if r.residual_risk), "")
    if residual:
        click.echo(f"\n  ⚠ residual risk: {residual}")

    shown = [r for r in rows if action == "all" or r.decision.action == action]
    groups: dict[str, list] = {}
    for r in shown:
        groups.setdefault(r.decision.action, []).append(r)

    for a in ("register_now", "review_contested", "monitor",
              "already_registered", "no_sales_tax"):
        grp = groups.get(a)
        if not grp:
            continue
        click.echo(f"\n  {a.upper().replace('_', ' ')} ({len(grp)}):")
        if a in ("already_registered", "no_sales_tax"):
            click.echo("    " + ", ".join(r.facts.state_code for r in grp))
            continue
        for r in grp:
            f, d = r.facts, r.decision
            click.echo(f"    {f.state_code}  phys={d.physical_nexus:<9} "
                       f"econ={d.economic_nexus:<16} "
                       f"sales=${f.total_relevant_sales:>10,.0f}  [{d.confidence}]")
            click.echo(f"        {d.reason}")
            if r.entity_note:
                click.echo(f"        ({r.entity_note})")

    if csv_path:
        with open(csv_path, "w", newline="") as fh:
            _csv.writer(fh).writerows(to_csv_rows(rows))
        click.echo(f"\n  CSV written: {csv_path} ({len(rows)} rows)")

    click.echo("\n  Monitoring aid — not legal or tax advice. Confirm positions "
               "with a CPA before registering.")


@cli.command("entity-matrix")
@click.option("--status", type=click.Choice(["all", "verified_applies",
              "verified_none", "not_researched"]), default="all")
@click.option("--exposure-only", is_flag=True,
              help="Only obligations a remote seller can owe WITHOUT "
                   "foreign-qualifying (drops qualification-driven reports)")
def entity_matrix_cmd(status, exposure_only):
    """50 states + DC: non-sales-tax obligations for a remote ecommerce seller.

    `not_researched` means no verification has been done — it is NOT a finding
    that nothing is required.
    """
    from src.compliance.state_matrix import (
        counts, load_matrix, obligations, remote_seller_exposure,
        states_with_status,
    )

    m = load_matrix()
    c = counts(m)
    click.echo("Entity / business-activity matrix — 50 states + DC")
    click.echo(f"  jurisdictions   : {c['jurisdictions']}")
    for st in ("verified_applies", "verified_none", "not_researched"):
        click.echo(f"  {st:<16}: {c['by_status'][st]}")
    click.echo(f"  obligations     : {c['obligations']} "
               f"({c['by_mode'].get('schedulable', 0)} schedulable, "
               f"{c['by_mode'].get('review_only', 0)} review-only)")
    click.echo(f"  by trigger      : {c['by_trigger']}")

    obs = remote_seller_exposure(m) if exposure_only else obligations(m)
    if status in ("all", "verified_applies") and obs:
        label = ("Remote-seller exposure (no qualification needed)"
                 if exposure_only else "verified_applies")
        click.echo(f"\n  {label}:")
        click.echo(f"    {'st':<4}{'obligation':<44}{'trigger':<28}"
                   f"{'conf':<8}{'mode'}")
        for o in sorted(obs, key=lambda x: (x.state_code, x.name)):
            click.echo(f"    {o.state_code:<4}{o.name[:43]:<44}{o.trigger[:27]:<28}"
                       f"{o.confidence:<8}{o.mode}")

    if status in ("all", "not_researched"):
        nr = states_with_status("not_researched", m)
        click.echo(f"\n  not_researched ({len(nr)}) — NOT a finding of 'none':")
        click.echo("    " + ", ".join(nr))

    none_states = states_with_status("verified_none", m)
    click.echo(f"\n  verified_none ({len(none_states)}): "
               f"{', '.join(none_states) if none_states else 'none yet — no state has been affirmatively cleared'}")
    click.echo("\n  Monitoring aid — not legal or tax advice.")


@cli.command("entity-audit")
@click.option("--year", default=None, type=int, help="Limit to one tax year")
@click.option("--horizon", type=click.Choice(["12m", "24m", "all"]), default="12m",
              help="How far ahead to list (default 12m). Overdue is always shown.")
@click.option("--scope", type=click.Choice(["all", "registered", "home_foreign"]),
              default="all",
              help="all = every tracked state; registered = sales-tax registered "
                   "PLUS home and foreign-qualified; home_foreign = home and "
                   "foreign-qualified only.")
@click.option("--show-sources/--no-show-sources", default=False,
              help="Print the official source behind each obligation")
def entity_audit_cmd(year, horizon, scope, show_sources):
    """Entity, franchise and foreign-qualification obligations — not sales tax.

    Driven by config/entity_profile.json plus the sourced rules in
    config/seed_entity_obligations.json. Contested positions (e.g. whether FBA
    inventory makes the LLC 'doing business' in California) are listed for
    review and never given a due date until confirmed in the profile.

    Monitoring aid — not legal or tax advice.
    """
    from datetime import date as _date
    from src.compliance.entity_obligations import current_view

    today = _date.today()
    years = [year] if year else None
    v = current_view(today, years)
    profile = v["profile"]

    if not profile:
        raise click.ClickException(
            "No config/entity_profile.json — cannot tell which entity "
            "obligations apply.")

    # Horizon + scope use the same pure functions the dashboard does.
    from src.compliance.entity_filters import filter_view
    from src.db import fetch_all as _fa
    registered = {n["state_code"] for n in _fa("nexus_status")
                  if n.get("is_registered") is True}
    foreign = {str(e.get("state", "")).upper()
               for e in (profile.get("foreign_qualified") or [])}
    v = {**v, **filter_view(v, today, horizon=horizon, scope=scope,
                            registered=registered,
                            home_state=str(profile.get("home_state") or "") or None,
                            foreign_states=foreign)}

    fq = ", ".join(e.get("state", "?") for e in profile.get("foreign_qualified") or [])
    click.echo(f"Entity & compliance obligations — {today}")
    click.echo(f"  profile: {profile.get('entity_type', '?')} · home "
               f"{profile.get('home_state', '?')} · foreign-qualified "
               f"{fq or '(none)'}")
    click.echo(f"  filters: horizon {horizon} · scope {scope}")
    click.echo(f"  overdue {len(v['overdue'])} · upcoming {len(v['upcoming'])} · "
               f"needs profile data {len(v['undated'])} · "
               f"review-only {len(v['review'])} · settled {len(v['settled'])}")
    if v.get("hidden_by_horizon"):
        click.echo(f"  ({v['hidden_by_horizon']} further obligation(s) beyond the "
                   f"{horizon} horizon — use --horizon all to see them)")

    if v["overdue"]:
        click.echo("\n  OVERDUE (entity — NOT sales tax):")
        for o in v["overdue"]:
            click.echo(_entity_line(o))

    if v["upcoming"]:
        click.echo("\n  Upcoming:")
        for o in v["upcoming"]:
            click.echo(_entity_line(o))

    if v["undated"]:
        click.echo("\n  Applies, but no due date can be computed:")
        for o in v["undated"]:
            click.echo(_entity_line(o))
            click.echo(f"        → {o.due_note}")

    if v["review"]:
        click.echo("\n  Review with a CPA — applies only if the entity's facts "
                   "support it, so nothing is scheduled:")
        for r in v["review"]:
            amt = f" ~${r['amount_estimate']:,.0f}" if r.get("amount_estimate") else ""
            click.echo(f"    {r['state_code']} {r['form_code']} — {r['title']} "
                       f"[{r['confidence']}]{amt}")
            if r.get("confidence_note"):
                click.echo(f"        {r['confidence_note'][:150]}")
            click.echo(f"        → {r['reason']}")

    if v["settled"]:
        click.echo(f"\n  Settled: " + ", ".join(
            f"{o.state_code} {o.form_code} {o.period_label} ({o.status})"
            for o in v["settled"]))

    if show_sources:
        click.echo("\n  Sources:")
        seen = set()
        for o in v["overdue"] + v["upcoming"] + v["undated"]:
            k = (o.state_code, o.source_citation)
            if k in seen:
                continue
            seen.add(k)
            click.echo(f"    {o.state_code} {o.form_code}: {o.source_authority} — "
                       f"{o.source_citation}")
            click.echo(f"        {o.source_url}")
        for r in v["review"]:
            click.echo(f"    {r['state_code']} {r['form_code']}: "
                       f"{r['source_authority']} — {r['source_citation']}")
            click.echo(f"        {r['source_url']}")

    # Overlap with the sales-tax calendar. Reported, never auto-resolved.
    from src.compliance.entity_obligations import find_calendar_overlap
    from src.db import fetch_all as _fetch_all
    overlap = find_calendar_overlap(
        v["overdue"] + v["upcoming"] + v["undated"] + v["settled"],
        _fetch_all("filing_calendar"))
    if overlap:
        click.echo("\n  ⚠ Same period tracked in BOTH calendars — pick one:")
        for o in overlap:
            click.echo(f"    {o['state_code']} {o['period_label']}: "
                       f"filing_calendar annual row (due {o['calendar_due']}) "
                       f"vs entity {o['entity_form']} (due {o['entity_due']})")
        click.echo("      The entity row names the form and cites the source. To drop "
                   "the sales-tax duplicate:")
        for o in overlap:
            click.echo(f"      python -m src.main filing-mark {o['state_code']} "
                       f"{o['period_label']} --status not_required "
                       f"--notes \"tracked as {o['entity_form']} in entity calendar\"")

    click.echo("\n  Monitoring aid — not legal or tax advice.")


@cli.command("entity-calendar")
@click.option("--year", default=None, type=int, help="Tax year to sync (default: this year + next)")
@click.option("--dry-run/--apply", default=True, help="Dry run by default")
def entity_calendar_cmd(year, dry_run):
    """Write computed entity obligations to compliance_obligations.

    Recomputes due dates from the rules every time, but never reopens a period
    the user marked filed / not_required / dismissed.
    """
    from datetime import date as _date
    from src.compliance.entity_obligations import (
        build_obligations, load_profile, load_rules, prune_disabled,
        sync_obligations,
    )
    from src.db import fetch_all

    today = _date.today()
    years = [year] if year else [today.year, today.year + 1]
    registered = {n["state_code"] for n in fetch_all("nexus_status")
                  if n.get("is_registered") is True}
    scheduled, review = build_obligations(
        load_profile(), load_rules(), registered, years, today)

    r = sync_obligations(scheduled, dry_run=dry_run)
    if r.get("skipped"):
        raise click.ClickException(r["skipped"])

    # Reconcile the other direction: an obligation turned back off must stop
    # appearing as scheduled. Settled periods are never removed.
    pruned = prune_disabled(scheduled, dry_run=dry_run)

    verb = "would write" if r["dry_run"] else "wrote"
    click.echo(f"{'DRY RUN — ' if r['dry_run'] else ''}{verb} "
               f"{r['would_write'] if r['dry_run'] else r['written']} obligation row(s) "
               f"for {years}")
    click.echo(f"  settled periods preserved: {r['settled_preserved']}")
    if pruned["removed"]:
        click.echo(f"  {'would remove' if dry_run else 'removed'} "
                   f"{pruned['removed']} row(s) for obligations no longer enabled: "
                   f"{', '.join(pruned['rows'][:4])}")
    if pruned["kept_settled"]:
        click.echo(f"  kept {pruned['kept_settled']} settled row(s) for "
                   f"no-longer-tracked obligations (filing records are never deleted)")
    click.echo(f"  review-only (not written, needs confirmation): {len(review)}")
    for o in scheduled:
        due = o.due_date.isoformat() if o.due_date else "no date"
        click.echo(f"    {o.state_code} {o.obligation_type:<18} {o.form_code:<24} "
                   f"{o.period_label} due {due}")
    if r["dry_run"]:
        click.echo("\n  Re-run with --apply to write.")


@cli.command("entity-mark")
@click.argument("state_code")
@click.argument("obligation_type")
@click.argument("period_label")
@click.option("--status", type=click.Choice(["filed", "not_required", "dismissed", "open"]),
              default="filed")
@click.option("--notes", default=None, help="Why (kept on the row)")
@click.option("--filed-date", default=None,
              help="When it was actually filed (YYYY-MM-DD). Defaults to today.")
@click.option("--no-filed-date", "no_stamp", is_flag=True,
              help="Filed, but the exact date is unknown — record no filing date "
                   "rather than asserting today's.")
def entity_mark_cmd(state_code, obligation_type, period_label, status, notes,
                    filed_date, no_stamp):
    """Mark an entity obligation filed / not required / dismissed."""
    from src.compliance.entity_obligations import mark_obligation

    row = mark_obligation(state_code.upper(), obligation_type, period_label,
                          status, notes, filed_date=filed_date,
                          stamp_today=not no_stamp)
    if not row:
        raise click.ClickException(
            f"No obligation row for {state_code.upper()} {obligation_type} "
            f"{period_label} — run `entity-calendar --apply` first")
    click.echo(f"{state_code.upper()} {obligation_type} {period_label} → {status}")


@cli.command("filing-audit")
@click.option("--state", default=None, help="Limit to one state code")
@click.option("--show-excluded/--no-show-excluded", default=True,
              help="List rows that are not real obligations, with the reason")
def filing_audit_cmd(state, show_excluded):
    """Explain every filing_calendar row: live obligation, or excluded and why.

    Read-only. Sales-tax deadlines require the state to be registered; nexus
    alone produces register/review actions, never an overdue return.
    """
    from datetime import date as _date
    from src.calendar.filing_calendar import audit_filing_calendar

    r = audit_filing_calendar()
    today = _date.today()

    def keep(rows):
        return [x for x in rows if not state or x.get("state_code") == state.upper()]

    overdue, upcoming, excluded = keep(r["overdue"]), keep(r["upcoming"]), keep(r["excluded"])

    click.echo(f"filing_calendar audit — {today} "
               f"(sales-tax obligations for registered states only)")
    click.echo(f"  OVERDUE  : {len(overdue)}")
    click.echo(f"  upcoming : {len(upcoming)}")
    click.echo(f"  excluded : {len(excluded)}")

    if overdue:
        click.echo("\n  OVERDUE — registered, past due, not settled:")
        for x in overdue:
            click.echo(f"    {x['state_code']} {x['period_label']:<9} "
                       f"due {x['due_date']}  ({x['days_overdue']}d late)")

    if upcoming:
        click.echo("\n  Next 5 upcoming:")
        for x in upcoming[:5]:
            click.echo(f"    {x['state_code']} {x['period_label']:<9} "
                       f"due {x['due_date']}  (in {x['days_until_due']}d)")

    if show_excluded and excluded:
        from collections import Counter
        click.echo(f"\n  Excluded by reason: "
                   f"{dict(Counter(x['excluded_reason'] for x in excluded))}")
        for x in excluded:
            if x["excluded_reason"] == "settled":
                continue
            click.echo(f"    {x['state_code']} {x['period_label']:<9} "
                       f"{x['period_type']:<12} → {x['excluded_reason']}: "
                       f"{x['excluded_detail']}")


@cli.command("filing-mark-late")
@click.option("--dry-run/--apply", default=True,
              help="Dry run by default; --apply writes the changes")
def filing_mark_late_cmd(dry_run):
    """Flip pending filings to late when due_date is before today (AGENT_TZ).

    Does not touch filed or not_required. Sets reminder_sent on flipped
    rows; Telegram still goes through the existing daily summary.
    """
    from src.calendar.filing_calendar import mark_overdue_filings

    r = mark_overdue_filings(dry_run=dry_run)
    verb = "would flip" if r["dry_run"] else "flipped"
    click.echo(
        f"{'DRY RUN — ' if r['dry_run'] else ''}{verb} {r['flipped']} "
        f"pending→late (today {r['today']} AGENT_TZ)"
    )
    for c in r["changes"]:
        click.echo(
            f"  {c['state_code']} {c['period_label']:<9} "
            f"due {c['due_date']}  {c['from_status']} → late"
        )
    click.echo(
        f"  already late: {r['already_late']}, "
        f"past-due settled left alone: {r['skipped_settled']}"
    )
    if r["dry_run"] and r["flipped"]:
        click.echo("\n  Re-run with --apply to write these changes.")


@cli.command("filing-cleanup")
@click.option("--dry-run/--apply", default=True,
              help="Dry run by default; --apply writes the changes")
def filing_cleanup_cmd(dry_run):
    """Settle calendar rows that are not real obligations.

    Sets them to `not_required` with the reason in filed_notes. Nothing is
    deleted — the row is the evidence for why a period was dismissed, and the
    rebuild preserves settled periods so they stay settled.
    """
    from src.calendar.filing_calendar import cleanup_filing_calendar

    r = cleanup_filing_calendar(dry_run=dry_run)
    verb = "would settle" if r["dry_run"] else "settled"
    click.echo(f"{'DRY RUN — ' if r['dry_run'] else ''}{verb} {r['changed']} row(s)")
    for c in r["changes"]:
        click.echo(f"  {c['state_code']} {c['period_label']:<9} {c['period_type']:<12} "
                   f"due {c['due_date']}  {c['from_status']} → not_required  "
                   f"({c['reason']}: {c['detail']})")
    click.echo(f"  remaining: {r['still_overdue']} overdue, "
               f"{r['still_upcoming']} upcoming")
    if r["dry_run"] and r["changed"]:
        click.echo("\n  Re-run with --apply to write these changes.")


@cli.command("filing-mark")
@click.argument("state_code")
@click.argument("period_label")
@click.option("--status", type=click.Choice(["filed", "not_required"]),
              default="filed", help="What to record for this period")
@click.option("--amount", default=None, type=float, help="Filed amount")
@click.option("--zero-return", is_flag=True, help="Filed as a zero return")
@click.option("--notes", default=None, help="Reason or reference")
def filing_mark_cmd(state_code, period_label, status, amount, zero_return, notes):
    """Mark one period filed, or not required / dismissed as a false positive.

    A settled period is preserved by the nightly rebuild, so it will not come
    back as an overdue chip.
    """
    from src.calendar.filing_calendar import (
        mark_filing_complete, mark_filing_not_required,
    )

    sc = state_code.upper()
    if status == "filed":
        r = mark_filing_complete(sc, period_label, amount=amount,
                                 notes=notes, is_zero_return=zero_return)
    else:
        r = mark_filing_not_required(sc, period_label, reason=notes)

    if not r:
        raise click.ClickException(
            f"No filing_calendar row for {sc} {period_label}")
    click.echo(f"{sc} {period_label} → {status}")


@cli.command("pnl-sync")
@click.option("--days", default=30, help="Days to compute")
@click.option("--no-skus", is_flag=True,
              help="Account grain only — skips the orders report (COGS falls back to estimates)")
def pnl_sync_cmd(days, no_skus):
    """Compute daily contribution: sales - referral - FBA - ad spend - COGS.

    Use a large --days to backfill history, e.g. `pnl-sync --days 365`.
    """
    from src.pnl import compute_pnl

    click.echo(f"Computing P&L for last {days} days...")
    result = compute_pnl(days=days, with_skus=not no_skus)
    if not result.get("rows"):
        click.echo("No data in that window.")
        return
    click.echo(f"Days: {result['days']}, Rows: {result['rows']} account "
               f"+ {result.get('sku_rows', 0)} SKU, Inserted: {result['inserted']}")
    click.echo(f"  Sales        ${result['total_sales']:>12,.2f}")
    click.echo(f"  - Fees       ${result['total_fees']:>12,.2f}")
    click.echo(f"  - Ad spend   ${result['total_ads']:>12,.2f}")
    click.echo(f"  - COGS       ${result['total_cogs']:>12,.2f}")
    click.echo(f"  = Contribution ${result['total_contribution']:>10,.2f}")
    click.echo(f"Fee basis: {result['settled_days']} settled day(s), "
               f"{result['estimated_days']} estimated "
               f"(referral {result['referral_pct']*100:.0f}%, FBA ${result['fba_per_unit']:.2f}/unit)")
    click.echo(f"COGS: {'sku_costs x daily units' if result['has_cogs'] else 'not configured (set sku_costs table)'}")
    if result.get("missing_cost_skus"):
        click.echo(f"  ⚠ no sku_costs entry for {len(result['missing_cost_skus'])} SKU(s), "
                   f"used average unit cost: {', '.join(result['missing_cost_skus'][:5])}")
    if result.get("excluded_zero_revenue_units"):
        click.echo(f"  ⚠ excluded {result['excluded_zero_revenue_units']} unit(s) on order lines "
                   f"with no item-price (kept units on the same basis as gross_sales)")
    flagged = result.get("flagged_days") or []
    if flagged:
        click.echo(click.style(f"  ⚠ {len(flagged)} day(s) failed the sanity check "
                               f"— stored, but the inputs disagree:", fg="yellow"))
        for f in flagged[:10]:
            click.echo(f"      {f['date']}  sales ${f['sales']:,.2f}  units {f['units']}  "
                       f"contribution ${f['contribution']:,.2f}  — {'; '.join(f['issues'])}")
    else:
        click.echo("  Sanity check: all days within expected fee % and revenue/unit bands")
    skipped = result.get("skipped_days") or []
    if skipped:
        click.echo(click.style(
            f"  ⚠ skipped {len(skipped)} day(s) with units but $0 sales "
            f"(not stored): {', '.join(skipped[:12])}",
            fg="yellow",
        ))


@cli.command("pnl-monthly-sync")
def pnl_monthly_sync_cmd():
    """Store monthly Amazon contribution from sales_by_sku × sku_costs.

    No SP-API call. Covers every Amazon month already in sales_by_sku
    (today: 2024-08 through the current month). Months before ads
    coverage are labelled ads-unknown, not $0 spend.
    """
    from src.pnl_monthly import compute_monthly_pnl

    click.echo("Computing monthly Amazon SKU economics from sales_by_sku...")
    result = compute_monthly_pnl(persist=True)
    if not result.get("month_count"):
        click.echo("No Amazon sales_by_sku rows.")
        return
    click.echo(f"Months: {result['coverage_min']} → {result['coverage_max']} "
               f"({result['month_count']} months, {result.get('sku_row_count', 0)} SKU rows)")
    click.echo(f"  Sales        ${result['total_sales']:>12,.2f}")
    click.echo(f"  - Fees       ${result['total_fees']:>12,.2f}")
    click.echo(f"  - Ad spend   ${result['total_ads']:>12,.2f}  "
               f"({result['ads_known_months']} month(s) with ads, "
               f"{result['ads_unknown_months']} ads-unknown)")
    click.echo(f"  - COGS       ${result['total_cogs']:>12,.2f}")
    click.echo(f"  = Contribution ${result['total_contribution']:>10,.2f}")
    click.echo(f"Inserted: {result.get('inserted', 0)} month + "
               f"{result.get('sku_inserted', 0)} SKU rows")
    if result.get("missing_cost_skus"):
        click.echo(f"  ⚠ no sku_costs for {len(result['missing_cost_skus'])} SKU(s): "
                   f"{', '.join(result['missing_cost_skus'][:8])}")


@cli.command("paid-ads-freshness")
@click.option("--send/--dry-run", default=False,
              help="Actually send the Telegram nudge (default: print only).")
def paid_ads_freshness_cmd(send):
    """Report how stale each Shopify paid-ads CSV export is."""
    from src.alerts.paid_ads_freshness import (
        STALE_AFTER_DAYS, build_message, check_paid_ads_freshness,
        run_paid_ads_freshness_check,
    )

    if send:
        run_paid_ads_freshness_check()
        return

    result = check_paid_ads_freshness()
    click.echo(f"Today: {result['today']}  (stale after {STALE_AFTER_DAYS} days)")
    for s in result["sources"]:
        if s["missing"]:
            state = "never uploaded"
        elif s["stale"]:
            state = f"STALE {s['days_behind']}d"
        else:
            state = f"{s['days_behind']}d old"
        click.echo(f"  {s['label']:16} {str(s['max_date']):12} {state}")
    message = build_message(result)
    if message:
        click.echo("\nWould send:\n" + message)
    else:
        click.echo("\nNothing to send — every source is current.")


@cli.command("pulse-audit")
@click.option("--date", "target_date", required=True, help="PST date to audit YYYY-MM-DD")
def pulse_audit_cmd(target_date):
    """Audit daily sales for a specific PST date — compare all-statuses vs shipped-only."""
    from src.amazon_sp.client import request_and_download
    from src.amazon_sp.reports import _detect_delimiter, _build_header_lookup, _get
    from src.sales_daily import _to_tz_date, LA
    from datetime import date, timedelta
    from collections import defaultdict, Counter
    import csv, io

    d = date.fromisoformat(target_date)
    click.echo(f"Auditing Amazon sales for {target_date} (PST/PDT)...")

    content = request_and_download(
        "GET_FLAT_FILE_ALL_ORDERS_DATA_BY_ORDER_DATE_GENERAL",
        d - timedelta(days=1), d + timedelta(days=1))

    delimiter = _detect_delimiter(content.split("\n", 1)[0])
    reader = csv.DictReader(io.StringIO(content), delimiter=delimiter, quotechar='"')
    H = _build_header_lookup(reader.fieldnames)

    by_status: dict[str, dict] = defaultdict(lambda: {"count": 0, "price": 0.0})
    for row in reader:
        status = _get(row, H, "order-status").lower()
        pd_str = _get(row, H, "purchase-date")
        sale_date = _to_tz_date(pd_str, LA)
        if not sale_date or sale_date.isoformat() != target_date:
            continue
        price_str = _get(row, H, "item-price")
        try:
            price = float(price_str.replace(",", "")) if price_str else 0
        except (ValueError, TypeError):
            price = 0
        by_status[status]["count"] += 1
        by_status[status]["price"] += price

    total_all = sum(s["price"] for s in by_status.values())
    total_shipped = by_status.get("shipped", {}).get("price", 0)
    total_items = sum(s["count"] for s in by_status.values())

    click.echo(f"{'='*55}")
    click.echo(f"  Date: {target_date} (America/Los_Angeles)")
    click.echo(f"  Source: SP-API orders report (item-price)")
    click.echo(f"{'='*55}")
    for st, data in sorted(by_status.items()):
        click.echo(f"  {st:>20}: {data['count']:>4} items  ${data['price']:>10,.2f}")
    click.echo(f"  {'─'*50}")
    click.echo(f"  {'ALL STATUSES':>20}: {total_items:>4} items  ${total_all:>10,.2f}")
    click.echo(f"  {'SHIPPED ONLY':>20}:               ${total_shipped:>10,.2f}")
    click.echo(f"{'='*55}")
    click.echo(f"  Pulse should show: ${total_all:,.2f} (all statuses)")
    click.echo(f"  Gap if shipped-only: ${total_all - total_shipped:,.2f}")


@cli.command("spapi-probe")
def spapi_probe_cmd():
    """Probe which SP-API report types are authorized with current credentials."""
    from datetime import date as d, timedelta
    from src.amazon_sp.client import create_report

    end = d.today() - timedelta(days=1)
    start = end - timedelta(days=7)

    reports = [
        ("GET_FLAT_FILE_ALL_ORDERS_DATA_BY_ORDER_DATE_GENERAL", "All Orders"),
        ("GET_LEDGER_DETAIL_VIEW_DATA", "Inventory Ledger"),
        ("GET_LEDGER_SUMMARY_VIEW_DATA", "Ledger Summary (tax)"),
        ("GET_FBA_FULFILLMENT_CUSTOMER_RETURNS_DATA", "FBA Customer Returns"),
        ("GET_FBA_STORAGE_FEE_CHARGES_DATA", "FBA Storage Fees"),
        ("GET_FBA_REIMBURSEMENTS_DATA", "FBA Reimbursements"),
        ("GET_FBA_ESTIMATED_FBA_FEES_TXT_DATA", "FBA Fee Preview"),
        ("GET_SALES_AND_TRAFFIC_REPORT", "Sales & Traffic (BA)"),
        ("GET_BRAND_ANALYTICS_REPEAT_PURCHASE_REPORT", "Repeat Purchase (BA)"),
        ("GET_FBA_FULFILLMENT_CUSTOMER_SHIPMENT_SALES_DATA", "FBA Shipment Sales"),
        ("GET_FBA_SNS_FORECAST_DATA", "Subscribe & Save Forecast"),
        ("GET_FBA_SNS_PERFORMANCE_DATA", "Subscribe & Save Performance"),
    ]

    click.echo(f"Probing {len(reports)} report types ({start} to {end})")
    click.echo(f"{'Report':<40} {'Status'}")
    click.echo("-" * 55)
    for rt, label in reports:
        try:
            create_report(rt, start, end)
            click.echo(f"  {label:<40} OK")
        except Exception as e:
            err = str(e)
            status = "403 (not authorized)" if "403" in err or "Forbidden" in err else f"ERROR: {err[:40]}"
            click.echo(f"  {label:<40} {status}")


@cli.command("spapi-returns")
@click.option("--days", default=30, help="Days back to fetch")
@click.option("--dry-run", is_flag=True)
def spapi_returns_cmd(days, dry_run):
    """Fetch FBA customer returns via SP-API."""
    from datetime import date as d, timedelta
    from src.amazon_sp.reports import fetch_fba_returns

    end = d.today() - timedelta(days=1)
    start = end - timedelta(days=days)
    if dry_run:
        click.echo("DRY RUN\n")

    click.echo(f"Fetching FBA returns: {start} to {end}")

    def _on_poll(status, elapsed):
        click.echo(f"  [{elapsed}s] {status}")

    result = fetch_fba_returns(start, end, dry_run=dry_run, on_poll=_on_poll)
    click.echo(f"Returns: {result['rows_parsed']} parsed, {result['rows_inserted']} inserted")
    click.echo(f"SKUs: {', '.join(result['skus_found'][:10])}")
    if result["reasons"]:
        click.echo("Reasons:")
        for reason, count in sorted(result["reasons"].items(), key=lambda x: -x[1]):
            click.echo(f"  {reason}: {count}")


@cli.command("forecast-sku")
@click.option("--sku", required=True, help="SKU code")
@click.option("--end", "end_date", required=True, help="End date YYYY-MM-DD")
@click.option("--start", "start_date", default=None, help="Start date (default: today)")
@click.option("--safety", default=0.15, help="Safety stock %% (default 0.15)")
def forecast_sku_cmd(sku, end_date, start_date, safety):
    """Forecast demand and coverage for a single SKU."""
    from src.forecast.sku_demand import forecast_sku

    result = forecast_sku(sku, end_date, start_date, safety)
    if result.get("error"):
        click.echo(f"Error: {result['error']}")
        return

    click.echo(f"{'='*60}")
    click.echo(f"  SKU DEMAND FORECAST — {result['product_name']}")
    click.echo(f"  {result['start_date']} → {result['end_date']} ({result['num_weeks']} weeks)")
    click.echo(f"{'='*60}")

    click.echo(f"\n  Expected:  {result['expected_units']:>8,} units")
    click.echo(f"  Coverage:  {result['coverage_units']:>8,} units ({safety*100:.0f}% safety)")
    click.echo(f"  Low/High:  {result['low_band']:>8,} / {result['high_band']:>8,}")

    m = result["methods"]
    click.echo(f"\n  TRIPLE-CHECK:")
    click.echo(f"    A) Naive run-rate:     {m['A_naive_runrate']:>8,}")
    click.echo(f"    B) Seasonal+Holiday:   {m['B_seasonal_yoy']:>8,}")
    click.echo(f"    C) SnS+organic:        {m['C_sns_plus_organic']:>8,}")
    click.echo(f"    Spread: {m['spread_pct']}%{' ⚠️ WIDE' if m['spread_warning'] else ''}")

    b = result["breakdown"]
    click.echo(f"\n  BREAKDOWN:")
    click.echo(f"    Velocity: {b['blended_daily_velocity']} u/day")
    click.echo(f"    SnS: {b['sns_active_subs']} subs, {b['sns_weekly_shipped']} shipped/wk")
    click.echo(f"    Return rate: {b['return_rate_pct']}%")

    for h in result.get("holidays", []):
        click.echo(f"    {h}")

    click.echo(f"\n  Model: {result.get('model_version', 'default')}")
    click.echo(f"  Offpeak weights: {result.get('offpeak_weights', result.get('weights', {}))}")
    if result.get("has_peak_weeks"):
        click.echo(f"  Peak weights:    {result.get('peak_weights', {})}")
        click.echo(f"  Peak protection: {'ACTIVE' if result.get('peak_protection') else 'no'}")
        click.echo(f"  Effective safety: {result.get('effective_safety_pct', 0)*100:.0f}%")
        # Show how expected compares to pure method B
        m = result["methods"]
        click.echo(f"  Expected vs pure B: {result['expected_units']:,} vs {m['B_seasonal_yoy']:,} "
                   f"({result['expected_units']/max(m['B_seasonal_yoy'],1)*100:.0f}%)")
    click.echo(f"\n  {result['disclaimer']}")
    click.echo()


@cli.command("forecast-backfill")
@click.option("--start", "start_date", required=True, help="Start date YYYY-MM-DD")
@click.option("--end", "end_date", required=True, help="End date YYYY-MM-DD")
@click.option("--min-velocity", default=0.5, help="Min V30 to include SKU")
def forecast_backfill_cmd(start_date, end_date, min_velocity):
    """Backfill forecasts for all active physical/inventory SKUs."""
    from src.forecast.sku_demand import forecast_sku
    from src.forecast.inventory_filter import filter_inventory_skus
    from src.db import fetch_all

    vel_rows = fetch_all("sku_velocity")
    all_active = sorted(set(
        v["sku"] for v in vel_rows
        if v.get("sku") and float(v.get("total_u_30", 0) or 0) > min_velocity
    ))

    active_skus, skipped = filter_inventory_skus(all_active)
    active_skus.sort()

    click.echo(f"Backfilling {len(active_skus)} physical SKUs: {start_date} → {end_date}")
    if skipped:
        click.echo(f"Skipped {len(skipped)} non-inventory: {', '.join(skipped)}")
    click.echo(f"{'='*60}")

    ok = 0
    fail = 0
    for sku in active_skus:
        try:
            result = forecast_sku(sku, end_date, start_date)
            if result.get("error"):
                click.echo(f"  {sku:<18} SKIP  {result['error']}")
                fail += 1
            else:
                click.echo(f"  {sku:<18} OK    {result['num_weeks']}wk  expected={result['expected_units']:,}")
                ok += 1
        except Exception as e:
            click.echo(f"  {sku:<18} FAIL  {str(e)[:60]}")
            fail += 1

    click.echo(f"\nDone: {ok} OK, {fail} failed/skipped out of {len(active_skus)}")


@cli.command("forecast-reconcile")
@click.option("--sku", default=None, help="Single SKU or all")
def forecast_reconcile_cmd(sku):
    """Reconcile forecasts: ingest actuals → score → calibrate."""
    from src.forecast.reconcile import run_full_reconcile

    result = run_full_reconcile(sku)
    a = result["actuals"]
    r = result["reconciliation"]
    c = result["calibration"]

    click.echo(f"Actuals: {a.get('rows_upserted', 0)} rows, {a.get('skus', 0)} SKUs "
               f"(orders: {a.get('orders_weeks', 0)} weeks, proxy: {a.get('proxy_weeks', 0)} weeks, "
               f"{a.get('skus_with_orders', 0)} SKUs with real order data)")
    click.echo(f"Reconciliation: {r.get('runs_found', 0)} runs, "
               f"{r.get('weeks_scored', 0)} weeks scored, "
               f"{r.get('skus_scored', 0)} SKUs"
               + (f" ({r.get('skus_non_inventory_skipped', 0)} non-inventory skipped)" if r.get('skus_non_inventory_skipped') else ""))
    if r.get("message"):
        click.echo(f"  {r['message']}")

    # Show per-SKU season detail
    sw = r.get("season_weights", {})
    if sw:
        for s, info in sorted(sw.items()):
            op_n = info.get("offpeak_weeks", 0)
            pk_n = info.get("peak_weeks", 0)
            status = "calibrated" if (op_n >= 8 or pk_n >= 8) else "insufficient"
            click.echo(f"  {s:<18} offpeak={op_n}wk peak={pk_n}wk → {status}")

    click.echo(f"Calibration: {c.get('calibrated', 0)} SKUs calibrated")
    if c.get("message"):
        click.echo(f"  {c['message']}")
    for ch in c.get("changes", []):
        click.echo(f"  {ch}")


@cli.command("spapi-sns")
@click.option("--weeks", default=None, type=int, help="Weeks of history (default 13)")
@click.option("--start", "start_date", default=None, help="Start date YYYY-MM-DD (overrides --weeks)")
@click.option("--dry-run", is_flag=True)
def spapi_sns_cmd(weeks, start_date, dry_run):
    """Fetch Subscribe & Save metrics via Replenishment API."""
    from src.amazon_sp.replenishment import fetch_seller_metrics, fetch_offer_metrics

    if dry_run:
        click.echo("DRY RUN\n")

    label = f"from {start_date}" if start_date else f"{weeks or 13} weeks"
    click.echo(f"Fetching seller-level SnS metrics ({label})...")
    try:
        seller = fetch_seller_metrics(weeks=weeks, start_date=start_date, dry_run=dry_run)
        click.echo(f"  Weeks: {seller['weeks_fetched']} ({seller['complete_weeks']} complete)")
        click.echo(f"  Latest complete: wk {seller['latest_week']} — "
                   f"{seller['latest_subs']:,} subs, {seller['latest_shipped']} shipped, "
                   f"${seller['latest_revenue']:,.2f} rev")
        click.echo(f"  Inserted: {seller['rows_inserted']}")
    except PermissionError as e:
        click.echo(f"  {e}")
        return
    except Exception as e:
        click.echo(f"  Error: {e}")

    click.echo("Fetching offer-level SnS metrics (latest week)...")
    try:
        offers = fetch_offer_metrics(dry_run=dry_run)
        click.echo(f"  Offers: {offers['offers_fetched']}, Week: {offers['week']}, Inserted: {offers['rows_inserted']}")
    except Exception as e:
        click.echo(f"  Error: {e}")


@cli.command("spapi-traffic")
@click.option("--days", default=7, help="Days back to fetch")
@click.option("--dry-run", is_flag=True)
def spapi_traffic_cmd(days, dry_run):
    """Fetch Amazon Sales & Traffic report via SP-API."""
    from datetime import date as d, timedelta
    from src.amazon_sp.reports import fetch_sales_traffic

    end = d.today() - timedelta(days=1)
    start = end - timedelta(days=days)
    if dry_run:
        click.echo("DRY RUN\n")

    click.echo(f"Fetching Sales & Traffic: {start} to {end}")
    result = fetch_sales_traffic(start, end, dry_run=dry_run)
    click.echo(f"Days: {result['days']}, ASINs: {result['asins']}, Inserted: {result['rows_inserted']}")


@cli.command("spapi-reimbursements")
@click.option("--days", default=90, help="Days back to fetch (chunked ≤30)")
@click.option("--dry-run", is_flag=True)
def spapi_reimbursements_cmd(days, dry_run):
    """Fetch FBA reimbursements via SP-API."""
    from datetime import timedelta
    from src.amazon_sp.reports import fetch_reimbursements
    from src.rules import amazon_as_of

    end = amazon_as_of()
    start = end - timedelta(days=days)
    if dry_run:
        click.echo("DRY RUN\n")

    click.echo(f"Fetching reimbursements: {start} to {end}")
    result = fetch_reimbursements(start, end, dry_run=dry_run)
    click.echo(f"Parsed: {result['rows_parsed']}, Total: ${result['total_amount']:,.2f}, Inserted: {result['rows_inserted']}")


@cli.command("spapi-refresh")
@click.option("--days", default=30, help="Number of days back to fetch (default 30)")
@click.option("--dry-run", is_flag=True)
def spapi_refresh(days, dry_run):
    """Refresh both orders and inventory from SP-API (last N days)."""
    from datetime import date as d, timedelta
    end = d.today() - timedelta(days=1)
    start = end - timedelta(days=days)

    if dry_run:
        click.echo("DRY RUN — no data will be written.\n")

    click.echo(f"SP-API refresh: {start} to {end}\n")

    def _on_poll(status, elapsed):
        click.echo(f"  [{elapsed}s] {status}")

    from src.amazon_sp.reports import fetch_orders, fetch_inventory

    click.echo("--- Orders ---")
    try:
        orders = fetch_orders(start, end, dry_run=dry_run, on_poll=_on_poll)
        _print_spapi_result("Orders", orders)
    except Exception as e:
        click.echo(f"  Orders failed: {e}")

    click.echo("\n--- Inventory Ledger ---")
    try:
        inv = fetch_inventory(start, end, dry_run=dry_run, on_poll=_on_poll)
        _print_spapi_result("Inventory", inv)
    except Exception as e:
        click.echo(f"  Inventory failed: {e}")

    click.echo("\nDone.")


def _print_spapi_result(label: str, result: dict):
    click.echo(f"\n  {label} Results:")
    if result.get("chunks", 0) > 1:
        click.echo(f"  Chunks:        {result['chunks']}")
    click.echo(f"  Rows total:    {result.get('rows_total', 0):,}")
    click.echo(f"  Rows parsed:   {result.get('rows_parsed', 0):,}")
    click.echo(f"  Rows inserted: {result.get('rows_inserted', 0):,}")
    if result.get("unique_orders"):
        click.echo(f"  Unique orders: {result['unique_orders']:,}")
    if result.get("sales_periods"):
        click.echo(f"  Sales periods: {result['sales_periods']}")
    if result.get("total_gross_sales"):
        click.echo(f"  Gross sales:   ${result['total_gross_sales']:,.2f}")
    if result.get("total_tax"):
        click.echo(f"  Tax:           ${result['total_tax']:,.2f}")
    if result.get("ship_to_states"):
        click.echo(f"  States (to):   {result['ship_to_states']}")
    if result.get("states_found"):
        click.echo(f"  States found:  {result['states_found']}")
    if result.get("unknown_fcs"):
        click.echo(f"  Unknown FCs:   {result['unknown_fcs']}")
    if result.get("total_cogs") is not None and result.get("report_type") == "ledger_summary":
        click.echo(f"  COGS value:    ${result['total_cogs']:,.2f}")
        click.echo(f"  Units:         {result.get('total_units', 0):,}")
    if result.get("missing_cost_skus"):
        click.echo(f"  Missing cost:  {len(result['missing_cost_skus'])} SKUs, "
                   f"{result.get('missing_cost_units', 0)} units excluded from $")
    for w in result.get("warnings", []):
        click.echo(f"  Warning: {w}")

    # Show sample records during dry-run
    samples = result.get("_samples", [])
    if samples:
        click.echo(f"\n  Sample records ({len(samples)}):")
        for s in samples:
            if hasattr(s, "state_code"):
                click.echo(
                    f"    {s.state_code} {s.period_start}..{s.period_end}: "
                    f"${s.gross_sales:,.2f} sales, ${s.tax_collected:,.2f} tax, "
                    f"{s.order_count} orders"
                )
            elif isinstance(s, dict):
                click.echo(
                    f"    {s.get('state_code', '?')} "
                    f"{s.get('period_start', '?')}..{s.get('period_end', '?')}: "
                    f"${s.get('gross_sales', 0):,.2f} sales"
                )


@cli.command("playbook")
@click.argument("state_code", required=False, default=None)
@click.option("--export", "export_path", default=None,
              help="Write playbook to a file (e.g., ./exports/CA_playbook.md)")
@click.option("--unregistered", is_flag=True,
              help="Generate playbooks for ALL unregistered states")
@click.option("--nexus-only", is_flag=True,
              help="With --unregistered: only states with nexus/franchise flags")
def playbook_cmd(state_code, export_path, unregistered, nexus_only):
    """Generate compliance playbooks. Single state or batch unregistered."""
    from src.compliance.playbook import build_playbook
    from pathlib import Path

    if unregistered:
        from src.db import fetch_all
        nexus_rows = fetch_all("nexus_status")
        flags = {f["state_code"] for f in fetch_all("franchise_tax_flags")
                 if f.get("status") == "open"}

        targets = []
        for n in nexus_rows:
            if n.get("is_registered"):
                continue
            sc = n["state_code"]
            has_nexus = n.get("has_physical_nexus") or n.get("has_economic_nexus")
            has_flag = sc in flags
            if nexus_only and not (has_nexus or has_flag):
                continue
            targets.append(sc)

        out_dir = Path(export_path) if export_path else Path("exports/playbooks")
        out_dir.mkdir(parents=True, exist_ok=True)

        for sc in sorted(targets):
            md = build_playbook(sc)
            p = out_dir / f"{sc}_playbook.md"
            p.write_text(md, encoding="utf-8")
            click.echo(f"  {sc} → {p}")

        click.echo(f"\nExported {len(targets)} playbooks to {out_dir}/")
        return

    if not state_code:
        click.echo("Usage: playbook STATE or playbook --unregistered")
        return

    state_code = state_code.upper()
    md = build_playbook(state_code)

    if export_path:
        p = Path(export_path)
        p.parent.mkdir(parents=True, exist_ok=True)
        p.write_text(md, encoding="utf-8")
        click.echo(f"Playbook exported to {p}")
    else:
        click.echo(md)


@cli.command("inventory-presence-export")
@click.option("--format", "fmt", type=click.Choice(["md", "csv", "pdf", "all"]),
              default="all", help="Output format")
@click.option("--out", "out_dir", default="exports/cpa",
              help="Output directory")
@click.option("--state", default=None, help="Single state filter (e.g., CA)")
@click.option("--validate-only", is_flag=True,
              help="Run validation checks only, no export")
@click.option("--no-upload", is_flag=True,
              help="Skip uploading to Supabase Storage")
def inventory_presence_export(fmt, out_dir, state, validate_only, no_upload):
    """CPA Export: FBA Inventory Presence by State (MD + CSV + PDF)."""
    from pathlib import Path
    from datetime import datetime, timezone
    from src.exports.inventory_presence import (
        build_markdown, build_csv, build_pdf, build_metadata,
        upload_exports, _gather_state_evidence, _run_validation,
    )

    if validate_only:
        evidence = _gather_state_evidence()
        results = _run_validation(evidence)
        for v in results:
            icon = "PASS" if v["status"] == "PASS" else "WARN"
            click.echo(f"  [{icon}] {v['check']}: {v['details']}")
        failed = any(v["status"] != "PASS" for v in results)
        click.echo(f"\n{'WARNINGS present' if failed else 'All checks passed'}")
        return

    out = Path(out_dir)
    out.mkdir(parents=True, exist_ok=True)

    ts = datetime.now(timezone.utc).strftime("%Y-%m-%d")
    suffix = f"_{state}" if state else ""
    base = f"Tallowbourn_FBA_Inventory_Presence_{ts}"

    md_content = None
    csv_content = None
    pdf_bytes = None

    if fmt in ("md", "all"):
        md_content = build_markdown(state_filter=state)
        # Latest
        (out / f"inventory_presence_latest{suffix}.md").write_text(md_content, encoding="utf-8")
        # Timestamped
        p = out / f"{base}{suffix}.md"
        p.write_text(md_content, encoding="utf-8")
        click.echo(f"Markdown: {p}")

    if fmt in ("csv", "all"):
        csv_content = build_csv(state_filter=state)
        (out / f"inventory_presence_latest{suffix}.csv").write_text(csv_content, encoding="utf-8")
        p = out / f"{base}{suffix}.csv"
        p.write_text(csv_content, encoding="utf-8")
        click.echo(f"CSV: {p}")

    if fmt in ("pdf", "all"):
        pdf_bytes = build_pdf(state_filter=state)
        (out / f"inventory_presence_latest{suffix}.pdf").write_bytes(pdf_bytes)
        p = out / f"{base}{suffix}.pdf"
        p.write_bytes(pdf_bytes)
        click.echo(f"PDF: {p} ({len(pdf_bytes):,} bytes)")

    # Write metadata sidecar
    if fmt == "all":
        meta = build_metadata(state_filter=state)
        import json
        (out / "inventory_presence_meta.json").write_text(
            json.dumps(meta, indent=2), encoding="utf-8")
        click.echo(f"Metadata: {out / 'inventory_presence_meta.json'}")

    # Upload to Supabase Storage
    if not no_upload and fmt == "all" and not state:
        if md_content and csv_content and pdf_bytes:
            click.echo("Uploading to Supabase Storage...")
            try:
                meta = build_metadata()
                results = upload_exports(md_content, csv_content, pdf_bytes, meta)
                ok = sum(1 for v in results.values() if v)
                click.echo(f"  Uploaded {ok}/{len(results)} files to cpa-exports bucket")
            except Exception as e:
                click.echo(f"  Storage upload failed (non-fatal): {e}", err=True)


@cli.command("economic-nexus-audit")
@click.option("--format", "fmt", type=click.Choice(["pdf", "csv", "json", "all"]),
              default="all", help="Output format")
@click.option("--out", "out_dir", default="exports/cpa",
              help="Output directory")
@click.option("--no-upload", is_flag=True, help="Skip Supabase Storage upload")
def economic_nexus_audit(fmt, out_dir, no_upload):
    """CPA Export: Economic Nexus Audit — transparent per-state analysis."""
    from pathlib import Path
    from datetime import datetime, timezone
    from src.exports.economic_nexus_audit import (
        build_audit, build_pdf, build_csv, build_meta, upload_exports,
    )

    audit = build_audit()
    exceeded = [s["state_code"] for s in audit["states"] if s["status"] == "exceeded"]
    approaching = [s["state_code"] for s in audit["states"] if s["status"] == "approaching"]

    # Print validation
    for v in audit["validation"]:
        icon = "PASS" if v["status"] == "PASS" else "WARN"
        click.echo(f"  [{icon}] {v['check']}: {v['details']}")

    click.echo(f"\nExceeded ({len(exceeded)}): {', '.join(exceeded)}")
    click.echo(f"Approaching ({len(approaching)}): {', '.join(approaching)}")

    ts = datetime.now(timezone.utc).strftime("%Y-%m-%d")
    out = Path(out_dir)
    out.mkdir(parents=True, exist_ok=True)

    if fmt in ("pdf", "all"):
        pdf_bytes = build_pdf(audit)
        p = out / f"Economic_Nexus_Audit_{ts}.pdf"
        p.write_bytes(bytes(pdf_bytes) if isinstance(pdf_bytes, bytearray) else pdf_bytes)
        click.echo(f"PDF: {p} ({len(pdf_bytes):,} bytes)")

    if fmt in ("csv", "all"):
        csv_content = build_csv(audit)
        p = out / f"Economic_Nexus_Audit_{ts}.csv"
        p.write_text(csv_content, encoding="utf-8")
        click.echo(f"CSV: {p}")

    if fmt in ("json", "all"):
        import json
        p = out / f"Economic_Nexus_Audit_{ts}.json"
        p.write_text(json.dumps(audit, indent=2, default=str), encoding="utf-8")
        click.echo(f"JSON: {p}")

    if not no_upload and fmt == "all":
        click.echo("Uploading to Supabase Storage...")
        try:
            results = upload_exports(audit)
            ok = sum(1 for v in results.values() if v)
            click.echo(f"  Uploaded {ok}/{len(results)} files to cpa-exports/economic-nexus/")
        except Exception as e:
            click.echo(f"  Storage upload failed (non-fatal): {e}", err=True)


@cli.command("kintsugi-compare")
@click.option("--file", "filepath", default=None,
              help="Path to Kintsugi XLSX (auto-detects newest in incoming/kintsugi/)")
@click.option("--window", "window", default="2024-01-01:",
              help="Date window as START:END (e.g. 2024-01-01:2026-08-16)")
@click.option("--out", "out_dir", default="exports/cpa",
              help="Output directory for CSV")
@click.option("--dry-run", is_flag=True, help="Parse Kintsugi only, no agent query")
@click.option("--no-upload", is_flag=True, help="Skip Storage upload")
def kintsugi_compare(filepath, window, out_dir, dry_run, no_upload):
    """Reconcile Kintsugi transaction report against agent sales data."""
    from pathlib import Path
    from src.exports.kintsugi_compare import (
        parse_jurisdiction_summary, build_comparison, build_csv, print_report,
    )
    from src.db import upload_to_storage

    # Auto-detect file
    if not filepath:
        kdir = Path("incoming/kintsugi")
        xlsx_files = sorted(kdir.glob("*.xlsx"), key=lambda p: p.stat().st_mtime, reverse=True)
        if not xlsx_files:
            click.echo("No .xlsx files found in incoming/kintsugi/", err=True)
            return
        filepath = str(xlsx_files[0])
        click.echo(f"Auto-detected: {filepath}")

    # Parse window
    parts = window.split(":")
    w_start = parts[0] if parts[0] else "2024-01-01"
    w_end = parts[1] if len(parts) > 1 and parts[1] else None

    if dry_run:
        rows = parse_jurisdiction_summary(filepath)
        click.echo(f"Kintsugi: {len(rows)} US states")
        total = sum(r["txn_amount"] for r in rows)
        click.echo(f"Grand total txn amount: ${total:,.2f}")
        for r in sorted(rows, key=lambda x: x["txn_amount"], reverse=True)[:10]:
            click.echo(f"  {r['state_code']}: ${r['txn_amount']:>12,.0f}  txns: {r['txn_count']:>6,}")
        return

    comparison = build_comparison(filepath, w_start, w_end)
    report = print_report(comparison)
    click.echo(report)

    # Write CSV
    out = Path(out_dir)
    out.mkdir(parents=True, exist_ok=True)
    from datetime import datetime, timezone
    ts = datetime.now(timezone.utc).strftime("%Y-%m-%d")
    csv_content = build_csv(comparison)
    csv_path = out / f"Kintsugi_Reconciliation_{ts}.csv"
    csv_path.write_text(csv_content, encoding="utf-8")
    click.echo(f"\nCSV: {csv_path}")

    # JSON
    import json
    json_path = out / f"Kintsugi_Reconciliation_{ts}.json"
    json_path.write_text(json.dumps(comparison, indent=2, default=str), encoding="utf-8")
    click.echo(f"JSON: {json_path}")

    # Upload
    if not no_upload:
        click.echo("Uploading to Supabase Storage...")
        try:
            upload_to_storage("cpa-exports", "kintsugi-compare/latest.csv",
                              csv_content.encode("utf-8"), "text/csv")
            upload_to_storage("cpa-exports", "kintsugi-compare/latest.json",
                              json.dumps(comparison, indent=2, default=str).encode("utf-8"),
                              "application/json")
            meta = {
                "generated_at": comparison["generated_at"],
                "window": comparison["window"],
                "totals": comparison["totals"],
                "flags_count": len(comparison["flags"]),
            }
            upload_to_storage("cpa-exports", "kintsugi-compare/meta.json",
                              json.dumps(meta, indent=2).encode("utf-8"), "application/json")
            click.echo("  Uploaded 3 files to cpa-exports/kintsugi-compare/")
        except Exception as e:
            click.echo(f"  Upload failed (non-fatal): {e}", err=True)


@cli.command("registration-triage")
@click.option("--format", "fmt", type=click.Choice(["md", "csv", "all"]),
              default="all", help="Output format")
@click.option("--out", "out_dir", default="exports/cpa",
              help="Output directory")
def registration_triage(fmt, out_dir):
    """CPA Export: Registration Triage by state (research aid, not advice)."""
    from pathlib import Path
    from src.exports.registration_triage import build_markdown, build_csv, build_triage_rows

    rows = build_triage_rows()

    # Print summary to console
    from collections import Counter
    buckets = Counter(r["triage_bucket"] for r in rows)
    click.echo("Registration Triage Summary:")
    for b in ["A_discuss", "D_entity_tax", "C_economic_watch", "B_monitor"]:
        if buckets.get(b):
            states = [r["state_code"] for r in rows if r["triage_bucket"] == b]
            click.echo(f"  {b}: {buckets[b]} states — {', '.join(states)}")
    click.echo()

    out = Path(out_dir)
    out.mkdir(parents=True, exist_ok=True)

    if fmt in ("md", "all"):
        md = build_markdown()
        p = out / "registration_triage.md"
        p.write_text(md, encoding="utf-8")
        click.echo(f"Markdown: {p}")

    if fmt in ("csv", "all"):
        csv_content = build_csv()
        p = out / "registration_triage.csv"
        p.write_text(csv_content, encoding="utf-8")
        click.echo(f"CSV: {p}")


@cli.command("github-backup")
@click.option("--dry-run", is_flag=True, help="Show what would be committed without pushing")
def github_backup(dry_run):
    """Push project state to a dated backup/* branch on GitHub."""
    from src.maintenance.github_backup import run_backup
    result = run_backup(dry_run=dry_run)
    if result["status"] == "success":
        click.echo(f"Backup: {result['message']}")
    elif result["status"] == "dry_run":
        click.echo(f"Dry run: {result['message']}")
    elif result["status"] == "nothing_to_backup":
        click.echo("Nothing to back up — working tree matches HEAD.")
    else:
        click.echo(f"Backup failed: {result.get('error', result['status'])}")


@cli.command("git-auto-update")
@click.option("--dry-run", is_flag=True,
              help="Fetch and report; never pull, never exit")
@click.option("--restart", is_flag=True,
              help="After a HEAD-changing pull, exit so launchd KeepAlive respawns")
def git_auto_update_cmd(dry_run, restart):
    """Fast-forward origin/main if the working tree is clean.

    The scheduler runs this daily at 04:30 ET and once shortly after
    startup. Manual use is for inspection: --dry-run never writes.
    Never pushes. Never resets. Aborts if dirty or diverged.
    """
    from src.maintenance.git_auto_update import run_auto_update

    result = run_auto_update(dry_run=dry_run, restart=restart if restart else False)
    status = result.get("status")
    detail = result.get("message") or result.get("error") or ""
    click.echo(f"{status}: {detail}".rstrip(": "))
    if result.get("restart"):
        from src.maintenance.git_auto_update import request_process_exit
        click.echo("HEAD changed — exiting so launchd KeepAlive can respawn.")
        request_process_exit(delay_seconds=0.2)


@cli.command("inventory-sync")
@click.option("--dry-run", is_flag=True)
def inventory_sync_cmd(dry_run):
    """Pull FBA inventory: restock, planning, stock levels, AWD."""
    from src.inventory.sync import sync_all
    if dry_run:
        click.echo("DRY RUN\n")
    click.echo(f"inventory-sync {_code_revision()}")
    click.echo("Inventory sync started (each step prints when it finishes)...")
    results = sync_all(dry_run=dry_run, echo=click.echo)
    sync_names = [
        "fba_summaries", "awd", "restock", "planning",
        "awd_replenishments", "inbound_shipments", "awd_inbound",
    ]
    for name in sync_names:
        r = results.get(name, {})
        if "error" in r:
            click.echo(f"{name}: ERROR — {r['error'][:200]}")
        elif r.get("skipped"):
            click.echo(f"{name}: SKIPPED — {r.get('skip_reason')}")
        elif r.get("upsert_error"):
            click.echo(
                f"{name}: {r.get('shipments_found', r.get('orders_found', 0))} found, "
                f"upsert warning — {r['upsert_error'][:200]}"
            )
        else:
            extra = ""
            if name == "fba_summaries" and r.get("daily"):
                extra = f", daily={r['daily'].get('rows', 0)}"
            click.echo(f"{name}: {r.get('rows_total', r.get('shipments_found', r.get('orders_found', 0)))} rows "
                       f"({r.get('rows_inserted', r.get('rows_upserted', 0))} upserted){extra}")


@cli.command("inventory-awd-test")
def inventory_awd_test_cmd():
    """Probe AWD API access (inventory, replenishment orders, inbound shipments)."""
    import json
    from src.inventory.awd_client import awd_probe, AWD_ROLE_HINT

    click.echo("Probing AWD SP-API endpoints…")
    result = awd_probe()
    for key, val in result.items():
        if key == "all_ok":
            continue
        if isinstance(val, dict) and val.get("ok"):
            click.echo(f"  {key}: OK (sample_count={val.get('sample_count', 0)})")
        else:
            click.echo(f"  {key}: FAIL — {(val or {}).get('error', val)[:240]}")
    if result.get("all_ok"):
        click.echo("\nAll AWD endpoints reachable.")
    else:
        click.echo(f"\nOne or more endpoints failed. {AWD_ROLE_HINT}")
    click.echo(json.dumps(result, indent=2))


@cli.command("inventory-supply-plan")
@click.option("--until", "until_date", default="2027-01-15")
@click.option("--sku", multiple=True, help="SKU filter (default: all active)")
def inventory_supply_plan_cmd(until_date, sku):
    """Four-number supply plan: manufacture, warehouse ship, FBA DOS, network OOS."""
    from src.inventory.supply_plan import build_four_numbers_plan, format_four_numbers_text
    skus = list(sku) if sku else None
    plan = build_four_numbers_plan(skus=skus, until_date=until_date)
    click.echo(format_four_numbers_text(plan))


@cli.command("inventory-inbound-plan")
@click.option("--until", "until_date", default="2027-01-15")
@click.option("--sku", multiple=True, help="SKU filter (default: lip balm SKUs)")
def inventory_inbound_plan_cmd(until_date, sku):
    """Warehouse → FBA inbound waves with receiving lead time."""
    from src.inventory.inbound_waves import build_inbound_wave_plan, format_inbound_plan_text
    skus = list(sku) if sku else None
    plan = build_inbound_wave_plan(skus=skus, until_date=until_date)
    click.echo(format_inbound_plan_text(plan))


@cli.command("inventory-holiday-surge")
@click.option("--year", default=None, type=int, help="Prior holiday year (default: last calendar year)")
def inventory_holiday_surge_cmd(year):
    """Recompute per-SKU holiday surge from prior-year Nov–Dec sales_by_sku."""
    from src.inventory.holiday_surge import compute_surge_from_sales
    click.echo("Computing per-SKU holiday surge from sales_by_sku...\n")
    r = compute_surge_from_sales(prior_year=year)
    if r.get("error"):
        click.echo(f"ERROR: {r['error']}")
        return
    click.echo(f"Prior year:        {r['prior_year']}")
    click.echo(f"Holiday mode:      {r['holiday_mode']}")
    click.echo(f"SKUs updated:      {r['skus_updated']}")
    click.echo(f"Surge SKUs (>1.05): {r['surge_skus']}")
    for t in r.get("top_surge", [])[:8]:
        click.echo(
            f"  {t['sku']}: {t['surge']:.2f}×  "
            f"holiday {t['holiday_daily']:.0f}/d  "
            f"plan {t['planning_u_30']:.0f}/d"
        )


@cli.command("inventory-calibrate")
@click.option("--dry-run", is_flag=True)
def inventory_calibrate_cmd(dry_run):
    """Recompute dual-rate signals + pull inbound shipment history."""
    from src.inventory.inbound_shipments import sync_inbound_shipments
    from src.inventory.rate_signals import sync_sku_signals
    click.echo(f"inventory-calibrate {_code_revision()}")
    if dry_run:
        click.echo("DRY RUN\n")
    else:
        from src.inventory.awd_replenishments import recompute_stored_replenish_days
        r = sync_inbound_shipments(days_back=180, dry_run=False)
        click.echo(f"Inbound shipments: {r.get('shipments_found', 0)} found, "
                   f"{r.get('rows_upserted', 0)} upserted")
        recomputed = recompute_stored_replenish_days()
        click.echo(f"AWD replenish recompute: {recomputed.get('updated', 0)} rows")
    sig = sync_sku_signals()
    click.echo(f"Rate signals: {sig.get('skus', 0)} SKUs")
    if sig.get("account_receive_days") is not None:
        click.echo(f"Account median receive: {sig['account_receive_days']}d "
                   f"(n={sig.get('account_receive_n', 0)})")
    if sig.get("account_replenish_days") is not None:
        click.echo(f"Account median AWD→FBA: {sig['account_replenish_days']}d "
                   f"(n={sig.get('account_replenish_n', 0)})")


@cli.command("inventory-velocity")
@click.option("--days", default=400)
@click.option("--dry-run", is_flag=True)
def inventory_velocity_cmd(days, dry_run):
    """Recompute SKU velocity + seasonality from order history."""
    from src.inventory.velocity import compute_velocity
    click.echo(f"Computing velocity from {days} days of history...\n")
    r = compute_velocity(amazon_days=days, shopify_days=days, dry_run=dry_run)
    click.echo(f"Amazon SKUs: {r['amazon_skus']}")
    click.echo(f"Shopify SKUs: {r['shopify_skus']}")
    click.echo(f"Total: {r['skus']}, seasonality weeks: {r['seasonality_weeks']}")
    click.echo(f"Forward multiplier: {r['avg_forward_mult']:.3f}")
    click.echo(f"Rows inserted: {r['rows_inserted']}")
    if not dry_run:
        from src.inventory.holiday_surge import compute_surge_from_sales
        click.echo("\nApplying prior-year holiday surge from sales_by_sku...")
        surge = compute_surge_from_sales()
        if surge.get("error"):
            click.echo(f"  Holiday surge ERROR: {surge['error']}")
        else:
            click.echo(
                f"  Prior year {surge['prior_year']}: {surge['surge_skus']} surge SKUs, "
                f"{surge['skus_updated']} velocity rows patched"
            )


@cli.command("inventory-report")
@click.option("--top", default=20)
def inventory_report_cmd(top):
    """Terminal summary: at-risk SKUs, reorder list."""
    from src.inventory.report import build_report
    report = build_report()
    s = report["summary"]
    click.echo(f"Active SKUs:     {s['active_skus']}")
    click.echo(f"FBA <60d cover:  {s['at_risk_skus']}")
    click.echo(f"Reorder total:   {s['total_our_reorder']:,} units")
    click.echo(f"Portfolio cover: {s['portfolio_weeks_cover']} weeks\n")
    for r in report["rows"][:top]:
        dos_str = f"{r['dos']:.0f}" if r['dos'] < 9999 else "—"
        click.echo(f"  {r['sku'][:22]:<22} {r['flag']:<9} FBA={r['fba_on_hand']:>5} V30={r['total_u_30']:>5.1f} DOS={dos_str:>4}")


@cli.command("inventory-3pl-sync")
@click.option("--dry-run", is_flag=True)
def inventory_3pl_sync_cmd(dry_run):
    """Pull 3PL inventory from Ship Sidekick."""
    from src.shipsidekick.client import sync_3pl
    r = sync_3pl(dry_run=dry_run)
    status = r.get("status") or "success"
    msg = r.get("message") or (
        f"{r['rows_total']} SKUs, {r['rows_inserted']} upserted"
    )
    click.echo(f"{status}: {msg}")
    if r.get("omitted_skus") and status == "success":
        click.echo("3PL omitted catalog SKUs: " + ", ".join(r["omitted_skus"]))


@cli.command("plan-sku")
@click.option("--sku", required=True)
@click.option("--until", "until_date", default="2027-03-31")
@click.option("--scenario", default="correction_factor",
              type=click.Choice(["correction_factor", "actual_2025", "optimistic"]))
def plan_sku_cmd(sku, until_date, scenario):
    """Forward sell-through plan for a single SKU.

    Uses imported weekly forecast (forecast_weekly) when available.
    Falls back to velocity × seasonality for weeks without forecast data.
    """
    from datetime import date as d, timedelta
    from src.db import fetch_all

    end = d.fromisoformat(until_date) + timedelta(days=14)
    today = d.today()

    vels = {r["sku"]: r for r in fetch_all("sku_velocity")}
    snaps = {r["sku"]: r for r in fetch_all("inventory_snapshots")}
    awds = {r["sku"]: r for r in fetch_all("inventory_awd")}
    tpls = {}
    try:
        tpls = {r["sku"]: r for r in fetch_all("inventory_3pl_snapshots")}
    except Exception:
        pass

    vel = vels.get(sku)
    snap = snaps.get(sku)
    if not vel:
        click.echo(f"SKU {sku} not found in sku_velocity")
        return

    base = float(vel.get("total_u_30", 0) or 0)
    fba = sum(int(snap.get(k, 0) or 0) for k in
              ["fulfillable", "reserved", "researching", "unfulfillable"]) if snap else 0
    inbound = sum(int(snap.get(k, 0) or 0) for k in
                  ["inbound_working", "inbound_shipped", "inbound_receiving"]) if snap else 0
    awd_oh = int(awds.get(sku, {}).get("awd_on_hand", 0) or 0)
    tpl_oh = int(tpls.get(sku, {}).get("available", 0) or 0)
    owned = fba + inbound + awd_oh + tpl_oh

    # Load forecast weekly series (range-based lookup: cursor within
    # [week_start, week_start+6] matches that forecast week)
    forecast_weeks: list[tuple[d, float]] = []
    try:
        fc_rows = fetch_all("forecast_weekly")
        for r in fc_rows:
            if r.get("sku") == sku and r.get("scenario") == scenario:
                ws = d.fromisoformat(str(r["week_start"]))
                forecast_weeks.append((ws, float(r.get("units", 0) or 0)))
        forecast_weeks.sort()
    except Exception:
        pass

    has_forecast = len(forecast_weeks) > 0

    def _get_forecast(cursor_date: d, cursor_end: d) -> float | None:
        """Find forecast that overlaps the plan week [cursor_date, cursor_end].
        Uses the forecast whose week_start is closest to cursor_date."""
        best = None
        best_dist = 999
        for ws, units in forecast_weeks:
            we = ws + timedelta(days=6)
            # Overlap: forecast [ws, we] intersects plan [cursor_date, cursor_end]
            if ws <= cursor_end and we >= cursor_date:
                dist = abs((ws - cursor_date).days)
                if dist < best_dist:
                    best = units
                    best_dist = dist
        return best

    # For summary display
    forecast_map = {ws.isoformat(): u for ws, u in forecast_weeks}

    # Seasonality fallback
    season: dict[int, float] = {}
    for r in fetch_all("seasonality_weekly"):
        if r.get("sku") == "_account_" and r.get("year") == 0:
            season[int(r["week"])] = float(r.get("multiplier", 1.0) or 1.0)

    # Header
    click.echo(f"{'='*65}")
    click.echo(f"  PLAN: {sku}")
    click.echo(f"  {today} → {end}")
    click.echo(f"  Demand source: {'forecast_weekly (' + scenario + ')' if has_forecast else 'velocity × seasonality'}")
    click.echo(f"{'='*65}")
    click.echo(f"  V30={base:.1f} u/day  FBA={fba:,}  inbound={inbound:,}  AWD={awd_oh:,}  3PL={tpl_oh:,}  owned={owned:,}")
    if has_forecast:
        fc_total = sum(forecast_map.values())
        click.echo(f"  Forecast weeks: {len(forecast_map)}  total={fc_total:,.0f} units ({scenario})")
    click.echo()

    # Weekly walk
    remaining = fba
    total_demand = 0
    forecast_demand = 0
    velocity_demand = 0
    stockout_week = None

    click.echo(f"  {'Wk':<4} {'Dates':<24} {'Source':<6} {'Demand':>7} {'FBA':>7}")
    click.echo(f"  {'-'*52}")

    cursor = today
    while cursor <= end:
        wk_end = min(cursor + timedelta(days=6), end)
        days = (wk_end - cursor).days + 1
        fc_units = _get_forecast(cursor, wk_end)

        if fc_units is not None:
            demand = round(fc_units)
            source = "FC"
            forecast_demand += demand
        else:
            mult = season.get(cursor.isocalendar()[1], 1.0)
            demand = round(base * days * mult)
            source = "V×S"
            velocity_demand += demand

        total_demand += demand
        remaining -= demand

        note = ""
        if remaining <= 0 and stockout_week is None:
            stockout_week = cursor
            note = " ← STOCKOUT"

        iso_wk = cursor.isocalendar()[1]
        click.echo(
            f"  W{iso_wk:<3} {cursor} → {wk_end}  {source:<6} {demand:>7,} {max(remaining, 0):>7,}{note}"
        )

        cursor = wk_end + timedelta(days=1)

    click.echo(f"  {'-'*52}")
    click.echo(f"\n  Total demand:    {total_demand:>8,}")
    if has_forecast:
        click.echo(f"    from forecast:  {forecast_demand:>8,} ({len(forecast_map)} weeks)")
        click.echo(f"    from velocity:  {velocity_demand:>8,} (remainder)")
    click.echo(f"  FBA on-hand:     {fba:>8,}")
    click.echo(f"  Owned total:     {owned:>8,}")
    click.echo(f"  Gap (FBA):       {max(total_demand - fba, 0):>8,}")
    click.echo(f"  Gap (owned):     {max(total_demand - owned, 0):>8,}")
    if stockout_week:
        click.echo(f"  FBA stockout:    {stockout_week}")
    click.echo()


@cli.command("pallet-plan")
@click.option("--scenario", default="correction_factor",
              type=click.Choice(["correction_factor", "actual_2025", "optimistic"]))
@click.option("--pallet-max", default=17550, help="Units per Amazon pallet (65×270)")
@click.option("--target", default="2026-10-31", help="In-Amazon-by date")
@click.option("--no-3pl", is_flag=True, help="Exclude 3PL from supply")
@click.option("--no-awd", is_flag=True, help="Exclude AWD from supply")
def pallet_plan_cmd(scenario, pallet_max, target, no_3pl, no_awd):
    """Lip Balm Monthly Pallet Planner — Nov+Dec holiday build."""
    from src.inventory.pallet_planner import build_pallet_plan

    plan = build_pallet_plan(
        pallet_max=pallet_max,
        amazon_in_by=target,
        scenario=scenario,
        include_3pl=not no_3pl,
        include_awd=not no_awd,
    )

    click.echo(f"{'='*65}")
    click.echo(f"  LIP BALM PALLET PLANNER — Holiday {plan['config']['scenario']}")
    click.echo(f"  All units in Amazon by: {plan['config']['amazon_in_by']}")
    click.echo(f"  Pallet max: {plan['config']['pallet_max_units']:,} units")
    click.echo(f"  3PL transfer: {'ON' if plan['config']['include_3pl_transfer'] else 'OFF'}")
    click.echo(f"  AWD in supply: {'ON' if plan['config']['include_awd'] else 'OFF'}")
    click.echo(f"{'='*65}")

    click.echo(f"\n  {'SKU':<16} {'NovDec':>8} {'FBA':>6} {'Inb':>5} {'AWD':>5} {'3PL':>6} {'Supply':>7} {'Gap':>7}")
    click.echo(f"  {'-'*62}")
    for p in plan["sku_plans"]:
        click.echo(
            f"  {p['sku']:<16} {p['nov_dec_demand']:>8,} {p['fba']:>6,} {p['inbound']:>5,} "
            f"{p['awd']:>5,} {p['tpl']:>6,} {p['amazon_supply']:>7,} {p['gap']:>7,}"
        )
    click.echo(f"  {'-'*62}")
    click.echo(f"  {'TOTAL':<16} {plan['total_nov_dec_demand']:>8,} "
               f"{'':>6} {'':>5} {'':>5} {'':>6} {plan['total_amazon_supply']:>7,} {plan['total_gap']:>7,}")

    click.echo(f"\n  Pallets needed: {plan['num_pallets']}")

    if plan["pallets"]:
        click.echo(f"\n  PALLET BREAKDOWN:")
        for p in plan["pallets"]:
            mix_str = " + ".join(f"{s} ×{q:,}" for s, q in p["mix"].items())
            click.echo(f"    Pallet {p['pallet_num']}: {p['total_units']:,} units — {mix_str}")

    if plan["monthly_schedule"]:
        click.echo(f"\n  MONTHLY BUILD SCHEDULE:")
        for month in plan["months"]:
            pallets_in_month = plan["monthly_schedule"].get(month, [])
            if pallets_in_month:
                total = sum(p["total_units"] for p in pallets_in_month)
                nums = ", ".join(f"P{p['pallet_num']}" for p in pallets_in_month)
                click.echo(f"    {month}: {total:,} units ({nums})")

    if plan["pallets"]:
        next_p = plan["pallets"][0]
        click.echo(f"\n  NEXT PALLET TO ORDER:")
        for s, q in next_p["mix"].items():
            click.echo(f"    {s}: {q:,} units")
        click.echo(f"    Total: {next_p['total_units']:,} units")
        click.echo(f"    Ship by: earliest possible (target receipt ≤ {plan['config']['amazon_in_by']})")

    click.echo()


@cli.command("fba-cover")
@click.option("--target-days", default=60, help="Cover target in days")
@click.option("--scenario", default="correction_factor",
              type=click.Choice(["correction_factor", "actual_2025", "optimistic"]))
@click.option("--no-3pl", is_flag=True, help="Exclude 3PL transfers")
@click.option("--no-awd", is_flag=True, help="Exclude AWD transfers")
def fba_cover_cmd(target_days, scenario, no_3pl, no_awd):
    """FBA Cover Projection — flag weeks below 60-day service target."""
    from src.inventory.pallet_planner import build_fba_cover_projection, LIP_BALM_SKUS

    SKU_SHORT = {
        "DDPE0001Shop": "Unscnt",
        "DDPE0002Shop": "Pepper",
        "DDPE0003Shop": "Orange",
        "DDPE0004Shop": "Assrtd",
    }

    proj = build_fba_cover_projection(
        cover_target_days=target_days,
        scenario=scenario,
        include_3pl=not no_3pl,
        include_awd=not no_awd,
    )

    click.echo(f"{'='*70}")
    click.echo(f"  FBA COVER PROJECTION — {target_days}-day target · {scenario}")
    click.echo(f"{'='*70}")

    for sp in proj["sku_projections"]:
        sku = sp["sku"]
        label = SKU_SHORT.get(sku, sku)
        click.echo(f"\n  {label} ({sku})  FBA now: {sp['fba_start']:,}  "
                   f"Inbound: {sp['inbound']:,}  AWD: {sp['awd']:,}  3PL: {sp['tpl']:,}")
        click.echo(f"  {'Week':<12} {'Demand':>7} {'Receipt':>8} {'FBA':>8} {'Rate/d':>7} {'Cover':>7}")
        click.echo(f"  {'-'*52}")
        for w in sp["weeks"]:
            cover_str = f"{w['cover_days']}d" if w["cover_days"] is not None else "—"
            flag = " <<<" if w["flagged"] else ""
            click.echo(
                f"  {w['week']:<12} {w['demand']:>7,} {w['receipt']:>8,} "
                f"{w['fba']:>8,} {w['daily_rate']:>7} {cover_str:>7}{flag}"
            )

    if proj["alerts"]:
        click.echo(f"\n  {'!'*50}")
        click.echo(f"  {len(proj['alerts'])} WEEKS BELOW {target_days}-DAY COVER TARGET:")
        for a in proj["alerts"]:
            click.echo(f"    {SKU_SHORT.get(a['sku'], a['sku'])} {a['week']}: "
                       f"{a['cover_days']}d cover ({a['fba']:,} units @ {a['daily_rate']}/day)")
        click.echo(f"  {'!'*50}")
    else:
        click.echo(f"\n  All weeks maintain ≥{target_days}-day FBA cover.")

    click.echo()


@cli.command("mfg-headsup")
@click.option("--commit", "commit_months", multiple=True, help="Months to mark FIRM (e.g. 2026-08)")
@click.option("--tpl-offsets", is_flag=True, help="Let 3PL reduce manufacture (aggressive low-produce view)")
@click.option("--no-jan", is_flag=True, help="Exclude January from holiday demand")
@click.option("--weights", default="25,35,40", help="Month weights pct (e.g. 25,35,40)")
@click.option("--csv", "csv_out", type=click.Path(), help="Export CSV to file")
@click.option("--sheet", "sheet_out", type=click.Path(), help="Export printable summary to file")
def mfg_headsup_cmd(commit_months, tpl_offsets, no_jan, weights, csv_out, sheet_out):
    """Manufacturer Heads-Up — rolling 3-month production schedule."""
    from src.inventory.pallet_planner import (
        build_manufacturer_headsup, format_manufacturer_csv,
        format_manufacturer_sheet,
    )

    w = tuple(float(x) / 100 for x in weights.split(","))

    headsup = build_manufacturer_headsup(
        month_weights=w,
        include_jan=not no_jan,
        tpl_offsets_production=tpl_offsets,
        committed_months=list(commit_months) if commit_months else None,
    )

    click.echo(format_manufacturer_sheet(headsup))

    if csv_out:
        with open(csv_out, "w") as f:
            f.write(format_manufacturer_csv(headsup))
        click.echo(f"CSV exported to {csv_out}")

    if sheet_out:
        with open(sheet_out, "w") as f:
            f.write(format_manufacturer_sheet(headsup))
        click.echo(f"Planning sheet exported to {sheet_out}")


@cli.command("import-forecast")
@click.argument("file_path")
@click.option("--dry-run", is_flag=True)
def import_forecast_cmd(file_path, dry_run):
    """Import holiday forecast from xlsx into forecast_weekly."""
    from src.parsers.forecast_xlsx import import_forecast

    if dry_run:
        click.echo("DRY RUN\n")

    result = import_forecast(file_path, dry_run=dry_run)
    click.echo(f"File: {result['file']}")
    click.echo(f"Variants: {', '.join(result['variants'])}")
    click.echo(f"SKUs: {', '.join(result['skus'])}")
    click.echo(f"Weeks: {result['weeks']}")
    click.echo(f"Rows: {result['rows_total']} parsed, {result['rows_inserted']} inserted")

    if dry_run:
        # Show sample
        from src.parsers.forecast_xlsx import parse_forecast
        parsed = parse_forecast(file_path)
        click.echo("\nSample rows:")
        for r in parsed["rows"][:9]:
            click.echo(f"  {r['sku']} {r['week_start']} {r['scenario']:<20} {r['units']:>8.0f}")


@cli.command("import-3pl")
@click.argument("file_path")
@click.option("--dry-run", is_flag=True)
def import_3pl_cmd(file_path, dry_run):
    """Import 3PL invoice CSV into cost tracking tables."""
    from src.parsers.tpl_invoice import import_tpl_invoice

    if dry_run:
        click.echo("DRY RUN\n")

    result = import_tpl_invoice(file_path, dry_run=dry_run)
    click.echo(f"File: {result['file']}")
    click.echo(f"Months: {', '.join(result['months_covered'])}")
    click.echo(f"Monthly rows: {result['monthly_count']}")
    click.echo(f"Fee rows: {result['fee_count']}")
    click.echo(f"Detail rows: {result['detail_count']}")
    click.echo(f"Rows inserted: {result['rows_inserted']}")


@cli.command("filing-packet")
@click.option("--state", default=None, help="Single state or all registered")
@click.option("--out", default="exports/filings", help="Output directory")
@click.option("--end-date", default=None, help="Period end (default: yesterday)")
def filing_packet_cmd(state, out, end_date):
    """Export filing packet CSV for registered states.

    Uses actual period end (yesterday or --end-date), not future month-end.
    Current month sales are prorated by days elapsed.
    """
    from datetime import date as d, timedelta
    from pathlib import Path
    from src.db import fetch_all
    from src.channels import normalize_channel, is_seller_responsible, display_label
    import calendar, csv

    yesterday = d.today() - timedelta(days=1)
    cutoff = d.fromisoformat(end_date) if end_date else yesterday

    nexus = fetch_all("nexus_status")
    sales = fetch_all("sales_by_state")
    registered = [n for n in nexus if n.get("is_registered")]
    if state:
        registered = [n for n in registered if n["state_code"] == state.upper()]

    Path(out).mkdir(parents=True, exist_ok=True)

    # Current month boundaries for partial-month detection
    current_month_start = cutoff.replace(day=1).isoformat()

    # Fetch Shopify orders for partial month (actual dated sales)
    partial_month_orders = {}  # {state_code: {channel: {gross, tax, count}}}
    try:
        from src.config import settings
        if settings.shopify_enabled:
            from src.shopify_auth import auth_headers
            from src.channels import classify_shopify_order
            import httpx

            headers = auth_headers()
            base_url = f"https://{settings.shopify_shop_domain}/admin/api/2024-01/orders.json"
            params = {
                "status": "any", "limit": 250,
                "fields": "id,created_at,subtotal_price,total_tax,shipping_address,source_name",
                "created_at_min": current_month_start + "T00:00:00Z",
                "created_at_max": (cutoff + timedelta(days=1)).isoformat() + "T00:00:00Z",
            }
            all_orders = []
            url = base_url
            while url:
                resp = httpx.get(url, headers=headers,
                                 params=params if url == base_url else None, timeout=30)
                if resp.status_code != 200:
                    break
                all_orders.extend(resp.json().get("orders", []))
                link_header = resp.headers.get("link", "")
                url = None
                if 'rel="next"' in link_header:
                    for part in link_header.split(","):
                        if 'rel="next"' in part:
                            url = part.split("<")[1].split(">")[0]
                            break

            for order in all_orders:
                addr = order.get("shipping_address") or {}
                prov = (addr.get("province_code") or "").upper()
                country = addr.get("country_code", "")
                if country != "US" or not prov:
                    continue
                created = order.get("created_at", "")[:10]
                if created > cutoff.isoformat():
                    continue
                source = order.get("source_name", "")
                ch = classify_shopify_order(source, d.fromisoformat(created))
                gross = float(order.get("subtotal_price", 0) or 0)
                tax = float(order.get("total_tax", 0) or 0)

                if prov not in partial_month_orders:
                    partial_month_orders[prov] = {}
                if ch not in partial_month_orders[prov]:
                    partial_month_orders[prov][ch] = {"gross": 0, "tax": 0, "count": 0}
                partial_month_orders[prov][ch]["gross"] += gross
                partial_month_orders[prov][ch]["tax"] += tax
                partial_month_orders[prov][ch]["count"] += 1
    except Exception as e:
        click.echo(f"  Warning: could not fetch partial-month orders: {e}")

    for n in registered:
        sc = n["state_code"]
        lft = n.get("last_filed_through") or ""
        period_start = (d.fromisoformat(lft) + timedelta(days=1)).isoformat() if lft else ""
        rows_out = []

        for s in sales:
            if s.get("state_code") != sc:
                continue
            ps = s.get("period_start", "")
            pe = s.get("period_end", "")
            if not ps:
                continue
            if lft and pe <= lft:
                continue
            if ps > cutoff.isoformat():
                continue

            ch = normalize_channel(s.get("channel", ""))

            # For the current partial month: use actual order data instead
            if pe > cutoff.isoformat() and ps <= cutoff.isoformat():
                continue  # skip the monthly aggregate; we'll add order-level below

            rows_out.append({
                "state": sc,
                "period_start": max(ps, period_start) if period_start else ps,
                "period_end": pe,
                "channel": ch,
                "channel_label": display_label(ch),
                "seller_responsible": is_seller_responsible(ch),
                "gross_sales": float(s.get("gross_sales", 0) or 0),
                "tax_collected": float(s.get("tax_collected", 0) or 0),
                "order_count": int(s.get("order_count", 0) or 0),
            })

        # Add actual partial-month Shopify data from orders
        if sc in partial_month_orders:
            for ch, agg in partial_month_orders[sc].items():
                rows_out.append({
                    "state": sc,
                    "period_start": max(current_month_start, period_start) if period_start else current_month_start,
                    "period_end": cutoff.isoformat(),
                    "channel": ch,
                    "channel_label": display_label(ch),
                    "seller_responsible": is_seller_responsible(ch),
                    "gross_sales": round(agg["gross"], 2),
                    "tax_collected": round(agg["tax"], 2),
                    "order_count": agg["count"],
                })

        if not rows_out:
            click.echo(f"  {sc}: no activity since {lft or 'never'}")
            continue

        fname = Path(out) / f"filing_{sc}_{cutoff.isoformat()}.csv"
        with open(fname, "w", newline="") as f:
            w = csv.DictWriter(f, fieldnames=list(rows_out[0].keys()))
            w.writeheader()
            w.writerows(rows_out)

        seller = sum(r["gross_sales"] for r in rows_out if r["seller_responsible"])
        total = sum(r["gross_sales"] for r in rows_out)
        click.echo(f"  {sc}: {fname.name} ({len(rows_out)} rows, seller=${seller:,.0f}, total=${total:,.0f})")

    click.echo(f"\nDisclaimer: Filing packets are reference data only. Does not submit to any DOR.")


@cli.command("integrity-check")
def integrity_check():
    """Verify data integrity: SKU case, duplicate keys, channel totals."""
    from src.db import fetch_all
    from src.channels import normalize_channel

    issues = 0

    # 1. SKU case check
    click.echo("Checking sales_by_sku SKU case normalization...")
    sku_rows = fetch_all("sales_by_sku")
    bad_case = [r for r in sku_rows if r.get("sku") and r["sku"] != r["sku"].strip().upper()]
    if bad_case:
        click.echo(f"  FAIL: {len(bad_case)} rows with non-uppercase SKU")
        for r in bad_case[:5]:
            click.echo(f"    {r['sku']!r} in {r.get('channel')} {r.get('period_start')}")
        issues += 1
    else:
        click.echo(f"  OK: {len(sku_rows)} rows, all uppercase")

    # 2. Duplicate key check
    click.echo("Checking for duplicate upsert keys in sales_by_sku...")
    seen: dict[tuple, int] = {}
    for r in sku_rows:
        key = (r.get("channel"), r.get("sku"), r.get("state_code"),
               r.get("period_start"), r.get("source"))
        seen[key] = seen.get(key, 0) + 1
    dupes = {k: v for k, v in seen.items() if v > 1}
    if dupes:
        click.echo(f"  FAIL: {len(dupes)} duplicate keys")
        for k, v in list(dupes.items())[:3]:
            click.echo(f"    {k}: {v} rows")
        issues += 1
    else:
        click.echo(f"  OK: {len(seen)} unique keys, no duplicates")

    # 3. Channel totals cross-check
    click.echo("Cross-checking sales_by_state vs sales_by_sku channel totals...")
    state_rows = fetch_all("sales_by_state")
    for ch in ["shopify", "amazon"]:
        state_total = sum(
            float(r.get("gross_sales", 0))
            for r in state_rows
            if normalize_channel(r.get("channel", "")) == ch
        )
        sku_total = sum(
            float(r.get("gross_sales", 0))
            for r in sku_rows
            if normalize_channel(r.get("channel", "")) == ch
        )
        diff = abs(state_total - sku_total)
        pct = (diff / state_total * 100) if state_total > 0 else 0
        # SKU totals can differ from state totals due to line-item vs
        # order-level aggregation (discounts, multi-item orders, etc.)
        # and different time range coverage.  WARN above 10%, NOTE above 50%.
        status = "OK" if pct < 10 else "WARN" if pct < 100 else "FAIL"
        click.echo(f"  {ch}: state=${state_total:,.0f} sku=${sku_total:,.0f} "
                    f"diff=${diff:,.0f} ({pct:.1f}%) [{status}]")
        if status == "FAIL":
            issues += 1

    # 4. Period start sanity
    click.echo("Checking period_start format (must be YYYY-MM-01)...")
    bad_periods_state = [r for r in state_rows
                         if r.get("period_start") and not str(r["period_start"]).endswith("-01")]
    bad_periods_sku = [r for r in sku_rows
                       if r.get("period_start") and not str(r["period_start"]).endswith("-01")]
    if bad_periods_state:
        click.echo(f"  FAIL: {len(bad_periods_state)} sales_by_state rows with non-01 period_start")
        issues += 1
    else:
        click.echo(f"  OK: sales_by_state all YYYY-MM-01")
    if bad_periods_sku:
        click.echo(f"  FAIL: {len(bad_periods_sku)} sales_by_sku rows with non-01 period_start")
        issues += 1
    else:
        click.echo(f"  OK: sales_by_sku all YYYY-MM-01")

    # 5. Refund semantics
    sku_with_refunds = sum(1 for r in sku_rows if float(r.get("refund_sales", 0)) > 0)
    sku_zero_refunds = sum(1 for r in sku_rows if float(r.get("refund_sales", 0)) == 0)
    click.echo(f"Refund coverage: {sku_with_refunds} rows with refund data, "
               f"{sku_zero_refunds} with zero/null (may be missing, not confirmed zero)")

    click.echo(f"\n{'PASSED' if issues == 0 else f'FAILED ({issues} issues)'}")


@cli.command()
def run():
    """Start the background agent (folder watcher + scheduled tasks)."""
    from src.watcher.folder_watcher import start_watcher

    click.echo("Starting Sales Tax Compliance Agent...")
    click.echo("Press Ctrl+C to stop.\n")

    observer = start_watcher(print_fn=click.echo)

    try:
        from apscheduler.schedulers.blocking import BlockingScheduler
        from src.rules import AGENT_TZ, AGENT_TZ_NAME

        # Every cron below fires on AGENT_TZ (config/business_rules.json →
        # agent.timezone), not on whatever the machine's clock is set to, so a
        # laptop that travels or a changed system setting cannot silently move
        # the sync window. Amazon day boundaries stay on America/Los_Angeles.
        scheduler = BlockingScheduler(timezone=AGENT_TZ)
        global _SCHEDULER
        _SCHEDULER = scheduler
        click.echo(f"[Scheduler] Timezone: {AGENT_TZ_NAME}")

        # ── Liveness, before anything else ──
        # The heartbeat is a local file stamped on a short interval. It detects
        # a WEDGED scheduler: a process that is alive while its job threads have
        # stopped. It cannot detect a dead process — a dead scheduler also sends
        # no morning check-in, so silence is that signal, and a missing 08:10
        # message is itself the alarm.
        from src.health import CONFIG as _HEALTH, write_heartbeat

        write_heartbeat()                      # stamp immediately on startup
        scheduler.add_job(
            write_heartbeat,
            "interval",
            minutes=int(_HEALTH["heartbeat"]["interval_minutes"]),
            id="heartbeat",
            coalesce=True,
            max_instances=1,
        )
        click.echo(f"[Scheduler] Heartbeat every "
                   f"{_HEALTH['heartbeat']['interval_minutes']}m "
                   f"-> {_HEALTH['heartbeat']['path']}")

        _hs = _HEALTH["schedule"]
        scheduler.add_job(
            _run_health_ping,
            "cron",
            hour=int(_hs["hour"]),
            minute=int(_hs["minute"]),
            timezone=_hs["timezone"],
            id="health_ping",
            misfire_grace_time=3600,
            coalesce=True,
        )
        click.echo(f"[Scheduler] Health check-in daily at {int(_hs['hour']):02d}:"
                   f"{int(_hs['minute']):02d} {_hs['timezone']} "
                   f"(one message/day; HEALTH_TELEGRAM=0 mutes)")

        from src.config import settings

        if settings.shopify_enabled:
            from src.parsers.shopify_orders import fetch_shopify_orders_api
            scheduler.add_job(
                _run_shopify_poll,
                "interval",
                hours=settings.shopify_poll_interval_hours,
                id="shopify_poll",
            )
            click.echo(f"[Scheduler] Shopify polling every {settings.shopify_poll_interval_hours}h")

        scheduler.add_job(
            _run_daily_analysis,
            "cron",
            hour=8,
            minute=0,
            id="daily_analysis",
        )
        click.echo("[Scheduler] Daily analysis at 08:00")

        scheduler.add_job(
            _run_deadline_check,
            "cron",
            hour=9,
            minute=0,
            id="deadline_check",
        )
        click.echo("[Scheduler] Deadline check at 09:00")

        if settings.amazon_sp_enabled:
            scheduler.add_job(
                _run_spapi_refresh,
                "cron",
                hour=6,
                minute=0,
                id="spapi_refresh",
            )
            click.echo("[Scheduler] SP-API refresh daily at 06:00 "
                       "(orders 7d, inventory ledger 14d)")

            # Current-month sales_by_sku only. spapi_refresh writes
            # sales_daily / sales_by_state, not SKU months — without this
            # the in-progress month freezes at the last manual
            # backfill-amazon-skus. Window is month-start → yesterday
            # (one All Orders chunk). Never the CLI default of 2025-01-01.
            scheduler.add_job(
                _run_amazon_sku_month,
                "cron",
                hour=6,
                minute=20,
                id="amazon_sku_month",
                misfire_grace_time=3600,
                coalesce=True,
                max_instances=1,
            )
            click.echo("[Scheduler] Amazon SKU month daily at 06:20 "
                       "(current month → yesterday, after spapi_refresh)")

            # Weekly deeper ledger re-pull. Sunday 04:00 keeps it clear of the
            # 06:00 daily refresh and of the ads jobs at 05:00-05:30.
            scheduler.add_job(
                _run_inventory_ledger_backfill,
                "cron",
                day_of_week="sun",
                hour=4,
                minute=0,
                id="inventory_ledger_backfill",
                misfire_grace_time=7200,
                coalesce=True,
            )
            click.echo("[Scheduler] Inventory ledger backfill weekly Sunday 04:00 (90d)")

            # Weekly Brand Analytics SQP -> organic-rank gate. Monday 10:00
            # Pacific, after Amazon publishes the prior Sun-Sat week.
            _sqp_cfg = {}
            try:
                from src.amazon_ads.organic_rank import load_config as _rank_cfg
                _sqp_cfg = (_rank_cfg().get("sqp_auto") or {})
            except Exception:
                pass
            if _sqp_cfg.get("enabled"):
                _sched = _sqp_cfg.get("schedule") or {}
                scheduler.add_job(
                    _run_sqp_sync,
                    "cron",
                    day_of_week=_sched.get("day_of_week", "mon"),
                    hour=int(_sched.get("hour", 10)),
                    minute=int(_sched.get("minute", 0)),
                    timezone=_sched.get("timezone", "America/Los_Angeles"),
                    id="sqp_sync",
                    misfire_grace_time=21600,
                    coalesce=True,
                )
                click.echo(f"[Scheduler] SQP sync weekly {_sched.get('day_of_week','mon')} "
                           f"{int(_sched.get('hour',10)):02d}:"
                           f"{int(_sched.get('minute',0)):02d} "
                           f"{_sched.get('timezone','America/Los_Angeles')} "
                           f"({len(_sqp_cfg.get('asins') or [])} ASINs)")

        scheduler.add_job(
            _run_source_monitoring,
            "cron",
            day_of_week="mon",
            hour=7,
            minute=0,
            id="source_monitoring",
        )
        click.echo("[Scheduler] Source monitoring every Monday at 07:00")

        # Paid Ads CSV freshness: Monday 07:15 ET.
        #
        # /paid-ads is fed by manual exports, so its staleness banner only
        # appears once the page is already open. This is the nudge that arrives
        # without being asked for, at the start of the week when the exports
        # would be pulled anyway. Read-only: it never writes paid_* rows.
        scheduler.add_job(
            _run_paid_ads_freshness,
            "cron",
            day_of_week="mon",
            hour=7,
            minute=15,
            id="paid_ads_freshness",
        )
        click.echo("[Scheduler] Paid Ads CSV freshness every Monday at 07:15")

        # CPA export: daily at 06:30 ET
        scheduler.add_job(
            _run_cpa_exports,
            "cron",
            hour=6,
            minute=30,
            id="cpa_exports",
        )
        click.echo("[Scheduler] CPA exports daily at 06:30")

        # daily_digest is NOT scheduled.
        #
        # It rendered the same registration-aware sections as the 08:00 daily
        # summary — both call digest_sections.build_sections — and fired five
        # minutes after it. Three Telegram messages landed every morning
        # (08:00 tax summary, 08:05 near-identical digest, 08:10 health), which
        # is how a monitor gets muted, and a muted monitor is worse than none.
        #
        # Its two unique pieces — MTD sales by channel, and the 24h failed-job
        # list — were NOT in the 08:00 summary, so they moved into the 08:10
        # health check-in rather than being dropped. The failed-job list in
        # particular was the only place a broken inventory_sync was reported.
        # The module is kept and still runs on demand: `python -m src.main digest`.
        click.echo("[Scheduler] Daily digest: on demand only (was 08:05 — "
                   "duplicated the 08:00 summary)")

        # Publish the PPC brief at 07:30, after ads sync (05:00-05:30), the
        # actions rebuild (06:00) and the outcome snapshots (07:00) — so the
        # stored copy reflects the same day's queue.
        #
        # Without this the dashboard's stored fallback only ever held whatever
        # the operator last published by hand, which is exactly the weekly
        # terminal step this system is supposed to remove. It also keeps the
        # stored brief on the current format version.
        scheduler.add_job(
            _run_ppc_export_publish,
            "cron",
            hour=7,
            minute=30,
            id="ppc_export_publish",
            misfire_grace_time=3600,
            coalesce=True,
        )
        click.echo("[Scheduler] PPC brief publish daily at 07:30 "
                   "(feeds the dashboard Download/Copy buttons)")

        # Inventory sync daily at 06:30 (after SP-API refresh)
        if settings.amazon_sp_enabled:
            scheduler.add_job(
                _run_inventory_sync,
                "cron",
                hour=6,
                minute=30,
                id="inventory_sync",
            )
            click.echo("[Scheduler] Inventory sync daily at 06:30")

            scheduler.add_job(
                _run_ledger_summary,
                "cron",
                hour=6,
                minute=40,
                id="ledger_summary_daily",
                misfire_grace_time=3600,
                coalesce=True,
                max_instances=1,
            )
            click.echo("[Scheduler] Ledger summary (tax inventory $) daily at 06:40")

        # 3PL sync daily at 06:35
        scheduler.add_job(
            _run_3pl_sync,
            "cron",
            hour=6,
            minute=35,
            id="3pl_sync",
        )
        click.echo("[Scheduler] 3PL sync daily at 06:35")

        # ── Amazon Ads: three independent jobs, deliberately not one ──
        # Campaigns are fast and feed the /ppc KPIs + trends; search terms are
        # slow (up to 90 min) and feed the Actions queue. Splitting them means a
        # search-term timeout can never delay or cancel the campaign refresh.
        # misfire_grace_time keeps a job that was asleep at its slot from being
        # skipped outright; coalesce collapses a backlog into one run.
        if settings.amazon_ads_enabled:
            scheduler.add_job(
                _run_ads_campaigns_sync,
                "cron",
                hour=5,
                minute=0,
                id="ads_campaigns_sync",
                misfire_grace_time=3600,
                coalesce=True,
                max_instances=1,
            )
            click.echo("[Scheduler] Ads campaigns sync daily at 05:00 (SP 7d, SB/SD 7d×1d chunks, then placements + search terms)")

            scheduler.add_job(
                _run_ads_search_terms_sync,
                "cron",
                hour=5,
                minute=30,
                id="ads_search_terms_sync",
                misfire_grace_time=3600,
                coalesce=True,
                max_instances=1,   # a 90-min run must never overlap itself
            )
            click.echo("[Scheduler] Ads search terms sync daily at 05:30 (7d, 7d chunks)")

            # Midday SB/SD heal — catches overnight partials that the
            # post-failure retry also missed (lock busy, Amazon still slow).
            scheduler.add_job(
                _run_ads_sb_sd_heal,
                "cron",
                hour=13,
                minute=0,
                id="ads_sb_sd_heal",
                misfire_grace_time=3600,
                coalesce=True,
                max_instances=1,
            )
            click.echo("[Scheduler] Ads SB/SD heal daily at 13:00 (fills SP-only days)")

            scheduler.add_job(
                _run_ads_actions,
                "cron",
                hour=6,
                minute=0,
                id="ads_actions",
                misfire_grace_time=3600,
                coalesce=True,
                max_instances=1,
            )
            click.echo("[Scheduler] Ads actions daily at 06:00 (7d, break-even target from account_target_acos — sole owner of the queue)")

            scheduler.add_job(
                _run_ads_campaigns_backfill,
                "cron",
                day_of_week="sun",
                hour=3,
                minute=0,
                id="ads_campaigns_backfill",
                misfire_grace_time=7200,
                coalesce=True,
                max_instances=1,
            )
            click.echo("[Scheduler] Ads campaigns backfill weekly Sunday 03:00 (SP 90d, SB/SD 7d)")

            scheduler.add_job(
                _run_ads_search_terms_backfill,
                "cron",
                day_of_week="sun",
                hour=3,
                minute=30,
                id="ads_search_terms_backfill",
                misfire_grace_time=7200,
                coalesce=True,
                max_instances=1,
            )
            click.echo("[Scheduler] Ads search terms backfill weekly Sunday 03:30 (90d, 7d chunks; weekday 7d unchanged)")

            # Placement performance at 05:15 — between campaigns and search
            # terms. Second spCampaigns report, so it is its own job rather
            # than doubling the 05:00 run's time.
            scheduler.add_job(
                _run_ads_placements_sync,
                "cron",
                hour=5,
                minute=15,
                id="ads_placements_sync",
                misfire_grace_time=3600,
                coalesce=True,
                max_instances=1,
            )
            click.echo("[Scheduler] Ads placements sync daily at 05:15 (14d)")
        else:
            click.echo("[Scheduler] Amazon Ads not configured — ads jobs not scheduled")

        # Contribution P&L at 06:45 — deliberately after the 05:00 ads sync and
        # the 06:00 SP-API refresh that writes sales_daily, so the day's sales,
        # ad spend and settlement are all present before contribution is stored.
        # Not gated on Ads: without ads data the formula still holds, ads = 0.
        scheduler.add_job(
            _run_pnl_sync,
            "cron",
            hour=6,
            minute=45,
            id="pnl_sync",
            misfire_grace_time=3600,
            coalesce=True,
        )
        click.echo("[Scheduler] Contribution P&L daily at 06:45 (after sales + ads)")

        # Outcome snapshots at 07:00 — last in the chain, because it reads the
        # ads tables (05:00-05:30), the action decisions (06:00) and the P&L
        # contribution rows (06:45) that the earlier jobs write.
        scheduler.add_job(
            _run_ads_outcomes,
            "cron",
            hour=7,
            minute=0,
            id="ads_outcomes",
            misfire_grace_time=3600,
            coalesce=True,
        )
        click.echo("[Scheduler] Action outcome snapshots daily at 07:00 (after P&L)")

        # Safe ff-only pull of origin/main. One job inside this scheduler —
        # not a second launchd agent — so it cannot race the running process.
        # 04:30 is before ads_campaigns_sync (05:00). A startup pass catches
        # merges that landed while the Mini was off. See
        # src/maintenance/git_auto_update.py for the safety rules.
        from datetime import datetime, timedelta
        from src.maintenance.git_auto_update import (
            CRON_HOUR as _GIT_UPD_HOUR,
            CRON_MINUTE as _GIT_UPD_MINUTE,
            STARTUP_DELAY_SECONDS as _GIT_UPD_DELAY,
            is_enabled as _git_auto_update_enabled,
        )
        if _git_auto_update_enabled():
            scheduler.add_job(
                _run_git_auto_update,
                "cron",
                hour=_GIT_UPD_HOUR,
                minute=_GIT_UPD_MINUTE,
                id="git_auto_update",
                misfire_grace_time=3600,
                coalesce=True,
                max_instances=1,
            )
            click.echo(
                f"[Scheduler] Git auto-update daily at "
                f"{_GIT_UPD_HOUR:02d}:{_GIT_UPD_MINUTE:02d} "
                "(ff-only origin/main; KeepAlive respawn if HEAD moves)"
            )
            scheduler.add_job(
                _run_git_auto_update,
                "date",
                run_date=datetime.now(AGENT_TZ) + timedelta(seconds=_GIT_UPD_DELAY),
                id="git_auto_update_startup",
                misfire_grace_time=300,
            )
            click.echo(
                f"[Scheduler] Git auto-update once {_GIT_UPD_DELAY}s after startup"
            )
        else:
            click.echo("[Scheduler] Git auto-update disabled (GIT_AUTO_UPDATE=0)")

        # Weekly GitHub backup (Sunday 09:00)
        scheduler.add_job(
            _run_github_backup,
            "cron",
            day_of_week="sun",
            hour=9,
            minute=0,
            id="github_backup",
        )
        click.echo("[Scheduler] GitHub backup weekly Sunday 09:00")

        # Agent job worker: poll every 45 seconds
        scheduler.add_job(
            _run_job_worker,
            "interval",
            seconds=45,
            id="job_worker",
        )
        click.echo("[Scheduler] Job worker every 45s")

        # Print full schedule table
        click.echo("\n── Schedule ─────────────────────────────────────")
        for job in scheduler.get_jobs():
            click.echo(f"  {job.id:<20s}  trigger: {job.trigger}")
        click.echo("─────────────────────────────────────────────────")

        click.echo("\nAgent is running. Watching for files and running scheduled tasks.\n")

        def _shutdown(signum, frame):
            click.echo("\nShutting down...")
            observer.stop()
            scheduler.shutdown(wait=False)
            sys.exit(0)

        signal.signal(signal.SIGINT, _shutdown)
        signal.signal(signal.SIGTERM, _shutdown)

        scheduler.start()

    except (KeyboardInterrupt, SystemExit):
        observer.stop()
        click.echo("Agent stopped.")

    observer.join()


def _run_shopify_poll():
    from src.parsers.shopify_orders import fetch_shopify_orders_api
    from src.db import log_ingestion, job_start, job_finish
    from datetime import datetime, timezone
    import time

    run_id = job_start("shopify_poll")
    ts = datetime.now(timezone.utc).strftime("%Y-%m-%dT%H:%M:%SZ")
    attempts = 0
    last_err = None

    for attempt in range(2):  # try twice
        attempts = attempt + 1
        try:
            result = fetch_shopify_orders_api()
            orders = result.get("orders_fetched", 0)
            inserted = result.get("rows_inserted", 0)
            print(f"[Shopify Poll] {ts} OK — {orders} orders, {inserted} rows"
                  + (f" (retry {attempt})" if attempt else ""))
            log_ingestion(
                filename=f"shopify_api_poll_{ts}",
                file_type="shopify_api",
                rows_total=orders,
                rows_inserted=inserted,
                status="success",
            )
            # Also store the orders THEMSELVES, incrementally.
            #
            # fetch_shopify_orders_api aggregates into (state, channel, month)
            # buckets and discards the orders, which is why no person-level
            # metric existed. A 7-day re-pull is upserted on order_id, so it is
            # idempotent and catches late edits (refunds, cancellations) without
            # re-walking three years of history every two hours.
            stored = 0
            try:
                from datetime import date as _date, timedelta as _td

                from src.shopify_backfill import backfill
                r2 = backfill(since=_date.today() - _td(days=7),
                              progress=lambda _m: None)
                if r2.get("error"):
                    print(f"[Shopify Poll] order store skipped: {r2['error'][:120]}")
                else:
                    stored = r2.get("written", 0)
                    print(f"[Shopify Poll] {ts} stored {stored} order row(s)")
            except Exception as e2:
                # Never fail the tax-relevant aggregate because the analytics
                # side-table had a problem.
                print(f"[Shopify Poll] order store failed (aggregates unaffected): "
                      f"{str(e2)[:120]}")

            job_finish(run_id, "success", f"{orders} orders, {inserted} rows",
                       {"orders": orders, "rows_inserted": inserted,
                        "orders_stored": stored})
            return  # success — done
        except Exception as e:
            last_err = e
            print(f"[Shopify Poll] {ts} attempt {attempts} failed: {e}")
            if attempt == 0:
                time.sleep(30)  # wait 30s before retry

    # Both attempts failed
    err_text = str(last_err)[:500]
    print(f"[Shopify Poll] {ts} FAILED after {attempts} attempts: {err_text}")
    log_ingestion(
        filename=f"shopify_api_poll_{ts}",
        file_type="shopify_api",
        rows_total=0,
        rows_inserted=0,
        status="failed",
        error_message=err_text,
    )
    job_finish(run_id, "fail", err_text)


def _run_daily_analysis():
    """Run physical + economic nexus analysis, gather deadlines, and send
    a SINGLE Telegram summary.  Dedicated threshold-crossed alerts are
    sent only for states that newly exceed their threshold on THIS run.
    """
    from src.db import fetch_all, job_start, job_finish
    from src.engines.physical_nexus import evaluate_physical_nexus
    from src.engines.economic_nexus import evaluate_economic_nexus
    from src.calendar.filing_calendar import (
        get_upcoming_deadlines, mark_overdue_filings,
    )
    from src.alerts.telegram import (
        send_daily_summary,
        send_threshold_crossed,
        send_telegram,
    )
    from datetime import date

    run_id = job_start("daily_analysis")
    try:
        # Snapshot which states had economic nexus BEFORE this run
        prior_econ = {
            r["state_code"]
            for r in fetch_all("nexus_status")
            if r.get("has_economic_nexus")
        }

        # ── Run engines ──
        phys = evaluate_physical_nexus()
        econ = evaluate_economic_nexus()

        # ── Detect newly crossed thresholds ──
        current_econ = set(econ.get("exceeded_threshold", []))
        newly_crossed = sorted(current_econ - prior_econ)

        # Send dedicated alert per newly crossed state
        for sc in newly_crossed:
            detail = econ.get("details", {}).get(sc, {})
            notes = detail.get("action_notes", "")
            send_threshold_crossed(
                sc,
                detail.get("threshold_amount", 0),
                detail.get("threshold_amount_cfg", 100000),
                notes[:200] if notes else f"${detail.get('threshold_amount',0):,.0f}",
            )

        # Persist overdue status before the digest reads the calendar.
        # deadline_check at 09:00 does the same write; doing it here keeps
        # the 08:00 summary and SQL `status='late'` in step the first day
        # a return slips past due.
        mark_overdue_filings()

        # ── Gather deadlines ──
        deadlines = get_upcoming_deadlines()
        today = date.today().isoformat()
        overdue = [d for d in deadlines if d.get("days_overdue")]
        upcoming = [d for d in deadlines if d.get("days_until_due") is not None
                    and not d.get("days_overdue")]

        # ── Gather franchise flags ──
        flags = fetch_all("franchise_tax_flags", {"status": "open"})
        critical_flags = [f for f in flags if f.get("severity") == "critical"]
        warning_flags = [f for f in flags if f.get("severity") == "warning"]

        # ── Count action needed (unregistered nexus states) ──
        nexus_all = fetch_all("nexus_status")
        action_needed = sum(
            1 for n in nexus_all
            if not n.get("is_registered")
            and (n.get("has_physical_nexus") or n.get("has_economic_nexus"))
        )

        # ── Send single summary ──
        send_daily_summary(
            phys_nexus_count=len(phys.get("nexus_states", [])),
            new_phys_states=phys.get("new_nexus_states", []),
            econ_exceeded=sorted(current_econ),
            econ_approaching=econ.get("approaching_threshold", []),
            newly_crossed=newly_crossed,
            critical_flags=critical_flags,
            warning_flags=warning_flags,
            overdue_count=len(overdue),
            upcoming_deadlines=upcoming,
            action_needed=action_needed,
        )

        print(f"[Daily Analysis] Physical: {len(phys.get('nexus_states', []))} states, "
              f"Economic exceeded: {len(current_econ)}, "
              f"Newly crossed: {newly_crossed}, "
              f"Overdue: {len(overdue)}")
        job_finish(run_id, "success",
                   f"Phys {len(phys.get('nexus_states', []))}, Econ {len(current_econ)}, Overdue {len(overdue)}")

    except Exception as e:
        print(f"[Daily Analysis] Error: {e}")
        job_finish(run_id, "fail", str(e)[:500])
        try:
            send_telegram(f"🚨 <b>Daily Analysis Failed</b>\n\n{str(e)[:300]}")
        except Exception:
            pass


def _run_deadline_check():
    """Flip pending → late for past-due filings, then log the queue.

    Individual Telegram alerts are still the daily summary's job
    (08:00). This 09:00 pass is the write that used to be missing:
    get_upcoming_deadlines stamped status="late" only in memory.
    """
    from src.calendar.filing_calendar import (
        get_upcoming_deadlines, mark_overdue_filings,
    )
    from src.db import job_start, job_finish

    run_id = job_start("deadline_check")
    try:
        marked = mark_overdue_filings()
        deadlines = get_upcoming_deadlines()
        overdue = [d for d in deadlines if d.get("days_overdue")]
        print(
            f"[Deadline Check] flipped {marked['flipped']} pending→late "
            f"(today={marked['today']}); {len(overdue)} overdue, "
            f"{len(deadlines)} total (included in daily summary)"
        )
        job_finish(
            run_id, "success",
            f"flipped {marked['flipped']}, overdue {len(overdue)}",
        )
    except Exception as e:
        print(f"[Deadline Check] Error: {e}")
        job_finish(run_id, "fail", str(e)[:500])


def _spapi_refresh_outcome(
    errors: list[str],
    orders_inserted: int | None,
) -> tuple[str, str]:
    """Classify the morning SP-API refresh for job_runs.

    Hard errors (orders exception, inventory, daily sales) stay fail.
    Orders returning 0 rows while inventory/daily succeeded used to record
    success — that hid a stale sales_by_state tax SoT behind a green job.
    """
    if errors:
        return "fail", "; ".join(errors[:3])
    n = 0 if orders_inserted is None else orders_inserted
    if n <= 0:
        return "partial", (
            "Orders wrote 0 sales_by_state rows; inventory + daily sales synced"
        )
    return "success", (
        f"Orders {n} rows + inventory + daily sales synced"
    )


def _run_spapi_refresh():
    from datetime import date, timedelta, datetime, timezone
    from src.amazon_sp.reports import fetch_orders, fetch_inventory
    from src.db import log_ingestion, job_start, job_finish

    from src.rules import SPAPI_INVENTORY_LEDGER_DAYS, amazon_as_of

    run_id = job_start("spapi_refresh")
    ts = datetime.now(timezone.utc).strftime("%Y-%m-%dT%H:%M:%SZ")
    end = amazon_as_of()
    # Month-start of (as_of − 7d): full current month, and the prior month
    # for the first week after a roll so a wiped August can still rebuild.
    # dest-daily upsert + monthly rollup. A 7-day lookback used to REPLACE
    # the whole month in sales_by_state. Jan–Jul are not in this window.
    start = (end - timedelta(days=7)).replace(day=1)
    # The inventory ledger gets its own, wider window. It drives physical nexus
    # and therefore registration decisions, and Amazon restates ledger rows for
    # several days after the fact — a 7-day window silently missed those. The
    # write is an upsert on (source_file, event_date, fc_code, asin, event_type,
    # quantity), so re-pulling a wider range corrects rows instead of
    # duplicating them, and never removes anything older than the window.
    inv_start = end - timedelta(days=SPAPI_INVENTORY_LEDGER_DAYS)
    errors = []
    orders_inserted: int | None = None
    try:
        orders = fetch_orders(start, end)
        orders_inserted = orders.get("rows_inserted", 0)
        print(f"[SP-API] {ts} Orders: {orders_inserted} rows, "
              f"${orders.get('total_gross_sales', 0):,.0f} sales")
        log_ingestion(
            filename=f"spapi_orders_{ts}",
            file_type="amazon_spapi",
            rows_total=orders.get("rows_total", orders_inserted),
            rows_inserted=orders_inserted,
            status="success" if orders_inserted else "partial",
        )
    except Exception as e:
        errors.append(f"Orders: {e}")
        print(f"[SP-API] {ts} Orders error: {e}")
        log_ingestion(
            filename=f"spapi_orders_{ts}",
            file_type="amazon_spapi",
            rows_total=0, rows_inserted=0,
            status="failed",
            error_message=str(e)[:500],
        )
    try:
        inv = fetch_inventory(inv_start, end)
        inserted = inv.get("rows_inserted", 0)
        unknown = inv.get("unknown_fcs") or []
        print(f"[SP-API] {ts} Inventory ({inv_start}→{end}): {inserted} rows, "
              f"states: {inv.get('states_found', [])}")
        if unknown:
            # Never fatal: an unmapped FC is a config gap, not a sync failure.
            # It is surfaced because those events carry no state and are
            # invisible to the physical-nexus engine until the code is mapped.
            print(f"[SP-API] {ts} Inventory: {len(unknown)} unmapped FC code(s): "
                  f"{', '.join(unknown[:10])} — add to config/fc_codes.json, "
                  f"then `inventory-remap-fc --apply`")
        log_ingestion(
            filename=f"spapi_inventory_{ts}",
            file_type="amazon_inventory",
            rows_total=inv.get("rows_total", inserted),
            rows_inserted=inserted,
            status="success",
        )
    except Exception as e:
        errors.append(f"Inventory: {e}")
        print(f"[SP-API] {ts} Inventory error: {e}")
        log_ingestion(
            filename=f"spapi_inventory_{ts}",
            file_type="amazon_inventory",
            rows_total=0, rows_inserted=0,
            status="failed",
            error_message=str(e)[:500],
        )
    # ── Sync sales_daily (Amazon + Shopify) ──
    try:
        from src.sales_daily import sync_amazon_daily
        amz_daily = sync_amazon_daily(days=7)
        print(f"[SP-API] {ts} Daily sales: {amz_daily.get('rows_upserted', 0)} rows, "
              f"${amz_daily.get('total_gross', 0):,.0f}")
    except Exception as e:
        errors.append(f"Daily sales: {e}")
        print(f"[SP-API] {ts} Daily sales error: {e}")

    try:
        from src.sales_daily import sync_shopify_daily
        shop_daily = sync_shopify_daily(days=7)
        print(f"[SP-API] {ts} Shopify daily: {shop_daily.get('rows_upserted', 0)} rows")
    except Exception as e:
        print(f"[SP-API] {ts} Shopify daily error: {e}")

    # FBA returns (30d window, once daily)
    try:
        from src.amazon_sp.reports import fetch_fba_returns
        returns = fetch_fba_returns(start, end)
        print(f"[SP-API] {ts} Returns: {returns.get('rows_inserted', 0)} rows")
    except Exception as e:
        print(f"[SP-API] {ts} Returns error: {e}")

    # Sales & Traffic (7d window)
    try:
        from src.amazon_sp.reports import fetch_sales_traffic
        st = fetch_sales_traffic(start, end)
        print(f"[SP-API] {ts} Sales&Traffic: {st.get('days', 0)} days, {st.get('asins', 0)} ASINs")
    except Exception as e:
        print(f"[SP-API] {ts} Sales&Traffic error: {e}")

    # Reimbursements (90d window, chunked ≤30d). A 30d pull drops July
    # approvals by late August — the cash Dave wants visible on /profit.
    try:
        from src.amazon_sp.reports import fetch_reimbursements
        from src.rules import SPAPI_REIMBURSEMENTS_DAYS
        reimb_start = end - timedelta(days=SPAPI_REIMBURSEMENTS_DAYS)
        reimb = fetch_reimbursements(reimb_start, end)
        print(f"[SP-API] {ts} Reimbursements ({reimb_start}→{end}): "
              f"{reimb.get('rows_parsed', 0)} rows, ${reimb.get('total_amount', 0):,.2f}")
    except Exception as e:
        print(f"[SP-API] {ts} Reimbursements error: {e}")

    # Subscribe & Save (Replenishment API)
    try:
        from src.amazon_sp.replenishment import fetch_seller_metrics, fetch_offer_metrics
        seller = fetch_seller_metrics(weeks=4)
        print(f"[SP-API] {ts} SnS seller: {seller.get('weeks_fetched', 0)} weeks, "
              f"{seller.get('latest_subs', 0):,} subs")
        offers = fetch_offer_metrics()
        print(f"[SP-API] {ts} SnS offers: {offers.get('offers_fetched', 0)} ASINs")
    except PermissionError as e:
        print(f"[SP-API] {ts} SnS: {e}")
    except Exception as e:
        print(f"[SP-API] {ts} SnS error: {e}")

    # Ads are owned by ads_campaigns_sync (05:00), ads_placements_sync
    # (05:15) and ads_search_terms_sync (05:30). A second 14-day ads pull
    # here at 06:00 stacked reports on the same Ads queue, hung this job,
    # and produced nightly SB/SD timeouts. ads_actions (also 06:00) is the
    # sole owner of the recommendation queue.
    print(f"[Ads] {ts} Skipped — dedicated ads_* jobs own the pull")

    # P&L recompute (picks up fresh ad spend + COGS)
    try:
        from src.pnl import compute_pnl
        pnl = compute_pnl(days=14)
        print(f"[PnL] {ts} {pnl.get('days', 0)} days, contribution=${pnl.get('total_contribution', 0):,.0f}")
    except Exception as e:
        print(f"[PnL] {ts} Error: {e}")

    status, message = _spapi_refresh_outcome(errors, orders_inserted)
    # amazon_sku_month at 06:20 owns sales_by_sku. Do not pull All Orders
    # a second time here — fetch_orders already holds that report at 06:00.
    if status == "fail":
        job_finish(run_id, status, message)
        try:
            from src.alerts.telegram import send_telegram
            send_telegram(
                f"🚨 <b>SP-API Sync Failed</b>\n\n"
                + "\n".join(errors[:3])
            )
        except Exception:
            pass
    else:
        job_finish(run_id, status, message)


ALL_ORDERS_JOB_NAMES = ("spapi_refresh", "amazon_sku_month")


def _all_orders_busy(exclude_run_id: str | None = None) -> str | None:
    """job_name of an in-flight All Orders pull, or None.

    fetch_orders (spapi_refresh) and fetch_amazon_skus (amazon_sku_month)
    share GET_FLAT_FILE_ALL_ORDERS_DATA_BY_ORDER_DATE_GENERAL. A second
    request while one is polling is skipped, not stacked.
    """
    try:
        from src.db import get_client
        rows = (
            get_client().table("job_runs")
            .select("id,job_name,status")
            .in_("job_name", list(ALL_ORDERS_JOB_NAMES))
            .eq("status", "running")
            .execute().data
        ) or []
    except Exception:
        return None
    for r in rows:
        if exclude_run_id and r.get("id") == exclude_run_id:
            continue
        return r.get("job_name") or "all_orders"
    return None


def _run_amazon_sku_month():
    """06:20 — current-month Amazon sales_by_sku only.

    start = first day of the Amazon month containing yesterday, end =
    yesterday. That is one ≤30-day All Orders chunk for most of the
    month. Never the backfill-amazon-skus CLI default of 2025-01-01.
    Shopify SKU backfill is not scheduled here.
    """
    from src.db import job_start, job_finish
    from src.rules import current_month_sku_range
    from src.amazon_sp.reports import fetch_amazon_skus

    run_id = job_start("amazon_sku_month")
    start, end = current_month_sku_range()
    # Fail closed if a future edit forgets the current-month window and
    # would request the backfill-amazon-skus default (2025-01-01 → now).
    if start != end.replace(day=1) or (end - start).days > 31:
        job_finish(run_id, "fail",
                   f"refusing non-current-month All Orders window {start}→{end}")
        return

    busy = _all_orders_busy(exclude_run_id=run_id)
    if busy:
        msg = f"skipped — {busy} already pulling All Orders"
        print(f"[Amazon SKU month] {msg}")
        job_finish(run_id, "skipped", msg, stats={
            "start": start.isoformat(), "end": end.isoformat(), "chunks": 0,
        })
        return

    try:
        result = fetch_amazon_skus(start, end)
    except Exception as e:
        print(f"[Amazon SKU month] Failed {start} → {end}: {e}")
        job_finish(run_id, "fail", str(e)[:500], stats={
            "start": start.isoformat(), "end": end.isoformat(),
        })
        return

    chunks = int(result.get("chunks") or 0)
    inserted = int(result.get("rows_inserted") or 0)
    msg = (f"{start} → {end}: {chunks} chunk(s), "
           f"{result.get('unique_skus', 0)} SKUs, {inserted} rows")
    print(f"[Amazon SKU month] {msg}")
    job_finish(run_id, "success", msg, stats={
        "start": start.isoformat(),
        "end": end.isoformat(),
        "chunks": chunks,
        "sku_rows": result.get("sku_rows", 0),
        "unique_skus": result.get("unique_skus", 0),
        "rows_inserted": inserted,
    })


def _run_paid_ads_freshness():
    """Nudge when a Shopify paid-ads CSV export has gone stale."""
    from src.alerts.paid_ads_freshness import run_paid_ads_freshness_check

    try:
        run_paid_ads_freshness_check()
    except Exception as e:
        print(f"[Paid Ads Freshness] error: {e}")


def _run_source_monitoring():
    from src.intelligence.monitor import run_monitoring_cycle

    try:
        result = run_monitoring_cycle()
        changes = result.get("changes_detected", 0)
        print(f"[Source Monitor] Checked {result.get('sources_checked', 0)} sources, "
              f"{changes} changes detected")

        if changes > 0:
            try:
                from src.alerts.telegram import send_telegram
                details = [d for d in result.get("details", []) if d.get("change")]
                urls = ", ".join(d.get("title", d.get("url", "?"))[:40] for d in details[:5])
                send_telegram(
                    f"<b>Source Monitor</b>\n"
                    f"{changes} monitored source(s) changed:\n{urls}\n\n"
                    f"Run <code>python -m src.main research-tasks</code> to review.",
                    "source_monitor",
                )
            except Exception:
                pass
    except Exception as e:
        print(f"[Source Monitor] Error: {e}")


def _run_ppc_export_publish():
    """07:30 — build the PPC Command Brief and store it for the dashboard.

    Publishing is the only reason this runs on a timer. The brief itself is
    always built live when the operator asks for it; this copy exists so the
    dashboard's Copy/Download buttons work from Vercel, where there is no
    Python interpreter. It is never sent to Telegram.
    """
    from src.db import job_start, job_finish

    run_id = job_start("ppc_export_publish")
    try:
        from src.amazon_ads.export_brief import (build_brief, build_prompt,
                                                 gather, grade_window,
                                                 publish_brief)
        d = gather(days=7)
        brief = build_brief(d)
        res = publish_brief(d, brief, build_prompt(brief))
        grade = grade_window(d)
        if res.get("published"):
            print(f"[PPC export] Published — grade {grade.score:.0f}/100 "
                  f"({grade.letter})")
            job_finish(run_id, "success",
                       f"score {grade.score:.0f} {grade.letter}")
        else:
            print(f"[PPC export] NOT published: {res.get('error')}")
            job_finish(run_id, "partial", str(res.get("error"))[:200])
    except Exception as e:
        print(f"[PPC export] FAILED: {e}")
        job_finish(run_id, "fail", str(e)[:200])


def _run_health_ping():
    """Daily agent health check-in.

    Wrapped in job_start/job_finish like every other job, but note the ordering
    problem it creates: this run appears in job_runs, so the health check is
    itself part of the history it reports on. That is fine — it never reads its
    own job name — but it is why the check-in reports ADS job freshness
    specifically rather than "last job of any kind".
    """
    from src.db import job_start, job_finish
    run_id = job_start("health_ping")
    try:
        from src.health import run_health_ping
        r = run_health_ping(send=True)
        state = ("healthy" if r["healthy"] else f"faults: {'; '.join(r['faults'])}")
        if r["sent"]:
            print(f"[Health] Sent — {state}")
        else:
            print(f"[Health] Not sent ({r['reason']}) — {state}")
        job_finish(run_id, "success", f"{r['severity']}; sent={r['sent']}")
    except Exception as e:
        print(f"[Health] FAILED: {e}")
        job_finish(run_id, "fail", str(e)[:200])


def _run_daily_digest():
    """Send morning sales digest via Telegram."""
    from src.db import job_start, job_finish
    run_id = job_start("daily_digest")
    try:
        from src.alerts.daily_digest import send_digest
        result = send_digest()
        if result.get("sent"):
            print("[Digest] Sent")
            job_finish(run_id, "success", "Sent")
        elif result.get("reason"):
            print(f"[Digest] Skipped: {result['reason']}")
            job_finish(run_id, "success", f"Skipped: {result['reason']}")
    except Exception as e:
        print(f"[Digest] Error: {e}")
        job_finish(run_id, "fail", str(e)[:500])


def _run_inventory_ledger_backfill():
    """Weekly deeper inventory-ledger re-pull.

    Amazon settles some ledger rows after the daily window has already moved
    past them. This re-pulls a wider range so those land. It is an UPSERT on the
    ledger's natural key — it corrects and adds, and cannot remove events older
    than the window, which is what protects the 2024/2025 history the
    physical-nexus engine reads.
    """
    from datetime import date, timedelta
    from src.db import job_start, job_finish
    from src.amazon_sp.reports import fetch_inventory
    from src.rules import SPAPI_INVENTORY_BACKFILL_DAYS, amazon_as_of

    run_id = job_start("inventory_ledger_backfill")
    end = amazon_as_of()
    start = end - timedelta(days=SPAPI_INVENTORY_BACKFILL_DAYS)
    try:
        r = fetch_inventory(start, end)
    except Exception as e:
        print(f"[Inventory Ledger] Backfill failed: {e}")
        job_finish(run_id, "fail", str(e)[:500])
        return

    unknown = r.get("unknown_fcs") or []
    msg = (f"{SPAPI_INVENTORY_BACKFILL_DAYS}d ledger: {r.get('rows_inserted', 0)} rows, "
           f"{len(r.get('states_found') or [])} states")
    if unknown:
        msg += f", {len(unknown)} unmapped FC"
    print(f"[Inventory Ledger] {msg}")
    job_finish(run_id, "success", msg)


def _run_ledger_summary():
    """Nightly GET_LEDGER_SUMMARY_VIEW_DATA for tax inventory $ at COGS.

    Amazon's summary report lags 1–3 days; amazon_as_of() is the newest
    closed Amazon day. Pull that single complete day (upsert-only).
    """
    from src.db import job_start, job_finish
    from src.amazon_sp.ledger_summary import fetch_ledger_summary
    from src.rules import amazon_as_of

    run_id = job_start("ledger_summary_daily")
    end = amazon_as_of()
    start = end
    try:
        r = fetch_ledger_summary(start, end)
    except Exception as e:
        print(f"[Ledger Summary] failed: {e}")
        job_finish(run_id, "fail", str(e)[:500])
        return

    msg = (
        f"{start}: {r.get('rows_inserted', 0)} rows, "
        f"${r.get('total_cogs', 0):,.2f} COGS, "
        f"{len(r.get('states_found') or [])} states"
    )
    unknown = r.get("unknown_fcs") or []
    if unknown:
        msg += f", {len(unknown)} unmapped FC"
    print(f"[Ledger Summary] {msg}")
    job_finish(run_id, "success", msg)


def _run_inventory_sync():
    """Daily inventory sync: FBA summaries + restock + AWD + velocity."""
    from src.db import job_start, job_finish
    run_id = job_start("inventory_sync")
    errors = []
    skip_notes: list[str] = []
    try:
        from src.inventory.sync import sync_all
        from src.inventory.freshness import collect_skip_reasons
        results = sync_all()
        for name in [
            "fba_summaries", "awd", "restock", "planning",
            "inbound_shipments", "awd_replenishments", "awd_inbound",
        ]:
            r = results.get(name, {})
            if "error" in r:
                print(f"[Inventory] {name}: {r['error'][:100]}")
                errors.append(f"{name}: {r['error'][:80]}")
            elif r.get("skipped"):
                print(f"[Inventory] {name}: SKIPPED — {r.get('skip_reason')}")
            else:
                print(f"[Inventory] {name}: {r.get('rows_total', r.get('shipments_found', r.get('orders_found', 0)))} rows")
        skip_notes.extend(collect_skip_reasons(results))
    except Exception as e:
        print(f"[Inventory Sync] Error: {e}")
        errors.append(str(e)[:200])

    try:
        from src.inventory.holiday_surge import approaching_peak, compute_surge_from_sales
        if approaching_peak():
            surge = compute_surge_from_sales()
            print(f"[HolidaySurge] {surge.get('skus_updated', 0)} SKUs, "
                  f"{surge.get('surge_skus', 0)} surge")
    except Exception as e:
        print(f"[HolidaySurge] Error: {e}")
        errors.append(f"holiday_surge: {e}")

    try:
        from src.inventory.velocity import compute_velocity
        r = compute_velocity(amazon_days=100, shopify_days=100)
        print(f"[Velocity] {r['skus']} SKUs, mult={r['avg_forward_mult']:.2f}")
    except Exception as e:
        print(f"[Velocity] Error: {e}")
        errors.append(f"velocity: {e}")

    try:
        from src.inventory.rate_signals import sync_sku_signals
        sig = sync_sku_signals()
        if sig.get("skipped"):
            print(f"[RateSignals] SKIPPED — {sig.get('skip_reason')}")
            skip_notes.append(f"sku_signals: {sig.get('skip_reason')}")
        else:
            print(f"[RateSignals] {sig.get('skus', 0)} SKUs, "
                  f"account receive ~{sig.get('account_receive_days')}d "
                  f"(n={sig.get('account_receive_n', 0)})")
    except Exception as e:
        print(f"[RateSignals] Error: {e}")
        errors.append(f"rate_signals: {e}")

    if errors:
        job_finish(run_id, "fail", "; ".join(errors))
    else:
        msg = "FBA + AWD + restock + velocity synced"
        if skip_notes:
            msg += " | skipped: " + "; ".join(skip_notes)
        job_finish(run_id, "success", msg)


def _run_3pl_sync():
    """Daily 3PL inventory sync from Ship Sidekick."""
    from src.db import job_start, job_finish
    run_id = job_start("3pl_sync")
    try:
        from src.shipsidekick.client import sync_3pl
        r = sync_3pl()
        status = r.get("status") or "success"
        msg = r.get("message") or (
            f"{r['rows_total']} SKUs, {r['rows_inserted']} upserted"
        )
        print(f"[3PL] {status}: {msg}")
        job_finish(run_id, status, msg)
    except Exception as e:
        print(f"[3PL] Error: {e}")
        job_finish(run_id, "fail", str(e)[:500])


# ── Amazon Ads scheduled jobs ────────────────────────────────
#
# Each runs in its own APScheduler job and swallows its own exceptions, so one
# failing never prevents the others from firing. Each writes its own job_runs
# row; the /ppc "last sync" label reads the newest of them.


def _ads_alert(subject: str, detail: str) -> None:
    """Telegram on hard failure only — partials are normal and stay in the log."""
    try:
        from src.config import settings
        if not settings.telegram_enabled:
            return
        from src.alerts.telegram import send_telegram
        send_telegram(f"⚠️ {subject}\n\n{detail[:600]}")
    except Exception as e:  # never let alerting break the job
        print(f"[Ads] Telegram alert failed: {e}")


_SCHEDULER = None
# 2026-08-24: campaigns held the lock past 08:51 ET. Three 20-minute
# retries from 05:15 would have given up at 06:15. Eighteen covers a
# Sunday backfill that overruns the 05:00 window (6 hours).
_ADS_RETRY_MAX = 18
_ADS_RETRY_SECONDS = 20 * 60


def _schedule_ads_retry(fn, retry: int, job_id: str) -> None:
    """Re-run a skipped ads job after the lock holder finishes."""
    if _SCHEDULER is None or retry >= _ADS_RETRY_MAX:
        return
    from datetime import datetime, timedelta
    from src.rules import AGENT_TZ
    when = datetime.now(AGENT_TZ) + timedelta(seconds=_ADS_RETRY_SECONDS)
    rid = f"{job_id}_retry_{retry + 1}"
    try:
        _SCHEDULER.add_job(
            fn, "date", run_date=when, id=rid,
            misfire_grace_time=3600, replace_existing=True)
        print(f"[Ads] Scheduled {rid} at {when.isoformat(timespec='minutes')}")
    except Exception as e:
        print(f"[Ads] Could not schedule {rid}: {e}")


def _run_ads_sync_job(job_name: str, *, days: int, campaigns_only: bool = False,
                      search_terms_only: bool = False,
                      placements_only: bool = False, label: str,
                      sb_sd_days: int | None = None,
                      skip_existing_search_term_weeks: bool = False,
                      newest_first_search_terms: bool = False,
                      retry: int = 0) -> str:
    """Shared body for the ads sync jobs. Returns the settled status."""
    from src.db import job_start, job_finish
    from src.amazon_ads.reports import AdsSyncBusy, sync_ads

    status = "fail"
    run_id = job_start(job_name)
    finished = False
    try:
        result = sync_ads(days=days, campaigns_only=campaigns_only,
                          search_terms_only=search_terms_only,
                          placements_only=placements_only,
                          sb_sd_days=sb_sd_days,
                          skip_existing_search_term_weeks=skip_existing_search_term_weeks,
                          newest_first_search_terms=newest_first_search_terms)
        status, message = _ads_sync_outcome(result, days)
        camp = result.get("campaigns")
        if isinstance(camp, dict):
            for t, v in (camp.get("by_type") or {}).items():
                print(f"[Ads {label}] {t}: {v['rows']} rows, ${v['spend']:,.2f} spend, "
                      f"{v['clicks']:,} clicks, {'ok' if v['ok'] else 'FAILED'}")
        for key in ("campaigns", "search_terms", "placements"):
            val = result.get(key)
            if isinstance(val, dict):
                if "error" in val:
                    print(f"[Ads {label}] {key}: ERROR — {val['error'][:120]}")
                else:
                    print(f"[Ads {label}] {key}: {val.get('rows', 0)} rows, "
                          f"{val.get('inserted', 0)} inserted, "
                          f"{val.get('chunks', 0)} chunks, {len(val.get('errors', []))} chunk errors")
        print(f"[Ads {label}] {status}: {message}")

        job_finish(run_id, status, message,
                   stats=result.get("campaigns") or result.get("search_terms"))
        finished = True
        # A chunk that timed out and will be re-fetched tomorrow is routine and
        # stays in the log. Two things do get a push: a total failure, and an ad
        # product that dropped out entirely — the latter under-reports account
        # spend against the Amazon console for as long as it goes unnoticed, which
        # is exactly how the SP-only sync hid ~6% of spend.
        lost_products = []
        if isinstance(camp, dict):
            for p in (camp.get("products_failed") or []):
                v = (camp.get("by_type") or {}).get(p) or {}
                if not v.get("rows") and not v.get("spend"):
                    lost_products.append(p)
        if status == "fail":
            _ads_alert(f"Ads {label} sync failed", message)
        elif lost_products:
            _ads_alert(f"Ads {label} sync partial — {'+'.join(lost_products)} missing",
                       message)
            # Do not wait for tomorrow. A 20-minute SB/SD-only heal with
            # 1-day chunks usually lands what the jumbo nightly chunk lost.
            if any(p in ("SB", "SD") for p in lost_products):
                _schedule_ads_retry(_run_ads_sb_sd_heal, 0, "ads_sb_sd_heal")
        return status
    except AdsSyncBusy as e:
        print(f"[Ads {label}] Skipped: {e}")
        status = "skipped"
        job_finish(run_id, status, str(e)[:500])
        finished = True
        nxt = retry + 1
        if job_name == "ads_campaigns_sync":
            _schedule_ads_retry(
                lambda n=nxt: _run_ads_campaigns_sync(retry=n), retry, job_name)
        else:
            _schedule_ads_retry(
                lambda n=nxt: _run_ads_sync_job(
                    job_name, days=days, campaigns_only=campaigns_only,
                    search_terms_only=search_terms_only,
                    placements_only=placements_only, label=label,
                    sb_sd_days=sb_sd_days,
                    skip_existing_search_term_weeks=skip_existing_search_term_weeks,
                    newest_first_search_terms=newest_first_search_terms,
                    retry=n),
                retry, job_name)
        return status
    except Exception as e:
        print(f"[Ads {label}] Failed: {e}")
        status = "fail"
        job_finish(run_id, status, str(e)[:500])
        finished = True
        _ads_alert(f"Ads {label} sync failed", str(e))
        return status
    finally:
        # SIGTERM / scheduler shutdown used to leave status=running for days
        # ("stuck running since …"). Mark those so the next night is not
        # mistaken for an in-flight pull.
        if not finished:
            job_finish(run_id, "fail", "interrupted before job_finish")
    return status


def _run_sqp_sync():
    """Weekly Brand Analytics SQP pull — feeds the PPC organic-rank gate.

    Scheduled after Amazon publishes the prior Sunday-Saturday week. A missing
    Brand Analytics role fails loudly rather than writing zero rows, because a
    silent no-op is indistinguishable from "no data this week" and would leave
    the gate permanently blind.
    """
    from src.amazon_sp.sqp import BrandAnalyticsRoleError, sync_sqp
    from src.db import job_finish, job_start, log_ingestion

    run_id = job_start("sqp_sync")
    try:
        r = sync_sqp()
    except BrandAnalyticsRoleError as e:
        print(f"[SQP] Role/permission problem: {e}")
        job_finish(run_id, "fail", str(e)[:500])
        log_ingestion(filename="sqp_spapi", file_type="brand_analytics_sqp",
                      rows_total=0, rows_inserted=0, status="failed",
                      error_message=str(e)[:500])
        _ads_alert("SQP sync blocked — Brand Analytics role", str(e))
        return
    except Exception as e:
        print(f"[SQP] Failed: {e}")
        job_finish(run_id, "fail", str(e)[:500])
        log_ingestion(filename="sqp_spapi", file_type="brand_analytics_sqp",
                      rows_total=0, rows_inserted=0, status="failed",
                      error_message=str(e)[:500])
        return

    written = r.get("written", 0)
    msg = (f"{r['period']} {r['start']}→{r['end']}: {len(r['rows'])} keyword(s), "
           f"{written} written, {len(r.get('asins') or [])} ASIN(s)")
    if r.get("errors"):
        msg += f", {len(r['errors'])} batch error(s)"
    print(f"[SQP] {msg}")
    log_ingestion(filename=f"sqp_spapi_{r['start']}_{r['end']}",
                  file_type="brand_analytics_sqp",
                  rows_total=r.get("parsed", 0), rows_inserted=written,
                  status="success" if not r.get("errors") else "partial",
                  warnings=r.get("warnings") or None)
    job_finish(run_id, "partial" if r.get("errors") else "success", msg)

    # Fresh ranks only reach the plan when recommendations are rebuilt.
    from src.amazon_ads.organic_rank import load_config
    if written and (load_config().get("sqp_auto") or {}).get(
            "on_success_refresh_ads_actions", False):
        try:
            _run_ads_actions()
            print("[SQP] Rebuilt ads recommendations so the gate sees new ranks")
        except Exception as e:
            print(f"[SQP] Recommendation refresh failed: {e}")


def _run_ads_campaigns_sync(retry: int = 0):
    """05:00 — campaign dailies for the KPI cards and trend chart.

    SP upserts the last 7 closed days on (date, campaign_id), same window
    as SB/SD. SB/SD still use 1-day chunks so one timed-out report cannot
    wipe the whole week. After campaigns release the lock this job runs
    placements and search terms itself; those crons are a backup.
    """
    from src.rules import ADS_SB_SD_DAILY_DAYS
    status = _run_ads_sync_job(
        "ads_campaigns_sync", days=7, campaigns_only=True,
        label="campaigns", sb_sd_days=ADS_SB_SD_DAILY_DAYS, retry=retry)
    # A skipped run is Sunday-backfill (or a late leftover) still holding
    # the lock. Chaining here would just skip twice more. The retry of
    # this wrapper re-chains once campaigns actually run.
    if status == "skipped":
        return
    _run_ads_placements_sync()
    _run_ads_search_terms_sync()


def _run_ads_sb_sd_heal():
    """13:00 (and post-partial retry) — fill SP-only days with SB/SD.

    Does nothing when the lookback already has Brands/Display on every day
    that has Sponsored Products. Safe to run while other ads jobs are idle;
    if the lock is held it skips via AdsSyncBusy like every other ads job.
    """
    from src.amazon_ads.heal import sync_missing_sb_sd
    from src.amazon_ads.reports import AdsSyncBusy
    from src.db import job_finish, job_start

    run_id = job_start("ads_sb_sd_heal")
    try:
        out = sync_missing_sb_sd(on_progress=lambda m: print(f"[Ads heal] {m}"))
    except AdsSyncBusy as e:
        print(f"[Ads heal] Skipped: {e}")
        job_finish(run_id, "skipped", str(e)[:500])
        _schedule_ads_retry(_run_ads_sb_sd_heal, 0, "ads_sb_sd_heal")
        return
    except Exception as e:
        print(f"[Ads heal] Failed: {e}")
        job_finish(run_id, "fail", str(e)[:500])
        _ads_alert("Ads SB/SD heal failed", str(e))
        return

    if not out.get("healed"):
        msg = out.get("reason") or "no gaps"
        print(f"[Ads heal] {msg}")
        job_finish(run_id, "success", msg)
        return

    camp = (out.get("result") or {}).get("campaigns") or {}
    bad = [p for p in (camp.get("products_failed") or [])
           if not ((camp.get("by_type") or {}).get(p) or {}).get("rows")]
    status = "partial" if bad else "success"
    msg = (f"healed {','.join(out.get('products') or [])} "
           f"{out['window']['start']}→{out['as_of']}")
    if bad:
        msg += f" — {'+'.join(bad)} still missing"
        _ads_alert(f"Ads SB/SD heal partial — {'+'.join(bad)} still missing", msg)
    print(f"[Ads heal] {status}: {msg}")
    job_finish(run_id, status, msg, stats=camp if isinstance(camp, dict) else None)


def _run_ads_search_terms_sync():
    """05:30 — 7 days of search terms in 7-day chunks (one chunk, 90-min cap)."""
    _run_ads_sync_job("ads_search_terms_sync", days=7, search_terms_only=True,
                      label="search terms")


def _run_ads_search_terms_backfill():
    """Sunday 03:30 — 90 days of search terms in existing 7-day chunks.

    fetch_search_terms upserts each 7-day chunk as it returns so a later
    chunk kill/timeout still leaves earlier weeks in ads_search_terms_daily.
    Newest-first + skip weeks already stored so a STOP on 425 keeps recent
    coverage and the next Sunday (or one-shot) resumes. Weekday
    ads_search_terms_sync stays 7 closed days. This job is the scheduled
    90d search-term pull — not chained from nightly campaigns, not a
    campaigns-only retry, must not run in CI or on merge.
    """
    _run_ads_sync_job(
        "ads_search_terms_backfill", days=90, search_terms_only=True,
        skip_existing_search_term_weeks=True,
        newest_first_search_terms=True,
        label="search terms 90d")


def _run_ads_campaigns_backfill():
    """Sunday 03:00 — 90 days of SP campaigns for long trends.

    SB/SD stay on the short backfill window so this job cannot still be
    holding the process lock when the 05:00 daily jobs fire.
    """
    from src.rules import ADS_SB_SD_BACKFILL_DAYS
    _run_ads_sync_job("ads_campaigns_backfill", days=90, campaigns_only=True,
                      label="campaigns 90d", sb_sd_days=ADS_SB_SD_BACKFILL_DAYS)


def _run_ads_placements_sync():
    """05:15 — placement (Top of Search / Detail Page / Other) performance.

    No-ops with a clear message until supabase/migration_ads_placement.sql has
    been run, so it never alerts nightly for a setup step.
    """
    _run_ads_sync_job("ads_placements_sync", days=14, placements_only=True,
                      label="placements")


def _run_ads_actions():
    """06:00 — regenerate the Actions queue from the freshest search terms.

    This is what makes the dashboard's "Generate Recommendations" button
    optional: the queue is already rebuilt before the user looks at it.
    """
    from src.db import job_start, job_finish
    from src.amazon_ads.actions_engine import generate_recommendations
    from src.amazon_ads.strategy import account_target_acos

    run_id = job_start("ads_actions")
    try:
        # Break-even derived from the same COGS/fee inputs the CLI and the
        # strategy layer use — never a hardcoded target. A flat 30% was
        # cutting terms that are profitable at this account's ~37.7%
        # break-even, and left the dashboard reading a target the engine had
        # not actually applied.
        target_acos, target_basis = account_target_acos()
        recs = generate_recommendations(target_acos=target_acos, lookback_days=7)
    except Exception as e:
        print(f"[Ads actions] Failed: {e}")
        job_finish(run_id, "fail", str(e)[:500])
        _ads_alert("PPC action generation failed", str(e))
        return

    if not recs:
        msg = "No recommendations — no search term data in the last 7 days"
        print(f"[Ads actions] {msg}")
        # Still record the target so the dashboard's break-even line stays
        # current on a quiet night instead of falling back to config.
        job_finish(run_id, "success", msg,
                   stats={"count": 0, "target_acos": target_acos,
                          "target_basis": target_basis, "days": 7})
        return

    by_priority: dict[str, int] = {}
    for r in recs:
        by_priority[r["priority"]] = by_priority.get(r["priority"], 0) + 1
    msg = (f"{len(recs)} recs (" + ", ".join(f"{k} {v}" for k, v in sorted(by_priority.items()))
           + f") at {target_acos:.1f}% target")
    print(f"[Ads actions] {msg} [{target_basis}]")
    job_finish(run_id, "success", msg,
               stats={"count": len(recs), "by_priority": by_priority,
                      "target_acos": target_acos, "target_basis": target_basis,
                      "days": 7})


def _run_ads_outcomes():
    """07:00 — snapshot outcomes for actions whose +7/+14/+30 day has closed.

    Observation only: records what happened after each applied or dismissed
    action so action types can later be compared by contribution. Writes no
    bids and trains nothing.
    """
    from src.db import job_start, job_finish
    from src.amazon_ads.learning import snapshot_outcomes

    run_id = job_start("ads_outcomes")
    try:
        r = snapshot_outcomes()
    except Exception as e:
        print(f"[Ads outcomes] Failed: {e}")
        job_finish(run_id, "fail", str(e)[:500])
        return

    if r.get("skipped"):
        print(f"[Ads outcomes] Skipped: {r['skipped']}")
        job_finish(run_id, "success", f"skipped: {r['skipped']}")
        return
    if r.get("error"):
        print(f"[Ads outcomes] Error: {r['error']}")
        job_finish(run_id, "fail", r["error"][:400])
        return

    msg = (f"{r['written']} snapshot(s) written, {r['skipped_not_due']} not due, "
           f"{r['already_present']} already recorded ({r['decisions_considered']} decisions)")
    print(f"[Ads outcomes] {msg}")
    job_finish(run_id, "success", msg, stats=r)


def _run_pnl_sync():
    """06:45 — store daily contribution for every Amazon day.

    contribution = gross_sales - referral - fba - ad_spend - cogs
    Runs after the ads sync (05:00) and SP-API sales refresh (06:00) so the
    inputs are current. Amazon settlement is fetched for fee detail and the
    payout reconciliation figure, never as the daily grain.
    """
    from src.db import job_start, job_finish
    from src.pnl import compute_pnl

    run_id = job_start("pnl_sync")
    try:
        # 90 days covers the ads_campaigns_daily span this account actually
        # has. The profit table now reads every stored day (week / month /
        # year rollups), so the writer has to keep more than a month or the
        # lookback controls are empty. Do not jump to 365: days before ads
        # coverage would store $0 spend and overstate contribution.
        r = compute_pnl(days=90)
        from src.pnl_monthly import compute_monthly_pnl
        monthly = compute_monthly_pnl(persist=True)
    except Exception as e:
        print(f"[P&L] Failed: {e}")
        job_finish(run_id, "fail", str(e)[:500])
        _ads_alert("Contribution P&L sync failed", str(e))
        return

    if not r.get("rows") and not monthly.get("month_count"):
        job_finish(run_id, "success", "no data in window")
        return

    if r.get("rows"):
        msg = (f"{r['days']}d: sales ${r['total_sales']:,.0f} - fees ${r['total_fees']:,.0f} "
               f"- ads ${r['total_ads']:,.0f} - COGS ${r['total_cogs']:,.0f} "
               f"= ${r['total_contribution']:,.0f}")
    else:
        msg = "no daily rows"
    if monthly.get("month_count"):
        msg += (f"; monthly {monthly['coverage_min']}→{monthly['coverage_max']} "
                f"${monthly['total_contribution']:,.0f}")
    settled = r.get("settled_days", 0)
    estimated = r.get("estimated_days", 0)
    print(f"[P&L] {msg} ({settled} settled, {estimated} estimated)")
    job_finish(run_id, "success", msg, stats={
        "days": r.get("days", 0), "sales": r.get("total_sales", 0),
        "fees": r.get("total_fees", 0), "ads": r.get("total_ads", 0),
        "cogs": r.get("total_cogs", 0),
        "contribution": r.get("total_contribution", 0),
        "settled_days": settled,
        "monthly_months": monthly.get("month_count", 0),
    })


def _run_github_backup():
    """Weekly GitHub backup to backup/* branch."""
    try:
        from src.maintenance.github_backup import run_backup
        r = run_backup()
        if r["status"] == "success":
            print(f"[Backup] {r['message']}")
        elif r["status"] == "nothing_to_backup":
            print("[Backup] Nothing to back up")
        else:
            print(f"[Backup] {r['status']}: {r.get('error', '')[:200]}")
    except Exception as e:
        print(f"[Backup] Error: {e}")


def _run_git_auto_update():
    """ff-only pull of origin/main; exit so KeepAlive loads new code.

    Dirty, diverged, or off-main aborts without touching the tree. A
    no-op when already at origin/main. job_runs records every outcome;
    Telegram fires only when a human has to unstick the checkout.
    """
    from src.db import job_finish, job_start
    from src.maintenance.git_auto_update import (
        alert_if_needed, request_process_exit, run_auto_update,
    )

    run_id = job_start("git_auto_update")
    try:
        result = run_auto_update()
    except Exception as e:
        print(f"[git_auto_update] Error: {e}")
        job_finish(run_id, "fail", str(e)[:500])
        alert_if_needed({"status": "error", "error": str(e)})
        return

    status = result.get("status")
    msg = result.get("message") or result.get("error") or status
    print(f"[git_auto_update] {status}: {msg}")

    if status in {"up_to_date", "updated", "disabled", "skipped", "dry_run"}:
        job_finish(run_id, "success", str(msg)[:500])
    else:
        job_finish(run_id, "fail", str(msg)[:500])
        alert_if_needed(result)

    if result.get("restart"):
        print("[git_auto_update] HEAD changed — exiting so launchd "
              "KeepAlive respawns with the new checkout")
        request_process_exit()


def _run_cpa_exports():
    """Generate CPA exports (inventory presence + triage) and upload to Storage."""
    from src.db import job_start, job_finish
    run_id = job_start("cpa_exports")
    cpa_errors = []
    try:
        from src.exports.inventory_presence import (
            build_markdown, build_csv, build_pdf, build_metadata, upload_exports,
        )
        md = build_markdown()
        csv_content = build_csv()
        pdf_bytes = build_pdf()
        meta = build_metadata()
        results = upload_exports(md, csv_content, pdf_bytes, meta)
        ok = sum(1 for v in results.values() if v)
        print(f"[CPA Export] Inventory presence: {ok}/{len(results)} files uploaded")
    except Exception as e:
        print(f"[CPA Export] Inventory presence error: {e}")
        cpa_errors.append(f"inv presence: {e}")

    try:
        from src.exports.registration_triage import (
            build_markdown as triage_md, build_csv as triage_csv,
            build_triage_rows,
        )
        from src.db import upload_to_storage
        import json as _json
        from datetime import datetime, timezone

        rows = triage_rows = build_triage_rows()
        md = triage_md()
        csv_c = triage_csv()
        meta_t = {
            "generated_at": datetime.now(timezone.utc).isoformat(),
            "state_count": len(rows),
            "buckets": {},
        }
        from collections import Counter
        for b, cnt in Counter(r["triage_bucket"] for r in rows).items():
            meta_t["buckets"][b] = cnt

        upload_to_storage("cpa-exports", "registration-triage/latest.md",
                          md.encode("utf-8"), "text/markdown")
        upload_to_storage("cpa-exports", "registration-triage/latest.csv",
                          csv_c.encode("utf-8"), "text/csv")
        upload_to_storage("cpa-exports", "registration-triage/meta.json",
                          _json.dumps(meta_t, indent=2).encode("utf-8"), "application/json")
        print(f"[CPA Export] Registration triage: 3 files uploaded")
    except Exception as e:
        print(f"[CPA Export] Registration triage error: {e}")
        cpa_errors.append(f"reg triage: {e}")

    try:
        from src.exports.economic_nexus_audit import build_audit, upload_exports as upload_econ
        audit = build_audit()
        results = upload_econ(audit)
        ok = sum(1 for v in results.values() if v)
        exceeded = [s["state_code"] for s in audit["states"] if s["status"] == "exceeded"]
        print(f"[CPA Export] Economic nexus audit: {ok}/{len(results)} files, {len(exceeded)} exceeded")
    except Exception as e:
        print(f"[CPA Export] Economic nexus audit error: {e}")
        cpa_errors.append(f"econ audit: {e}")

    if cpa_errors:
        job_finish(run_id, "fail", "; ".join(cpa_errors))
    else:
        job_finish(run_id, "success", "All CPA exports uploaded")


def _payload_flag(value, default: bool = False) -> bool:
    """Parse a JSON / Slack / Dana payload flag into a bool."""
    if value is None:
        return default
    if isinstance(value, str):
        return value.strip().lower() in ("1", "true", "yes", "on")
    return bool(value)


def _ads_enqueue_kwargs(payload: dict | None) -> dict:
    """Resolve ads_sync / ppc_sync agent_jobs payload.

    Dashboard / Dana catch-up is a short campaigns refresh: 7 days,
    campaigns_only. Search terms (~90 min poll cap) and placements stay
    on their scheduled jobs unless the payload explicitly asks for them.

    A 90-minute search-term pull would occupy job_worker (APScheduler
    max_instances=1) and stall inventory_sync / other queued agent_jobs.

    Explicit overrides:
      campaigns_only: false — full suite (old enqueue behavior)
      search_terms_only     — search-term report only
      placements_only       — placements only
    """
    payload = payload or {}
    raw_days = payload.get("days", 7)
    days = int(7 if raw_days in (None, "") else raw_days)

    search_terms_only = _payload_flag(payload.get("search_terms_only"), False)
    placements_only = _payload_flag(payload.get("placements_only"), False)
    if "campaigns_only" in payload and payload["campaigns_only"] is not None:
        campaigns_only = _payload_flag(payload["campaigns_only"], True)
    else:
        campaigns_only = not (search_terms_only or placements_only)

    only_flags = (campaigns_only, search_terms_only, placements_only)
    if sum(bool(f) for f in only_flags) > 1:
        raise ValueError(
            "campaigns_only, search_terms_only and placements_only "
            "are mutually exclusive")

    return {
        "days": days,
        "campaigns_only": campaigns_only,
        "search_terms_only": search_terms_only,
        "placements_only": placements_only,
        "label": "dashboard-enqueue",
    }


def _run_job_worker():
    """Poll agent_jobs for pending work and execute."""
    from src.db import get_client

    try:
        client = get_client()
        # Claim oldest pending job
        result = client.table("agent_jobs") \
            .select("*") \
            .eq("status", "pending") \
            .order("created_at") \
            .limit(1) \
            .execute()

        if not result.data:
            return

        job = result.data[0]
        job_id = job["id"]
        job_type = job["job_type"]

        # Mark as running
        from datetime import datetime, timezone
        client.table("agent_jobs") \
            .update({"status": "running", "started_at": datetime.now(timezone.utc).isoformat()}) \
            .eq("id", job_id) \
            .execute()

        print(f"[Job Worker] Running job {job_id}: {job_type}")

        try:
            payload = job.get("payload") or {}
            if job_type in ("export_cpa", "export_triage", "export_economic_audit"):
                _run_cpa_exports()  # runs all three: inventory, triage, economic
            elif job_type == "spapi_refresh":
                # Dashboard /api/spapi-refresh enqueues this. Honour payload days.
                from datetime import date as d, timedelta
                days = int(payload.get("days", 30))
                end = d.today() - timedelta(days=1)
                start = end - timedelta(days=days)
                from src.amazon_sp.reports import fetch_orders, fetch_inventory
                print(f"[Job Worker] spapi_refresh {start} → {end}")
                fetch_orders(start, end)
                fetch_inventory(start, end)
            elif job_type in ("ads_sync", "ppc_sync"):
                # Honour payload flags. Default is 7d campaigns-only so a
                # dashboard / Dana catch-up cannot start search terms.
                kwargs = _ads_enqueue_kwargs(payload)
                _run_ads_sync_job("ads_sync", **kwargs)
            elif job_type == "inventory_sync":
                _run_inventory_sync()
            elif job_type == "sqp_sync":
                _run_sqp_sync()
            else:
                raise ValueError(f"Unknown job type: {job_type}")

            client.table("agent_jobs") \
                .update({"status": "done", "finished_at": datetime.now(timezone.utc).isoformat()}) \
                .eq("id", job_id) \
                .execute()
            print(f"[Job Worker] Job {job_id} completed")

        except Exception as e:
            client.table("agent_jobs") \
                .update({
                    "status": "error",
                    "finished_at": datetime.now(timezone.utc).isoformat(),
                    "error_text": str(e)[:500],
                }) \
                .eq("id", job_id) \
                .execute()
            print(f"[Job Worker] Job {job_id} failed: {e}")

    except Exception as e:
        # Table might not exist yet — silently skip
        msg = str(e)
        if "agent_jobs" in msg and ("not exist" in msg or "PGRST205" in msg):
            return  # table not created yet, skip silently
        print(f"[Job Worker] Error: {e}")


# Also trigger CPA export after successful ingest runs
if __name__ == "__main__":
    cli()
