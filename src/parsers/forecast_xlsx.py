"""Parse holiday forecast xlsx into forecast_weekly rows.

Expected format (2026 Lip Balm Forecast):
  Row 1: Title
  Row 2: Variant names at columns 1, 6, 10, 14
  Row 3: Headers (Date, 2026 Optimistic, 2025 actual, Estimate with Correction Factor)
  Rows 4+: Weekly data
  Row with empty date = totals row (skip)

Variant → SKU mapping (Tallowbourn lip balm 3pk):
  Orange / Sweet Orange → DDPE0003Shop
  Unscented → DDPE0001Shop
  Assorted → DDPE0004Shop
  Peppermint → DDPE0002Shop
"""
from __future__ import annotations

import logging
from datetime import date, datetime
from pathlib import Path

from src.db import upsert_rows

log = logging.getLogger(__name__)

# Variant name → seller SKU
VARIANT_SKU_MAP: dict[str, str] = {
    "orange": "DDPE0003Shop",
    "sweet orange": "DDPE0003Shop",
    "unscented": "DDPE0001Shop",
    "assorted": "DDPE0004Shop",
    "peppermint": "DDPE0002Shop",
}

# Column layout for 2026 forecast section (row 2 variant names at these columns)
# Each variant block: [date_col, optimistic_col, actual_col, correction_col]
# Orange=1, Unscented=6, Assorted=10, Peppermint=14 (1-indexed)
VARIANT_BLOCKS_2026 = [
    {"name_col": 1, "date_col": 1, "opt_col": 2, "act_col": 3, "corr_col": 4},
    {"name_col": 6, "date_col": 1, "opt_col": 6, "act_col": 7, "corr_col": 8},
    {"name_col": 10, "date_col": 1, "opt_col": 10, "act_col": 11, "corr_col": 12},
    {"name_col": 14, "date_col": 1, "opt_col": 14, "act_col": 15, "corr_col": 16},
]


def parse_forecast(file_path: str | Path) -> dict:
    """Parse holiday forecast xlsx and return rows for forecast_weekly.

    Returns dict with rows, variants found, and summary.
    """
    import openpyxl

    path = Path(file_path)
    if not path.exists():
        raise FileNotFoundError(f"File not found: {path}")

    wb = openpyxl.load_workbook(str(path), data_only=True)
    ws = wb[wb.sheetnames[0]]

    # Read variant names from row 2
    variants: list[dict] = []
    for block in VARIANT_BLOCKS_2026:
        name_val = ws.cell(2, block["name_col"]).value
        if not name_val:
            continue
        name = str(name_val).strip().lower()
        sku = VARIANT_SKU_MAP.get(name)
        if not sku:
            # Try partial match
            for k, v in VARIANT_SKU_MAP.items():
                if k in name:
                    sku = v
                    break
        if not sku:
            log.warning("Unknown variant: %s", name_val)
            continue
        variants.append({**block, "variant_name": str(name_val).strip(), "sku": sku})

    # Parse weekly data rows (row 4+, 2026 forecast section)
    rows: list[dict] = []
    source = path.name

    # Only parse the 2026 forecast section (rows 4 until first blank date)
    # The 2025 historical section has a different column layout
    for r in range(4, ws.max_row + 1):
        date_val = ws.cell(r, 1).value
        if not date_val:
            break  # end of 2026 section (totals row or blank)

        # Parse date
        if isinstance(date_val, datetime):
            week_start = date_val.date()
        elif isinstance(date_val, date):
            week_start = date_val
        elif isinstance(date_val, str):
            try:
                week_start = datetime.fromisoformat(date_val.split(" ")[0]).date()
            except (ValueError, IndexError):
                continue
        else:
            continue

        for v in variants:
            opt = _to_float(ws.cell(r, v["opt_col"]).value)
            act = _to_float(ws.cell(r, v["act_col"]).value)
            corr = _to_float(ws.cell(r, v["corr_col"]).value)

            if opt is not None:
                rows.append({
                    "sku": v["sku"],
                    "week_start": week_start.isoformat(),
                    "scenario": "optimistic",
                    "units": round(opt, 1),
                    "source_file": source,
                })
            if act is not None:
                rows.append({
                    "sku": v["sku"],
                    "week_start": week_start.isoformat(),
                    "scenario": "actual_2025",
                    "units": round(act, 1),
                    "source_file": source,
                })
            if corr is not None:
                rows.append({
                    "sku": v["sku"],
                    "week_start": week_start.isoformat(),
                    "scenario": "correction_factor",
                    "units": round(corr, 1),
                    "source_file": source,
                })

    return {
        "file": str(path),
        "variants": [v["variant_name"] for v in variants],
        "skus": sorted(set(v["sku"] for v in variants)),
        "weeks": len(set(r["week_start"] for r in rows)),
        "rows_total": len(rows),
        "rows": rows,
    }


def import_forecast(file_path: str | Path, dry_run: bool = False) -> dict:
    """Parse and upsert forecast data."""
    parsed = parse_forecast(file_path)

    result = {
        "file": parsed["file"],
        "variants": parsed["variants"],
        "skus": parsed["skus"],
        "weeks": parsed["weeks"],
        "rows_total": parsed["rows_total"],
        "rows_inserted": 0,
        "dry_run": dry_run,
    }

    if not dry_run and parsed["rows"]:
        result["rows_inserted"] = upsert_rows(
            "forecast_weekly", parsed["rows"],
            on_conflict="sku,week_start,scenario",
        )

    return result


def _to_float(v) -> float | None:
    if v is None:
        return None
    try:
        return float(str(v).replace(",", ""))
    except (ValueError, TypeError):
        return None
